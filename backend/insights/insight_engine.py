"""
Insight Engine - Generates AI-powered financial commentary
Analyzes variances and trends to provide actionable insights
"""

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any
import pandas as pd
import openpyxl
from openpyxl import load_workbook

import openai
import os

logger = logging.getLogger(__name__)

class InsightEngine:
    """Generate AI-powered insights from financial data"""
    
    def __init__(self, openai_api_key: str = None):
        self.api_key = openai_api_key or os.getenv('OPENAI_API_KEY')
        if self.api_key:
            openai.api_key = self.api_key
        else:
            logger.warning("No OpenAI API key found - insights will use rule-based fallback")
    
    def extract_metrics_from_template(self, template_path: Path, template_type: str) -> Dict[str, Any]:
        """Extract key metrics from a populated template"""
        
        wb = load_workbook(template_path, data_only=True)
        metrics = {}
        
        if template_type == "3_statement_model":
            metrics = self._extract_3statement_metrics(wb)
        elif template_type == "kpi_dashboard":
            metrics = self._extract_kpi_metrics(wb)
        else:
            # Generic extraction
            metrics = self._extract_generic_metrics(wb)
        
        return metrics
    
    def _extract_3statement_metrics(self, workbook) -> Dict[str, Any]:
        """Extract metrics from 3-statement model"""
        metrics = {
            'income_statement': {},
            'balance_sheet': {},
            'variances': {},
            'ratios': {}
        }
        
        # Income Statement Analysis
        if 'Income Statement' in workbook.sheetnames:
            pl = workbook['Income Statement']
            
            # Find the last two periods for comparison
            period_cols = []
            for col in range(2, pl.max_column + 1):  # Start from column B
                if pl.cell(row=3, column=col).value:  # Period header row
                    period_cols.append({
                        'col': col,
                        'period': pl.cell(row=3, column=col).value
                    })
            
            if len(period_cols) >= 2:
                current = period_cols[-1]
                prior = period_cols[-2]
                
                # Extract key P&L metrics
                metric_rows = {
                    'revenue': 5,
                    'cogs': 6,
                    'gross_profit': 7,
                    'opex': 9,
                    'ebitda': 11,
                    'net_income': 14
                }
                
                for metric, row in metric_rows.items():
                    current_val = pl.cell(row=row, column=current['col']).value or 0
                    prior_val = pl.cell(row=row, column=prior['col']).value or 0
                    
                    metrics['income_statement'][metric] = {
                        'current': float(current_val),
                        'prior': float(prior_val),
                        'change': float(current_val) - float(prior_val),
                        'change_pct': ((float(current_val) / float(prior_val)) - 1) * 100 if prior_val else 0,
                        'current_period': current['period'],
                        'prior_period': prior['period']
                    }
                
                # Calculate key ratios
                if metrics['income_statement']['revenue']['current'] > 0:
                    metrics['ratios']['gross_margin'] = (
                        metrics['income_statement']['gross_profit']['current'] / 
                        metrics['income_statement']['revenue']['current'] * 100
                    )
                    metrics['ratios']['ebitda_margin'] = (
                        metrics['income_statement']['ebitda']['current'] / 
                        metrics['income_statement']['revenue']['current'] * 100
                    )
                    metrics['ratios']['net_margin'] = (
                        metrics['income_statement']['net_income']['current'] / 
                        metrics['income_statement']['revenue']['current'] * 100
                    )
        
        # Balance Sheet Analysis
        if 'Balance Sheet' in workbook.sheetnames:
            bs = workbook['Balance Sheet']
            
            # Extract latest balance sheet metrics
            if period_cols:
                latest_col = period_cols[-1]['col']
                
                balance_metrics = {
                    'cash': 5,
                    'ar': 6,
                    'current_assets': 8,
                    'total_assets': 12,
                    'current_liabilities': 14,
                    'total_liabilities': 17,
                    'equity': 18
                }
                
                for metric, row in balance_metrics.items():
                    value = bs.cell(row=row, column=latest_col).value
                    if value:
                        metrics['balance_sheet'][metric] = float(value)
                
                # Calculate liquidity ratios
                if metrics['balance_sheet'].get('current_liabilities', 0) > 0:
                    metrics['ratios']['current_ratio'] = (
                        metrics['balance_sheet'].get('current_assets', 0) / 
                        metrics['balance_sheet']['current_liabilities']
                    )
                    metrics['ratios']['quick_ratio'] = (
                        (metrics['balance_sheet'].get('current_assets', 0) - 
                         metrics['balance_sheet'].get('inventory', 0)) / 
                        metrics['balance_sheet']['current_liabilities']
                    )
        
        # Identify significant variances
        for metric, data in metrics['income_statement'].items():
            if abs(data['change_pct']) > 10:  # 10% threshold
                metrics['variances'][metric] = {
                    'value': data['change'],
                    'percent': data['change_pct'],
                    'direction': 'increase' if data['change'] > 0 else 'decrease'
                }
        
        return metrics
    
    def _extract_kpi_metrics(self, workbook) -> Dict[str, Any]:
        """Extract metrics from KPI dashboard"""
        metrics = {'kpis': {}, 'trends': {}}
        
        # Find KPI sheet
        kpi_sheet = None
        for sheet_name in ['DASH_KPI', 'KPIs', 'Dashboard']:
            if sheet_name in workbook.sheetnames:
                kpi_sheet = workbook[sheet_name]
                break
        
        if kpi_sheet:
            # Common KPI locations (customize based on your template)
            kpi_cells = {
                'mrr': 'B5',
                'arr': 'B6',
                'customer_count': 'B8',
                'churn_rate': 'B9',
                'ltv': 'B11',
                'cac': 'B12',
                'ltv_cac_ratio': 'B13',
                'burn_rate': 'B15',
                'runway_months': 'B16'
            }
            
            for kpi, cell in kpi_cells.items():
                try:
                    value = kpi_sheet[cell].value
                    if value is not None:
                        metrics['kpis'][kpi] = float(value)
                except:
                    continue
        
        return metrics
    
    def _extract_generic_metrics(self, workbook) -> Dict[str, Any]:
        """Generic metric extraction for unknown templates"""
        return {
            'sheets': workbook.sheetnames,
            'data_available': True
        }
    
    def generate_insights(self, metrics: Dict[str, Any], template_type: str, 
                         company_context: Optional[Dict] = None) -> Dict[str, Any]:
        """Generate AI-powered insights from metrics"""
        
        # Identify key findings
        findings = self._identify_key_findings(metrics)
        
        # Generate narrative
        if self.api_key:
            narrative = self._generate_gpt4_narrative(metrics, findings, company_context)
        else:
            narrative = self._generate_rule_based_narrative(metrics, findings)
        
        # Generate recommendations
        recommendations = self._generate_recommendations(metrics, findings)
        
        return {
            'summary': self._generate_executive_summary(findings),
            'findings': findings,
            'narrative': narrative,
            'recommendations': recommendations,
            'metrics': metrics,
            'generated_at': datetime.now().isoformat()
        }
    
    def _identify_key_findings(self, metrics: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Identify the most significant findings from metrics"""
        findings = []
        
        # Check revenue trends
        if 'income_statement' in metrics and 'revenue' in metrics['income_statement']:
            revenue_data = metrics['income_statement']['revenue']
            if abs(revenue_data['change_pct']) > 5:
                findings.append({
                    'type': 'revenue_variance',
                    'metric': 'Revenue',
                    'change': revenue_data['change'],
                    'change_pct': revenue_data['change_pct'],
                    'significance': 'high' if abs(revenue_data['change_pct']) > 15 else 'medium',
                    'direction': 'positive' if revenue_data['change'] > 0 else 'negative'
                })
        
        # Check profitability
        if 'ratios' in metrics:
            if 'ebitda_margin' in metrics['ratios']:
                margin = metrics['ratios']['ebitda_margin']
                findings.append({
                    'type': 'profitability',
                    'metric': 'EBITDA Margin',
                    'value': margin,
                    'significance': 'high' if margin < 10 else 'medium',
                    'assessment': 'strong' if margin > 20 else 'weak' if margin < 10 else 'moderate'
                })
        
        # Check liquidity
        if 'ratios' in metrics and 'current_ratio' in metrics['ratios']:
            current_ratio = metrics['ratios']['current_ratio']
            findings.append({
                'type': 'liquidity',
                'metric': 'Current Ratio',
                'value': current_ratio,
                'significance': 'high' if current_ratio < 1.0 else 'low',
                'assessment': 'strong' if current_ratio > 2.0 else 'weak' if current_ratio < 1.0 else 'adequate'
            })
        
        # Check expense control
        if 'income_statement' in metrics and 'opex' in metrics['income_statement']:
            opex_data = metrics['income_statement']['opex']
            if opex_data['change_pct'] > 10:
                findings.append({
                    'type': 'expense_control',
                    'metric': 'Operating Expenses',
                    'change': opex_data['change'],
                    'change_pct': opex_data['change_pct'],
                    'significance': 'high' if opex_data['change_pct'] > 20 else 'medium',
                    'concern': True if opex_data['change_pct'] > metrics['income_statement'].get('revenue', {}).get('change_pct', 0) else False
                })
        
        # Sort by significance
        findings.sort(key=lambda x: {'high': 3, 'medium': 2, 'low': 1}.get(x.get('significance', 'low'), 0), reverse=True)
        
        return findings[:5]  # Top 5 findings
    
    def _generate_gpt4_narrative(self, metrics: Dict[str, Any], findings: List[Dict], 
                                company_context: Optional[Dict] = None) -> str:
        """Generate narrative using GPT-4"""
        
        # Build context for GPT-4
        context_parts = []
        
        # Add financial metrics
        if 'income_statement' in metrics:
            pl = metrics['income_statement']
            context_parts.append(f"Financial Performance for {pl.get('revenue', {}).get('current_period', 'latest period')}:")
            context_parts.append(f"- Revenue: ${pl.get('revenue', {}).get('current', 0):,.0f} ({pl.get('revenue', {}).get('change_pct', 0):+.1f}% vs prior period)")
            context_parts.append(f"- Gross Profit: ${pl.get('gross_profit', {}).get('current', 0):,.0f}")
            context_parts.append(f"- EBITDA: ${pl.get('ebitda', {}).get('current', 0):,.0f}")
            context_parts.append(f"- Net Income: ${pl.get('net_income', {}).get('current', 0):,.0f}")
        
        if 'ratios' in metrics:
            context_parts.append("\nKey Ratios:")
            for ratio, value in metrics['ratios'].items():
                context_parts.append(f"- {ratio.replace('_', ' ').title()}: {value:.1f}%")
        
        # Add findings
        context_parts.append("\nKey Findings:")
        for finding in findings:
            if finding['type'] == 'revenue_variance':
                context_parts.append(f"- {finding['metric']} {finding['direction']} by {abs(finding['change_pct']):.1f}%")
            elif finding['type'] == 'profitability':
                context_parts.append(f"- {finding['metric']} is {finding['value']:.1f}% ({finding['assessment']})")
        
        # Add company context if provided
        if company_context:
            context_parts.append(f"\nCompany Context:")
            context_parts.append(f"- Industry: {company_context.get('industry', 'Not specified')}")
            context_parts.append(f"- Stage: {company_context.get('stage', 'Not specified')}")
            context_parts.append(f"- Employees: {company_context.get('employee_count', 'Not specified')}")
        
        context = "\n".join(context_parts)
        
        # Generate prompt
        prompt = f"""You are a CFO providing financial commentary for a board meeting. 
Based on the following financial data and findings, write a concise narrative (3-4 sentences) that:
1. Highlights the most important trend or variance
2. Provides context for why this might be happening
3. Suggests what to watch in the coming period

{context}

Write in a professional but conversational tone. Focus on insights, not just restating numbers."""

        try:
            response = openai.ChatCompletion.create(
                model="gpt-4",
                messages=[
                    {"role": "system", "content": "You are an experienced CFO providing financial analysis."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=200,
                temperature=0.7
            )
            
            return response.choices[0].message.content.strip()
            
        except Exception as e:
            logger.error(f"GPT-4 generation failed: {e}")
            # Fallback to rule-based
            return self._generate_rule_based_narrative(metrics, findings)
    
    def _generate_rule_based_narrative(self, metrics: Dict[str, Any], findings: List[Dict]) -> str:
        """Generate narrative using rules when GPT-4 is not available"""
        
        narratives = []
        
        # Revenue narrative
        revenue_finding = next((f for f in findings if f['type'] == 'revenue_variance'), None)
        if revenue_finding:
            direction = "increased" if revenue_finding['direction'] == 'positive' else "decreased"
            narratives.append(
                f"Revenue {direction} by {abs(revenue_finding['change_pct']):.1f}% compared to the prior period, "
                f"representing a ${abs(revenue_finding['change']):,.0f} change."
            )
        
        # Profitability narrative
        profit_finding = next((f for f in findings if f['type'] == 'profitability'), None)
        if profit_finding:
            assessment = profit_finding['assessment']
            narratives.append(
                f"EBITDA margin of {profit_finding['value']:.1f}% indicates {assessment} profitability."
            )
        
        # Expense narrative
        expense_finding = next((f for f in findings if f['type'] == 'expense_control'), None)
        if expense_finding and expense_finding.get('concern'):
            narratives.append(
                f"Operating expenses grew {expense_finding['change_pct']:.1f}%, outpacing revenue growth, "
                f"which warrants attention to cost control measures."
            )
        
        # Combine narratives
        if narratives:
            narrative = " ".join(narratives[:2])  # Take top 2
            narrative += " Management should focus on "
            
            # Add focus areas
            focus_areas = []
            if revenue_finding and revenue_finding['direction'] == 'negative':
                focus_areas.append("revenue recovery strategies")
            if expense_finding and expense_finding.get('concern'):
                focus_areas.append("expense optimization")
            if not focus_areas:
                focus_areas.append("maintaining current momentum")
            
            narrative += " and ".join(focus_areas) + "."
            
            return narrative
        
        return "Financial performance remains stable with no significant variances requiring immediate attention."
    
    def _generate_recommendations(self, metrics: Dict[str, Any], findings: List[Dict]) -> List[str]:
        """Generate actionable recommendations based on findings"""
        recommendations = []
        
        # Revenue-based recommendations
        revenue_finding = next((f for f in findings if f['type'] == 'revenue_variance'), None)
        if revenue_finding:
            if revenue_finding['direction'] == 'negative':
                recommendations.append("Review sales pipeline and customer retention strategies")
                recommendations.append("Analyze pricing and competitive positioning")
            else:
                recommendations.append("Invest in growth initiatives to maintain momentum")
        
        # Profitability recommendations
        if 'ratios' in metrics and 'ebitda_margin' in metrics['ratios']:
            margin = metrics['ratios']['ebitda_margin']
            if margin < 15:
                recommendations.append("Implement cost reduction initiatives to improve margins")
            elif margin > 25:
                recommendations.append("Consider strategic investments to accelerate growth")
        
        # Liquidity recommendations
        liquidity_finding = next((f for f in findings if f['type'] == 'liquidity'), None)
        if liquidity_finding and liquidity_finding['assessment'] == 'weak':
            recommendations.append("Improve working capital management and cash collection")
            recommendations.append("Consider additional financing options to strengthen balance sheet")
        
        # Expense recommendations
        expense_finding = next((f for f in findings if f['type'] == 'expense_control'), None)
        if expense_finding and expense_finding.get('concern'):
            recommendations.append("Conduct detailed expense review to identify optimization opportunities")
        
        return recommendations[:3]  # Top 3 recommendations
    
    def _generate_executive_summary(self, findings: List[Dict]) -> str:
        """Generate one-line executive summary"""
        if not findings:
            return "Financial performance is stable with no significant variances."
        
        top_finding = findings[0]
        
        if top_finding['type'] == 'revenue_variance':
            direction = "strong growth" if top_finding['direction'] == 'positive' else "concerning decline"
            return f"Revenue shows {direction} of {abs(top_finding['change_pct']):.1f}%, requiring strategic attention."
        elif top_finding['type'] == 'profitability':
            return f"Profitability is {top_finding['assessment']} with EBITDA margin at {top_finding['value']:.1f}%."
        elif top_finding['type'] == 'expense_control':
            return f"Operating expenses increased {top_finding['change_pct']:.1f}%, impacting overall profitability."
        else:
            return "Multiple financial metrics show significant variances requiring management attention."


# Convenience function for API integration
def generate_template_insights(template_path: str, template_type: str, 
                             company_context: Optional[Dict] = None) -> Dict[str, Any]:
    """Generate insights for a populated template file"""
    
    engine = InsightEngine()
    
    # Extract metrics
    metrics = engine.extract_metrics_from_template(Path(template_path), template_type)
    
    # Generate insights
    insights = engine.generate_insights(metrics, template_type, company_context)
    
    return insights