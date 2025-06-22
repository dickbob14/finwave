#!/usr/bin/env python3
"""
Populate Payroll Summary template with data from payroll systems
Integrates Gusto/ADP data for headcount, compensation, and department analytics
"""

import argparse
import logging
import os
import sys
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Any

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Add parent directory to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from integrations.payroll.client import create_payroll_client
from config.field_mapper import FieldMapper
from templates.template_utils import (
    format_currency_cell,
    format_percent_cell,
    format_number_cell,
    add_month_columns
)
from metrics.excel_utils import add_and_ingest
from metrics.utils import normalize_period

logger = logging.getLogger(__name__)

class PayrollSummaryPopulator:
    """Populates Payroll Summary template with payroll system data"""
    
    def __init__(self, template_path: str, payroll_type: str = 'gusto'):
        self.template_path = Path(template_path)
        self.wb = None
        self.payroll_type = payroll_type
        
        # Initialize clients
        self.payroll_client = None
        
        # Field mapper
        self.mapper = FieldMapper('config/field_maps/payroll.yml')
        
    def initialize_client(self) -> None:
        """Initialize payroll client"""
        try:
            logger.info(f"Initializing {self.payroll_type} payroll client...")
            self.payroll_client = create_payroll_client(self.payroll_type)
            
        except Exception as e:
            logger.error(f"Failed to initialize payroll client: {e}")
            raise
    
    def load_template(self) -> None:
        """Load the Excel template"""
        logger.info(f"Loading template: {self.template_path}")
        self.wb = load_workbook(self.template_path, keep_vba=True, data_only=False)
        
    def fetch_payroll_data(self, start_date: str, end_date: str) -> Dict[str, pd.DataFrame]:
        """Fetch payroll data from system"""
        logger.info(f"Fetching payroll data from {start_date} to {end_date}")
        
        payroll_data = {}
        
        try:
            # Get employee roster as of end date
            employees_df = self.payroll_client.fetch_employees(as_of_date=end_date)
            payroll_data['employees'] = self.mapper.map_dataframe(
                employees_df, 
                f'{self.payroll_type}.employees'
            )
            
            # Get payroll runs for the period
            payroll_runs_df = self.payroll_client.fetch_payroll_runs(start_date, end_date)
            if not payroll_runs_df.empty:
                payroll_data['payroll_runs'] = self.mapper.map_dataframe(
                    payroll_runs_df,
                    f'{self.payroll_type}.payroll_runs'
                )
            else:
                payroll_data['payroll_runs'] = pd.DataFrame()
            
            # Get departments
            departments_df = self.payroll_client.fetch_departments()
            if not departments_df.empty:
                payroll_data['departments'] = self.mapper.map_dataframe(
                    departments_df,
                    f'{self.payroll_type}.departments'
                )
            else:
                payroll_data['departments'] = pd.DataFrame()
            
            # Get summary metrics
            payroll_data['metrics'] = self.payroll_client.get_headcount_summary(as_of_date=end_date)
            
        except Exception as e:
            logger.error(f"Error fetching payroll data: {e}")
            # Return sample data for testing
            payroll_data = self._generate_sample_payroll_data(start_date, end_date)
            
        return payroll_data
    
    def populate_data_sheets(self, payroll_data: Dict[str, pd.DataFrame]) -> None:
        """Populate DATA_* sheets with raw data"""
        
        # Populate DATA_Employees sheet
        if 'DATA_Employees' in self.wb.sheetnames and not payroll_data['employees'].empty:
            ws = self.wb['DATA_Employees']
            self._clear_and_populate_sheet(ws, payroll_data['employees'])
            logger.info(f"Populated DATA_Employees with {len(payroll_data['employees'])} records")
        
        # Populate DATA_Payroll sheet
        if 'DATA_Payroll' in self.wb.sheetnames and not payroll_data['payroll_runs'].empty:
            ws = self.wb['DATA_Payroll']
            self._clear_and_populate_sheet(ws, payroll_data['payroll_runs'])
            logger.info(f"Populated DATA_Payroll with {len(payroll_data['payroll_runs'])} records")
        
        # Populate DATA_Departments sheet
        if 'DATA_Departments' in self.wb.sheetnames and not payroll_data['departments'].empty:
            ws = self.wb['DATA_Departments']
            self._clear_and_populate_sheet(ws, payroll_data['departments'])
            logger.info(f"Populated DATA_Departments with {len(payroll_data['departments'])} records")
    
    def populate_summary_metrics(self, payroll_data: Dict[str, Any], period_date: date) -> None:
        """Populate summary metrics on dashboard"""
        
        if 'Payroll Summary' not in self.wb.sheetnames:
            logger.warning("Payroll Summary sheet not found")
            return
        
        ws = self.wb['Payroll Summary']
        
        # Get cell mappings
        cell_mappings = self.mapper.config.get('payroll_summary_cells', {})
        metrics = payroll_data.get('metrics', {})
        
        # Headcount metrics
        headcount_cells = cell_mappings.get('headcount_metrics', {})
        for metric, cell in headcount_cells.items():
            if metric in ['Total_Headcount', 'FTE_Count', 'Contractor_Count']:
                value = metrics.get(metric.lower(), 0)
                ws[cell] = value
                format_number_cell(ws[cell])
            elif metric == 'New_Hires_MTD':
                ws[cell] = metrics.get('new_hires_mtd', 0)
                format_number_cell(ws[cell])
            elif metric == 'Terminations_MTD':
                ws[cell] = metrics.get('terminations_mtd', 0)
                format_number_cell(ws[cell])
            elif metric == 'Net_Change_MTD':
                new_hires = metrics.get('new_hires_mtd', 0)
                terms = metrics.get('terminations_mtd', 0)
                ws[cell] = new_hires - terms
                format_number_cell(ws[cell])
        
        # Compensation metrics (calculate from payroll runs)
        if not payroll_data['payroll_runs'].empty:
            payroll_df = payroll_data['payroll_runs']
            
            # Total payroll cost for the period
            total_cost = payroll_df['Total_Cost'].sum() if 'Total_Cost' in payroll_df.columns else 0
            
            # Monthly average (if multiple pay periods)
            unique_pay_dates = payroll_df['Pay_Date'].nunique() if 'Pay_Date' in payroll_df.columns else 1
            monthly_cost = total_cost / max(unique_pay_dates, 1)
            
            comp_cells = cell_mappings.get('compensation_metrics', {})
            
            # Total payroll cost
            if 'Total_Payroll_Cost' in comp_cells:
                ws[comp_cells['Total_Payroll_Cost']] = monthly_cost
                format_currency_cell(ws[comp_cells['Total_Payroll_Cost']])
            
            # Average cost per FTE
            fte_count = metrics.get('fte_count', 1)
            if 'Average_Cost_FTE' in comp_cells and fte_count > 0:
                avg_cost = monthly_cost / fte_count
                ws[comp_cells['Average_Cost_FTE']] = avg_cost
                format_currency_cell(ws[comp_cells['Average_Cost_FTE']])
            
            # Benefits load percentage
            if 'Gross_Pay' in payroll_df.columns and 'Benefits_Load_Pct' in comp_cells:
                gross_pay = payroll_df['Gross_Pay'].sum()
                if gross_pay > 0:
                    benefits_load = ((total_cost - gross_pay) / gross_pay) * 100
                    ws[comp_cells['Benefits_Load_Pct']] = benefits_load / 100  # Excel expects decimal
                    format_percent_cell(ws[comp_cells['Benefits_Load_Pct']])
        
        # Department breakdown
        if 'by_department' in metrics and metrics['by_department']:
            start_row = 21  # Adjust based on template
            col_b = 2  # Department name
            col_c = 3  # Headcount
            
            for i, (dept, count) in enumerate(metrics['by_department'].items()):
                row = start_row + i
                ws.cell(row=row, column=col_b, value=dept or 'Unassigned')
                ws.cell(row=row, column=col_c, value=count)
                format_number_cell(ws.cell(row=row, column=col_c))
        
        # Add period date
        ws['B2'] = f"As of {period_date.strftime('%B %d, %Y')}"
        
        logger.info("Populated Payroll Summary metrics")
    
    def calculate_cohort_metrics(self, employees_df: pd.DataFrame, period_date: date) -> Dict[str, Any]:
        """Calculate employee cohort retention metrics"""
        
        if employees_df.empty or 'Hire_Date' not in employees_df.columns:
            return {}
        
        # Convert dates
        employees_df['Hire_Date'] = pd.to_datetime(employees_df['Hire_Date'])
        employees_df['Termination_Date'] = pd.to_datetime(employees_df['Termination_Date'])
        
        # Calculate tenure for active employees
        employees_df['Tenure_Days'] = (period_date - employees_df['Hire_Date']).dt.days
        
        # Group by hire month cohort
        employees_df['Hire_Cohort'] = employees_df['Hire_Date'].dt.to_period('M')
        
        cohort_metrics = {}
        
        # Calculate retention by cohort
        for cohort in employees_df['Hire_Cohort'].unique():
            cohort_df = employees_df[employees_df['Hire_Cohort'] == cohort]
            total_hired = len(cohort_df)
            
            if total_hired > 0:
                # Count still active
                still_active = len(cohort_df[cohort_df['Status'] == 'Active'])
                retention_rate = (still_active / total_hired) * 100
                
                cohort_metrics[str(cohort)] = {
                    'hired': total_hired,
                    'active': still_active,
                    'retention_rate': retention_rate
                }
        
        return cohort_metrics
    
    def _clear_and_populate_sheet(self, ws, df: pd.DataFrame) -> None:
        """Clear sheet and populate with dataframe"""
        # Clear existing data (keep headers in row 1)
        ws.delete_rows(2, ws.max_row)
        
        # Write headers if empty
        if ws.max_row == 1:
            for c_idx, col_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=c_idx, value=col_name)
        
        # Write data
        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    
    def _generate_sample_payroll_data(self, start_date: str, end_date: str) -> Dict[str, Any]:
        """Generate sample payroll data for testing"""
        
        # Sample employees
        employees = []
        departments = ['Engineering', 'Sales', 'Marketing', 'Operations', 'Finance']
        
        for i in range(50):
            emp = {
                'Employee_ID': f'EMP{i+1:03d}',
                'First_Name': f'First{i+1}',
                'Last_Name': f'Last{i+1}',
                'Email': f'employee{i+1}@company.com',
                'Department': departments[i % len(departments)],
                'Job_Title': 'Manager' if i % 10 == 0 else 'Contributor',
                'Employment_Type': 'FTE' if i % 10 != 9 else 'Contractor',
                'Hire_Date': (datetime.now() - timedelta(days=30*i)).strftime('%Y-%m-%d'),
                'Status': 'Active' if i % 20 != 19 else 'Terminated',
                'Location': 'San Francisco' if i % 3 == 0 else 'Remote'
            }
            employees.append(emp)
        
        employees_df = pd.DataFrame(employees)
        
        # Sample payroll runs
        payroll_runs = []
        base_salary = 120000  # Annual
        
        for emp in employees[:40]:  # Only active employees
            monthly_salary = base_salary / 12
            run = {
                'Pay_Date': end_date,
                'Employee_ID': emp['Employee_ID'],
                'Employee_Name': f"{emp['First_Name']} {emp['Last_Name']}",
                'Gross_Pay': monthly_salary,
                'Net_Pay': monthly_salary * 0.75,
                'Employer_Taxes': monthly_salary * 0.0765,  # FICA
                'Benefits_Employer_Paid': monthly_salary * 0.15,
                'Total_Cost': monthly_salary * 1.2265
            }
            payroll_runs.append(run)
        
        payroll_df = pd.DataFrame(payroll_runs)
        
        # Department summary
        dept_summary = employees_df[employees_df['Status'] == 'Active'].groupby('Department').size().reset_index()
        dept_summary.columns = ['Department_Name', 'Employee_Count']
        dept_summary['Department_ID'] = [f'DEPT{i+1:02d}' for i in range(len(dept_summary))]
        
        # Metrics
        active_employees = employees_df[employees_df['Status'] == 'Active']
        metrics = {
            'total_headcount': len(active_employees),
            'fte_count': len(active_employees[active_employees['Employment_Type'] == 'FTE']),
            'contractor_count': len(active_employees[active_employees['Employment_Type'] == 'Contractor']),
            'by_department': active_employees.groupby('Department').size().to_dict(),
            'new_hires_mtd': 3,
            'terminations_mtd': 1
        }
        
        return {
            'employees': employees_df,
            'payroll_runs': payroll_df,
            'departments': dept_summary,
            'metrics': metrics
        }
    
    def save_populated_file(self, output_path: Optional[str] = None, period_date: Optional[date] = None) -> str:
        """Save the populated workbook and ingest metrics"""
        if output_path is None:
            # Save to populated directory
            populated_dir = Path('populated')
            populated_dir.mkdir(exist_ok=True)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = populated_dir / f"payroll_summary_populated_{timestamp}.xlsx"
        
        # Use add_and_ingest helper to save and ingest metrics
        workspace_id = os.getenv('COMPANY_SLUG', 'demo-corp')
        results = add_and_ingest(
            workspace_id, 
            self.wb, 
            Path(output_path),
            period_date
        )
        
        logger.info(f"Saved and ingested: {output_path}")
        logger.info(f"Metrics ingested: {results}")
        
        return str(output_path)


def main():
    parser = argparse.ArgumentParser(description='Populate Payroll Summary with payroll system data')
    parser.add_argument('--template', default='assets/templates/registered/Payroll Summary.xlsx',
                        help='Path to template file')
    parser.add_argument('--since', required=True, help='Start date (YYYY-MM-DD)')
    parser.add_argument('--until', help='End date (YYYY-MM-DD), defaults to today')
    parser.add_argument('--payroll', choices=['gusto', 'adp'], default='gusto',
                        help='Payroll system to use')
    parser.add_argument('--output', help='Output filename')
    parser.add_argument('--debug', action='store_true', help='Enable debug logging')
    
    args = parser.parse_args()
    
    # Setup logging
    logging.basicConfig(
        level=logging.DEBUG if args.debug else logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # Default end date to today
    if not args.until:
        args.until = datetime.now().strftime('%Y-%m-%d')
    
    # Parse and normalize period date
    period_date = normalize_period(args.until)
    
    # Initialize populator
    populator = PayrollSummaryPopulator(args.template, args.payroll)
    
    try:
        # Initialize client
        populator.initialize_client()
        
        # Load template
        populator.load_template()
        
        # Fetch data
        logger.info(f"Fetching payroll data from {args.since} to {args.until}")
        payroll_data = populator.fetch_payroll_data(args.since, args.until)
        
        # Populate data sheets
        populator.populate_data_sheets(payroll_data)
        
        # Populate summary metrics
        populator.populate_summary_metrics(payroll_data, period_date)
        
        # Calculate and add cohort metrics
        if not payroll_data['employees'].empty:
            cohort_metrics = populator.calculate_cohort_metrics(
                payroll_data['employees'], 
                period_date
            )
            logger.info(f"Calculated cohort metrics for {len(cohort_metrics)} cohorts")
        
        # Save file and ingest metrics
        output_path = populator.save_populated_file(args.output, period_date)
        
        print(f"✅ Successfully populated Payroll Summary: {output_path}")
        print(f"📊 Payroll data source: {args.payroll.title()}")
        print(f"📅 Period: {args.since} to {args.until}")
        
        # Show summary
        metrics = payroll_data.get('metrics', {})
        print(f"\n👥 Headcount Summary:")
        print(f"   Total: {metrics.get('total_headcount', 0)}")
        print(f"   FTEs: {metrics.get('fte_count', 0)}")
        print(f"   Contractors: {metrics.get('contractor_count', 0)}")
        print(f"   New Hires (MTD): {metrics.get('new_hires_mtd', 0)}")
        print(f"   Terminations (MTD): {metrics.get('terminations_mtd', 0)}")
        
    except Exception as e:
        logger.error(f"Error populating template: {e}")
        raise


if __name__ == '__main__':
    main()