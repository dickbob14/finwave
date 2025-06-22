"""
Excel template generation and export system for financial reports
Supports both local Excel files and Google Sheets integration
"""
import io
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, BinaryIO
from decimal import Decimal
import json

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
import xlsxwriter
from sqlalchemy.orm import Session
from sqlalchemy import and_, func, extract

from database import get_db_session
from models.financial import GeneralLedger, Account, Customer, Vendor, FinancialPeriod

logger = logging.getLogger(__name__)

class ExcelTemplateGenerator:
    """Generate Excel templates with financial data and formulas"""
    
    def __init__(self):
        self.styles = self._create_styles()
    
    def _create_styles(self) -> Dict[str, Any]:
        """Create reusable Excel styles"""
        return {
            'header': {
                'font': Font(bold=True, color='FFFFFF'),
                'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center')
            },
            'subheader': {
                'font': Font(bold=True, color='333333'),
                'fill': PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid'),
                'alignment': Alignment(horizontal='left', vertical='center')
            },
            'currency': {
                'number_format': '$#,##0.00',
                'alignment': Alignment(horizontal='right')
            },
            'percentage': {
                'number_format': '0.00%',
                'alignment': Alignment(horizontal='right')
            },
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def generate_financial_summary(self, start_date: str, end_date: str, output_path: Optional[str] = None) -> BinaryIO:
        """
        Generate comprehensive financial summary Excel template
        
        Args:
            start_date: YYYY-MM-DD format
            end_date: YYYY-MM-DD format
            output_path: If provided, save to file. Otherwise return BytesIO
            
        Returns:
            BinaryIO object containing Excel file
        """
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        with get_db_session() as db:
            # Create individual sheets
            self._create_summary_sheet(workbook, db, start_date, end_date)
            self._create_trial_balance_sheet(workbook, db, start_date, end_date)
            self._create_pl_sheet(workbook, db, start_date, end_date)
            self._create_balance_sheet_template(workbook, db, start_date, end_date)
            self._create_cash_flow_sheet(workbook, db, start_date, end_date)
            self._create_variance_analysis_sheet(workbook, db, start_date, end_date)
        
        # Save to file or return BytesIO
        if output_path:
            workbook.save(output_path)
            logger.info(f"Financial summary saved to {output_path}")
            with open(output_path, 'rb') as f:
                return io.BytesIO(f.read())
        else:
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            return output
    
    def _create_summary_sheet(self, workbook: openpyxl.Workbook, db: Session, start_date: str, end_date: str):
        """Create executive summary sheet"""
        ws = workbook.create_sheet("Executive Summary", 0)
        
        # Header
        ws['A1'] = "Financial Executive Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Period: {start_date} to {end_date}"
        ws['A2'].font = Font(size=12, italic=True)
        
        # Key metrics section
        ws['A4'] = "Key Financial Metrics"
        self._apply_style(ws['A4'], self.styles['header'])
        
        # Calculate key metrics
        revenue = self._get_revenue(db, start_date, end_date)
        expenses = self._get_expenses(db, start_date, end_date)
        net_income = revenue - expenses
        
        # Cash balances
        cash_balance = self._get_cash_balance(db, end_date)
        ar_balance = self._get_ar_balance(db, end_date)
        ap_balance = self._get_ap_balance(db, end_date)
        
        metrics = [
            ("Total Revenue", revenue),
            ("Total Expenses", expenses),
            ("Net Income", net_income),
            ("Cash & Cash Equivalents", cash_balance),
            ("Accounts Receivable", ar_balance),
            ("Accounts Payable", ap_balance),
            ("Net Profit Margin", net_income / revenue if revenue != 0 else 0)
        ]
        
        row = 5
        for metric_name, value in metrics:
            ws[f'A{row}'] = metric_name
            ws[f'B{row}'] = value
            
            if 'Margin' in metric_name:
                self._apply_style(ws[f'B{row}'], self.styles['percentage'])
            else:
                self._apply_style(ws[f'B{row}'], self.styles['currency'])
            
            row += 1
        
        # Auto-fit columns
        self._auto_fit_columns(ws)
    
    def _create_trial_balance_sheet(self, workbook: openpyxl.Workbook, db: Session, start_date: str, end_date: str):
        """Create trial balance sheet with formulas"""
        ws = workbook.create_sheet("Trial Balance")
        
        # Headers
        headers = ["Account Code", "Account Name", "Account Type", "Debit", "Credit", "Balance"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_style(cell, self.styles['header'])
        
        # Get trial balance data
        trial_balance = self._get_trial_balance_data(db, start_date, end_date)
        
        row = 2
        total_debits = 0
        total_credits = 0
        
        for account_data in trial_balance:
            ws[f'A{row}'] = account_data['account_id']
            ws[f'B{row}'] = account_data['account_name']
            ws[f'C{row}'] = account_data['account_type']
            ws[f'D{row}'] = account_data['debit_total']
            ws[f'E{row}'] = account_data['credit_total']
            ws[f'F{row}'] = f"=D{row}-E{row}"  # Balance formula
            
            # Apply currency formatting
            for col in ['D', 'E', 'F']:
                self._apply_style(ws[f'{col}{row}'], self.styles['currency'])
            
            total_debits += account_data['debit_total']
            total_credits += account_data['credit_total']
            row += 1
        
        # Total row
        ws[f'B{row}'] = "TOTALS"
        ws[f'D{row}'] = total_debits
        ws[f'E{row}'] = total_credits
        ws[f'F{row}'] = f"=D{row}-E{row}"
        
        # Apply bold formatting to totals
        for col in ['B', 'D', 'E', 'F']:
            cell = ws[f'{col}{row}']
            cell.font = Font(bold=True)
            if col in ['D', 'E', 'F']:
                self._apply_style(cell, self.styles['currency'])
        
        self._auto_fit_columns(ws)
    
    def _create_pl_sheet(self, workbook: openpyxl.Workbook, db: Session, start_date: str, end_date: str):
        """Create Profit & Loss statement"""
        ws = workbook.create_sheet("P&L Statement")
        
        # Title
        ws['A1'] = "Profit & Loss Statement"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f"Period: {start_date} to {end_date}"
        
        # Get P&L data
        pl_data = self._get_pl_data(db, start_date, end_date)
        
        row = 4
        
        # Revenue section
        ws[f'A{row}'] = "REVENUE"
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 1
        
        total_revenue = 0
        for item in pl_data['revenue']:
            ws[f'A{row}'] = f"  {item['account_name']}"
            ws[f'B{row}'] = item['amount']
            self._apply_style(ws[f'B{row}'], self.styles['currency'])
            total_revenue += item['amount']
            row += 1
        
        # Total revenue
        ws[f'A{row}'] = "Total Revenue"
        ws[f'B{row}'] = total_revenue
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        self._apply_style(ws[f'B{row}'], self.styles['currency'])
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # Expenses section
        ws[f'A{row}'] = "EXPENSES"
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 1
        
        total_expenses = 0
        for item in pl_data['expenses']:
            ws[f'A{row}'] = f"  {item['account_name']}"
            ws[f'B{row}'] = item['amount']
            self._apply_style(ws[f'B{row}'], self.styles['currency'])
            total_expenses += item['amount']
            row += 1
        
        # Total expenses
        ws[f'A{row}'] = "Total Expenses"
        ws[f'B{row}'] = total_expenses
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        self._apply_style(ws[f'B{row}'], self.styles['currency'])
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # Net income
        net_income = total_revenue - total_expenses
        ws[f'A{row}'] = "NET INCOME"
        ws[f'B{row}'] = net_income
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'].font = Font(bold=True, size=12)
        self._apply_style(ws[f'B{row}'], self.styles['currency'])
        
        # Add border around net income
        for col in ['A', 'B']:
            ws[f'{col}{row}'].border = self.styles['border']
        
        self._auto_fit_columns(ws)
    
    def _create_balance_sheet_template(self, workbook: openpyxl.Workbook, db: Session, start_date: str, end_date: str):
        """Create Balance Sheet"""
        ws = workbook.create_sheet("Balance Sheet")
        
        # Title
        ws['A1'] = "Balance Sheet"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f"As of {end_date}"
        
        # Get balance sheet data
        bs_data = self._get_balance_sheet_data(db, end_date)
        
        row = 4
        
        # Assets section
        ws[f'A{row}'] = "ASSETS"
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 1
        
        # Current Assets
        ws[f'A{row}'] = "Current Assets"
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        row += 1
        
        current_assets_total = 0
        for item in bs_data['current_assets']:
            ws[f'A{row}'] = f"  {item['account_name']}"
            ws[f'B{row}'] = item['balance']
            self._apply_style(ws[f'B{row}'], self.styles['currency'])
            current_assets_total += item['balance']
            row += 1
        
        ws[f'A{row}'] = "Total Current Assets"
        ws[f'B{row}'] = current_assets_total
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        self._apply_style(ws[f'B{row}'], self.styles['currency'])
        row += 2
        
        # Fixed Assets
        ws[f'A{row}'] = "Fixed Assets"
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        row += 1
        
        fixed_assets_total = 0
        for item in bs_data['fixed_assets']:
            ws[f'A{row}'] = f"  {item['account_name']}"
            ws[f'B{row}'] = item['balance']
            self._apply_style(ws[f'B{row}'], self.styles['currency'])
            fixed_assets_total += item['balance']
            row += 1
        
        ws[f'A{row}'] = "Total Fixed Assets"
        ws[f'B{row}'] = fixed_assets_total
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        self._apply_style(ws[f'B{row}'], self.styles['currency'])
        row += 1
        
        # Total Assets
        total_assets = current_assets_total + fixed_assets_total
        ws[f'A{row}'] = "TOTAL ASSETS"
        ws[f'B{row}'] = total_assets
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'].font = Font(bold=True, size=12)
        self._apply_style(ws[f'B{row}'], self.styles['currency'])
        row += 3
        
        # Liabilities & Equity section
        ws[f'A{row}'] = "LIABILITIES & EQUITY"
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 1
        
        # Current Liabilities
        ws[f'A{row}'] = "Current Liabilities"
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        row += 1
        
        current_liabilities_total = 0
        for item in bs_data['current_liabilities']:
            ws[f'A{row}'] = f"  {item['account_name']}"
            ws[f'B{row}'] = item['balance']
            self._apply_style(ws[f'B{row}'], self.styles['currency'])
            current_liabilities_total += item['balance']
            row += 1
        
        ws[f'A{row}'] = "Total Current Liabilities"
        ws[f'B{row}'] = current_liabilities_total
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        self._apply_style(ws[f'B{row}'], self.styles['currency'])
        row += 2
        
        # Equity
        ws[f'A{row}'] = "Equity"
        self._apply_style(ws[f'A{row}'], self.styles['subheader'])
        row += 1
        
        equity_total = 0
        for item in bs_data['equity']:
            ws[f'A{row}'] = f"  {item['account_name']}"
            ws[f'B{row}'] = item['balance']
            self._apply_style(ws[f'B{row}'], self.styles['currency'])
            equity_total += item['balance']
            row += 1
        
        ws[f'A{row}'] = "Total Equity"
        ws[f'B{row}'] = equity_total
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        self._apply_style(ws[f'B{row}'], self.styles['currency'])
        row += 1
        
        # Total Liabilities & Equity
        total_liab_equity = current_liabilities_total + equity_total
        ws[f'A{row}'] = "TOTAL LIABILITIES & EQUITY"
        ws[f'B{row}'] = total_liab_equity
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'].font = Font(bold=True, size=12)
        self._apply_style(ws[f'B{row}'], self.styles['currency'])
        
        self._auto_fit_columns(ws)
    
    def _create_cash_flow_sheet(self, workbook: openpyxl.Workbook, db: Session, start_date: str, end_date: str):
        """Create Cash Flow statement"""
        ws = workbook.create_sheet("Cash Flow")
        
        # Title
        ws['A1'] = "Statement of Cash Flows"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f"Period: {start_date} to {end_date}"
        
        # Get cash flow data
        cf_data = self._get_cash_flow_data(db, start_date, end_date)
        
        row = 4
        
        # Operating Activities
        ws[f'A{row}'] = "CASH FLOWS FROM OPERATING ACTIVITIES"
        self._apply_style(ws[f'A{row}'], self.styles['header'])
        row += 1
        
        operating_total = 0
        for item in cf_data['operating']:
            ws[f'A{row}'] = f"  {item['description']}"
            ws[f'B{row}'] = item['amount']
            self._apply_style(ws[f'B{row}'], self.styles['currency'])
            operating_total += item['amount']
            row += 1
        
        ws[f'A{row}'] = "Net Cash from Operating Activities"
        ws[f'B{row}'] = operating_total
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        self._apply_style(ws[f'B{row}'], self.styles['currency'])
        row += 2
        
        # Add formulas for projections
        ws[f'A{row}'] = "=== PROJECTIONS ==="
        ws[f'A{row}'].font = Font(bold=True, italic=True)
        row += 1
        
        # Simple cash projection formula
        ws[f'A{row}'] = "Projected Monthly Operating Cash Flow"
        ws[f'B{row}'] = f"={get_column_letter(2)}{row-2}/DATEDIF(DATE(LEFT(\"{start_date}\",4),MID(\"{start_date}\",6,2),RIGHT(\"{start_date}\",2)),DATE(LEFT(\"{end_date}\",4),MID(\"{end_date}\",6,2),RIGHT(\"{end_date}\",2)),\"M\")"
        self._apply_style(ws[f'B{row}'], self.styles['currency'])
        
        self._auto_fit_columns(ws)
    
    def _create_variance_analysis_sheet(self, workbook: openpyxl.Workbook, db: Session, start_date: str, end_date: str):
        """Create variance analysis template"""
        ws = workbook.create_sheet("Variance Analysis")
        
        # Headers
        headers = ["Account", "Budget", "Actual", "Variance", "Variance %", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_style(cell, self.styles['header'])
        
        # Get actual data
        actual_data = self._get_pl_data(db, start_date, end_date)
        
        row = 2
        
        # Revenue variance
        for item in actual_data['revenue']:
            ws[f'A{row}'] = item['account_name']
            ws[f'B{row}'] = 0  # Budget placeholder
            ws[f'C{row}'] = item['amount']  # Actual
            ws[f'D{row}'] = f"=C{row}-B{row}"  # Variance formula
            ws[f'E{row}'] = f"=IF(B{row}=0,\"\",D{row}/B{row})"  # Variance % formula
            ws[f'F{row}'] = f"=IF(E{row}=\"\",\"\",IF(E{row}>0.1,\"Over Budget\",IF(E{row}<-0.1,\"Under Budget\",\"On Target\")))"
            
            # Apply formatting
            for col in ['B', 'C', 'D']:
                self._apply_style(ws[f'{col}{row}'], self.styles['currency'])
            self._apply_style(ws[f'E{row}'], self.styles['percentage'])
            
            row += 1
        
        # Expense variance
        for item in actual_data['expenses']:
            ws[f'A{row}'] = item['account_name']
            ws[f'B{row}'] = 0  # Budget placeholder
            ws[f'C{row}'] = item['amount']  # Actual
            ws[f'D{row}'] = f"=C{row}-B{row}"  # Variance formula
            ws[f'E{row}'] = f"=IF(B{row}=0,\"\",D{row}/B{row})"  # Variance % formula
            ws[f'F{row}'] = f"=IF(E{row}=\"\",\"\",IF(E{row}>0.1,\"Over Budget\",IF(E{row}<-0.1,\"Under Budget\",\"On Target\")))"
            
            # Apply formatting
            for col in ['B', 'C', 'D']:
                self._apply_style(ws[f'{col}{row}'], self.styles['currency'])
            self._apply_style(ws[f'E{row}'], self.styles['percentage'])
            
            row += 1
        
        self._auto_fit_columns(ws)
    
    def _apply_style(self, cell, style_dict):
        """Apply style dictionary to cell"""
        for key, value in style_dict.items():
            setattr(cell, key, value)
    
    def _auto_fit_columns(self, ws):
        """Auto-fit column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Data retrieval methods
    def _get_revenue(self, db: Session, start_date: str, end_date: str) -> Decimal:
        """Get total revenue for period"""
        result = db.query(func.sum(GeneralLedger.credit_amount)).filter(
            and_(
                GeneralLedger.transaction_date >= start_date,
                GeneralLedger.transaction_date <= end_date,
                GeneralLedger.account_type.in_(['Income', 'Revenue'])
            )
        ).scalar()
        return result or Decimal('0')
    
    def _get_expenses(self, db: Session, start_date: str, end_date: str) -> Decimal:
        """Get total expenses for period"""
        result = db.query(func.sum(GeneralLedger.debit_amount)).filter(
            and_(
                GeneralLedger.transaction_date >= start_date,
                GeneralLedger.transaction_date <= end_date,
                GeneralLedger.account_type.in_(['Expense'])
            )
        ).scalar()
        return result or Decimal('0')
    
    def _get_cash_balance(self, db: Session, end_date: str) -> Decimal:
        """Get cash balance as of date"""
        result = db.query(func.sum(GeneralLedger.amount)).filter(
            and_(
                GeneralLedger.transaction_date <= end_date,
                GeneralLedger.account_type.in_(['Bank', 'Cash'])
            )
        ).scalar()
        return result or Decimal('0')
    
    def _get_ar_balance(self, db: Session, end_date: str) -> Decimal:
        """Get accounts receivable balance"""
        result = db.query(func.sum(GeneralLedger.amount)).filter(
            and_(
                GeneralLedger.transaction_date <= end_date,
                GeneralLedger.account_type.in_(['Accounts Receivable'])
            )
        ).scalar()
        return result or Decimal('0')
    
    def _get_ap_balance(self, db: Session, end_date: str) -> Decimal:
        """Get accounts payable balance"""
        result = db.query(func.sum(GeneralLedger.amount)).filter(
            and_(
                GeneralLedger.transaction_date <= end_date,
                GeneralLedger.account_type.in_(['Accounts Payable'])
            )
        ).scalar()
        return result or Decimal('0')
    
    def _get_trial_balance_data(self, db: Session, start_date: str, end_date: str) -> List[Dict]:
        """Get trial balance data"""
        results = db.query(
            GeneralLedger.account_id,
            GeneralLedger.account_name,
            GeneralLedger.account_type,
            func.sum(GeneralLedger.debit_amount).label('debit_total'),
            func.sum(GeneralLedger.credit_amount).label('credit_total')
        ).filter(
            and_(
                GeneralLedger.transaction_date >= start_date,
                GeneralLedger.transaction_date <= end_date
            )
        ).group_by(
            GeneralLedger.account_id,
            GeneralLedger.account_name,
            GeneralLedger.account_type
        ).all()
        
        return [
            {
                'account_id': r.account_id,
                'account_name': r.account_name,
                'account_type': r.account_type,
                'debit_total': r.debit_total or Decimal('0'),
                'credit_total': r.credit_total or Decimal('0')
            }
            for r in results
        ]
    
    def _get_pl_data(self, db: Session, start_date: str, end_date: str) -> Dict:
        """Get P&L data grouped by revenue and expenses"""
        revenue_results = db.query(
            GeneralLedger.account_name,
            func.sum(GeneralLedger.credit_amount).label('amount')
        ).filter(
            and_(
                GeneralLedger.transaction_date >= start_date,
                GeneralLedger.transaction_date <= end_date,
                GeneralLedger.account_type.in_(['Income', 'Revenue'])
            )
        ).group_by(GeneralLedger.account_name).all()
        
        expense_results = db.query(
            GeneralLedger.account_name,
            func.sum(GeneralLedger.debit_amount).label('amount')
        ).filter(
            and_(
                GeneralLedger.transaction_date >= start_date,
                GeneralLedger.transaction_date <= end_date,
                GeneralLedger.account_type.in_(['Expense'])
            )
        ).group_by(GeneralLedger.account_name).all()
        
        return {
            'revenue': [{'account_name': r.account_name, 'amount': r.amount or Decimal('0')} for r in revenue_results],
            'expenses': [{'account_name': r.account_name, 'amount': r.amount or Decimal('0')} for r in expense_results]
        }
    
    def _get_balance_sheet_data(self, db: Session, end_date: str) -> Dict:
        """Get balance sheet data"""
        # Current Assets
        current_assets = db.query(
            GeneralLedger.account_name,
            func.sum(GeneralLedger.amount).label('balance')
        ).filter(
            and_(
                GeneralLedger.transaction_date <= end_date,
                GeneralLedger.account_type.in_(['Bank', 'Cash', 'Accounts Receivable', 'Other Current Asset'])
            )
        ).group_by(GeneralLedger.account_name).all()
        
        # Fixed Assets
        fixed_assets = db.query(
            GeneralLedger.account_name,
            func.sum(GeneralLedger.amount).label('balance')
        ).filter(
            and_(
                GeneralLedger.transaction_date <= end_date,
                GeneralLedger.account_type.in_(['Fixed Asset', 'Other Asset'])
            )
        ).group_by(GeneralLedger.account_name).all()
        
        # Current Liabilities
        current_liabilities = db.query(
            GeneralLedger.account_name,
            func.sum(GeneralLedger.amount).label('balance')
        ).filter(
            and_(
                GeneralLedger.transaction_date <= end_date,
                GeneralLedger.account_type.in_(['Accounts Payable', 'Credit Card', 'Other Current Liability'])
            )
        ).group_by(GeneralLedger.account_name).all()
        
        # Equity
        equity = db.query(
            GeneralLedger.account_name,
            func.sum(GeneralLedger.amount).label('balance')
        ).filter(
            and_(
                GeneralLedger.transaction_date <= end_date,
                GeneralLedger.account_type.in_(['Equity'])
            )
        ).group_by(GeneralLedger.account_name).all()
        
        return {
            'current_assets': [{'account_name': r.account_name, 'balance': r.balance or Decimal('0')} for r in current_assets],
            'fixed_assets': [{'account_name': r.account_name, 'balance': r.balance or Decimal('0')} for r in fixed_assets],
            'current_liabilities': [{'account_name': r.account_name, 'balance': abs(r.balance) or Decimal('0')} for r in current_liabilities],
            'equity': [{'account_name': r.account_name, 'balance': abs(r.balance) or Decimal('0')} for r in equity]
        }
    
    def _get_cash_flow_data(self, db: Session, start_date: str, end_date: str) -> Dict:
        """Get cash flow data"""
        # Simplified cash flow - operating activities only
        operating = [
            {'description': 'Net Income', 'amount': self._get_revenue(db, start_date, end_date) - self._get_expenses(db, start_date, end_date)},
            {'description': 'Accounts Receivable Change', 'amount': Decimal('0')},  # Placeholder
            {'description': 'Accounts Payable Change', 'amount': Decimal('0')},     # Placeholder
        ]
        
        return {
            'operating': operating
        }


# Google Sheets integration class
class GoogleSheetsExporter:
    """Export Excel templates to Google Sheets"""
    
    def __init__(self, credentials_path: str):
        """
        Initialize with Google Sheets credentials
        
        Args:
            credentials_path: Path to Google service account JSON credentials
        """
        self.credentials_path = credentials_path
        try:
            import pygsheets
            self.pygsheets = pygsheets
            self.gc = pygsheets.authorize(service_file=credentials_path)
        except ImportError:
            logger.error("pygsheets not installed. Run: pip install pygsheets")
            raise
        except Exception as e:
            logger.error(f"Failed to authorize Google Sheets: {e}")
            raise
    
    def export_to_google_sheets(self, excel_file: BinaryIO, sheet_title: str) -> str:
        """
        Export Excel file to Google Sheets
        
        Args:
            excel_file: Excel file as BytesIO
            sheet_title: Title for the Google Sheet
            
        Returns:
            URL of created Google Sheet
        """
        try:
            # Create new Google Sheet
            sh = self.gc.create(sheet_title)
            
            # Load Excel file
            excel_file.seek(0)
            workbook = openpyxl.load_workbook(excel_file)
            
            # Process each worksheet
            for i, worksheet_name in enumerate(workbook.sheetnames):
                excel_ws = workbook[worksheet_name]
                
                if i == 0:
                    # Use the first sheet that's created by default
                    gs_ws = sh.sheet1
                    gs_ws.title = worksheet_name
                else:
                    # Add new worksheet
                    gs_ws = sh.add_worksheet(worksheet_name)
                
                # Copy data from Excel to Google Sheets
                self._copy_worksheet_data(excel_ws, gs_ws)
            
            # Share with anyone with link
            sh.share('', role='reader', type='anyone')
            
            logger.info(f"Exported to Google Sheets: {sh.url}")
            return sh.url
            
        except Exception as e:
            logger.error(f"Failed to export to Google Sheets: {e}")
            raise
    
    def _copy_worksheet_data(self, excel_ws, gs_ws):
        """Copy data from Excel worksheet to Google Sheets worksheet"""
        # Get all data from Excel worksheet
        data = []
        for row in excel_ws.iter_rows(values_only=True):
            row_data = []
            for cell in row:
                if cell is None:
                    row_data.append('')
                elif isinstance(cell, datetime):
                    row_data.append(cell.strftime('%Y-%m-%d'))
                else:
                    row_data.append(str(cell))
            data.append(row_data)
        
        # Update Google Sheets with data
        if data:
            gs_ws.update_values('A1', data)


# Convenience functions
def generate_financial_excel(start_date: str, end_date: str, output_path: str = None) -> BinaryIO:
    """Generate financial Excel template"""
    generator = ExcelTemplateGenerator()
    return generator.generate_financial_summary(start_date, end_date, output_path)

def export_to_google_sheets(excel_file: BinaryIO, sheet_title: str, credentials_path: str) -> str:
    """Export Excel file to Google Sheets"""
    exporter = GoogleSheetsExporter(credentials_path)
    return exporter.export_to_google_sheets(excel_file, sheet_title)


if __name__ == "__main__":
    # Example usage
    import tempfile
    
    # Generate Excel file
    end_date = datetime.now().date().isoformat()
    start_date = (datetime.now().date() - timedelta(days=30)).isoformat()
    
    print(f"Generating financial Excel for {start_date} to {end_date}")
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        excel_file = generate_financial_excel(start_date, end_date, tmp.name)
        print(f"Excel file saved to: {tmp.name}")
        
        # Optionally export to Google Sheets
        # credentials_path = "path/to/google-credentials.json"
        # url = export_to_google_sheets(excel_file, f"Financial Report {end_date}", credentials_path)
        # print(f"Google Sheets URL: {url}")