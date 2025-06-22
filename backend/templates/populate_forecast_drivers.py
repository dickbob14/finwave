#!/usr/bin/env python3
"""
Populate Forecast Drivers from Excel template
Extracts driver assumptions and updates forecasts based on scenarios
"""

import argparse
import logging
import os
import sys
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from dateutil.relativedelta import relativedelta

# Add parent directory to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from metrics.excel_utils import add_metric_named_ranges, add_and_ingest
from metrics.utils import normalize_period
from forecast.engine import ForecastEngine
from scheduler.variance_watcher import VarianceWatcher
from core.database import get_db_session
from metrics.models import Metric

logger = logging.getLogger(__name__)

class ForecastDriverPopulator:
    """Populates driver assumptions and triggers forecast updates"""
    
    def __init__(self, template_path: str):
        self.template_path = Path(template_path)
        self.wb = None
        self.drivers = {}
        self.budget_data = {}
        
    def load_template(self) -> None:
        """Load the Excel template"""
        logger.info(f"Loading template: {self.template_path}")
        self.wb = load_workbook(self.template_path, keep_vba=True, data_only=False)
    
    def extract_driver_sheets(self) -> Dict[str, Dict[str, Any]]:
        """Extract data from DRIVER_ prefixed sheets"""
        driver_data = {}
        
        for sheet_name in self.wb.sheetnames:
            if sheet_name.startswith('DRIVER_'):
                logger.info(f"Processing driver sheet: {sheet_name}")
                ws = self.wb[sheet_name]
                
                # Common driver patterns
                driver_mappings = {
                    # Revenue drivers
                    'new_customer_growth': ['New Customer Growth', 'Customer Growth Rate', 'New Logo Growth'],
                    'churn_rate': ['Churn Rate', 'Monthly Churn', 'Customer Churn %'],
                    'arpu': ['ARPU', 'Average Revenue Per User', 'Avg Revenue/Customer'],
                    'price_increase': ['Price Increase %', 'Annual Price Increase', 'Pricing Growth'],
                    
                    # Cost drivers
                    'gross_margin_target': ['Gross Margin Target', 'Target GM %', 'Gross Margin Goal'],
                    'headcount_growth': ['Headcount Growth', 'HC Growth Rate', 'Hiring Plan Growth'],
                    'salary_inflation': ['Salary Inflation', 'Comp Increase %', 'Merit Increase'],
                    'benefits_load': ['Benefits Load', 'Benefits %', 'Employer Tax Rate'],
                    
                    # Operational drivers
                    'sales_productivity': ['Sales Productivity', 'Quota/Rep', 'Sales Efficiency'],
                    'cac_target': ['CAC Target', 'Target CAC', 'Customer Acquisition Cost'],
                    'r_and_d_percent': ['R&D %', 'R&D as % Revenue', 'Engineering %'],
                    's_and_m_percent': ['S&M %', 'Sales & Marketing %', 'GTM Spend %'],
                    'g_and_a_percent': ['G&A %', 'G&A as % Revenue', 'Admin %'],
                    
                    # Cash drivers
                    'dso': ['DSO', 'Days Sales Outstanding', 'Collection Days'],
                    'dpo': ['DPO', 'Days Payable Outstanding', 'Payment Days'],
                    'capex_percent': ['CapEx %', 'CapEx as % Revenue', 'Capital Expenditure %']
                }
                
                # Extract driver values
                sheet_drivers = {}
                
                # Search for driver names and values
                for row in range(1, min(50, ws.max_row + 1)):  # Check first 50 rows
                    cell_a = ws.cell(row=row, column=1).value
                    if cell_a:
                        cell_a_lower = str(cell_a).lower()
                        
                        for driver_id, search_terms in driver_mappings.items():
                            if any(term.lower() in cell_a_lower for term in search_terms):
                                # Look for value in adjacent cells
                                for col in range(2, min(10, ws.max_column + 1)):
                                    value = ws.cell(row=row, column=col).value
                                    if value and isinstance(value, (int, float)):
                                        sheet_drivers[driver_id] = float(value)
                                        logger.debug(f"Found {driver_id}: {value}")
                                        break
                                break
                
                # Also check for named ranges in this sheet
                if hasattr(self.wb, 'defined_names'):
                    for defined_name in self.wb.defined_names.definedName:
                        if sheet_name in str(defined_name.value):
                            name = defined_name.name.lower()
                            for driver_id in driver_mappings:
                                if driver_id.replace('_', '') in name.replace('_', ''):
                                    try:
                                        # Get the cell value
                                        cell_ref = defined_name.value.split('!')[1]
                                        cell_value = ws[cell_ref].value
                                        if cell_value and isinstance(cell_value, (int, float)):
                                            sheet_drivers[driver_id] = float(cell_value)
                                            logger.debug(f"Found {driver_id} from named range: {cell_value}")
                                    except:
                                        pass
                
                if sheet_drivers:
                    driver_data[sheet_name] = sheet_drivers
        
        return driver_data
    
    def extract_budget_sheets(self) -> Dict[str, pd.DataFrame]:
        """Extract data from BUDGET_ prefixed sheets"""
        budget_data = {}
        
        for sheet_name in self.wb.sheetnames:
            if sheet_name.startswith('BUDGET_'):
                logger.info(f"Processing budget sheet: {sheet_name}")
                ws = self.wb[sheet_name]
                
                # Read as dataframe
                data = []
                headers = []
                
                # Get headers from row 1
                for col in range(1, ws.max_column + 1):
                    header = ws.cell(row=1, column=col).value
                    if header:
                        headers.append(str(header))
                
                # Read data
                for row in range(2, ws.max_row + 1):
                    row_data = []
                    for col in range(1, len(headers) + 1):
                        value = ws.cell(row=row, column=col).value
                        row_data.append(value)
                    
                    if any(row_data):  # Skip empty rows
                        data.append(dict(zip(headers, row_data)))
                
                if data:
                    df = pd.DataFrame(data)
                    budget_data[sheet_name] = df
                    logger.info(f"Extracted {len(df)} rows from {sheet_name}")
        
        return budget_data
    
    def extract_scenario_assumptions(self) -> Dict[str, Dict[str, float]]:
        """Extract scenario assumptions (base/upside/downside)"""
        scenarios = {}
        
        # Look for scenario sheets
        scenario_sheets = [s for s in self.wb.sheetnames if 'SCENARIO' in s.upper()]
        
        for sheet_name in scenario_sheets:
            ws = self.wb[sheet_name]
            logger.info(f"Processing scenario sheet: {sheet_name}")
            
            # Extract scenario data
            scenario_data = {}
            
            # Common scenario structure: rows are drivers, columns are scenarios
            headers = []
            for col in range(2, min(10, ws.max_column + 1)):
                header = ws.cell(row=1, column=col).value
                if header:
                    headers.append(str(header).lower())
            
            # Read driver values for each scenario
            for row in range(2, min(50, ws.max_row + 1)):
                driver_name = ws.cell(row=row, column=1).value
                if driver_name:
                    driver_key = str(driver_name).lower().replace(' ', '_').replace('%', '_percent')
                    
                    for col_idx, scenario in enumerate(headers, 2):
                        value = ws.cell(row=row, column=col_idx).value
                        if value and isinstance(value, (int, float)):
                            if scenario not in scenario_data:
                                scenario_data[scenario] = {}
                            scenario_data[scenario][driver_key] = float(value)
            
            scenarios[sheet_name] = scenario_data
        
        return scenarios
    
    def create_driver_metrics(self, workspace_id: str, drivers: Dict[str, Any],
                            period_start: date, period_end: date) -> Dict[str, int]:
        """Create driver metrics in the metric store"""
        metrics_created = 0
        
        with get_db_session() as db:
            # Generate monthly periods
            current_date = period_start
            while current_date <= period_end:
                period = normalize_period(current_date)
                
                for driver_id, value in drivers.items():
                    metric_id = f"driver_{driver_id}"
                    
                    # Check if exists
                    existing = db.query(Metric).filter_by(
                        workspace_id=workspace_id,
                        metric_id=metric_id,
                        period_date=period
                    ).first()
                    
                    if existing:
                        existing.value = value
                        existing.source_template = self.template_path.name
                        existing.updated_at = datetime.utcnow()
                    else:
                        metric = Metric(
                            workspace_id=workspace_id,
                            metric_id=metric_id,
                            period_date=period,
                            value=value,
                            source_template=self.template_path.name,
                            unit='percentage' if 'percent' in driver_id or 'rate' in driver_id else None
                        )
                        db.add(metric)
                    
                    metrics_created += 1
                
                # Move to next month
                current_date = current_date + relativedelta(months=1)
            
            db.commit()
        
        logger.info(f"Created/updated {metrics_created} driver metrics")
        return {'created': metrics_created}
    
    def update_forecasts_with_drivers(self, workspace_id: str, 
                                    drivers: Dict[str, Any]) -> Dict[str, Any]:
        """Update forecasts based on new driver assumptions"""
        engine = ForecastEngine(workspace_id)
        
        # Apply driver-based adjustments to forecasts
        results = {}
        
        # Revenue forecast with drivers
        if 'new_customer_growth' in drivers or 'churn_rate' in drivers or 'arpu' in drivers:
            # Custom revenue forecast based on SaaS drivers
            revenue_forecast = self._forecast_revenue_with_drivers(
                engine, drivers, periods_ahead=12
            )
            results['revenue'] = revenue_forecast
        
        # Cost forecasts with drivers
        if 'headcount_growth' in drivers or 'salary_inflation' in drivers:
            opex_forecast = self._forecast_opex_with_drivers(
                engine, drivers, periods_ahead=12
            )
            results['opex'] = opex_forecast
        
        # Save all forecasts
        all_forecasts = {}
        for metric_type, forecast_data in results.items():
            forecast_metric_id = f"forecast_{metric_type}"
            all_forecasts[forecast_metric_id] = forecast_data
        
        saved_count = engine.save_forecasts(all_forecasts)
        
        # Also run standard forecast for other metrics
        standard_forecasts = engine.forecast_all_metrics(periods_ahead=12)
        engine.save_forecasts(standard_forecasts)
        
        return {
            'driver_based_forecasts': len(results),
            'standard_forecasts': len(standard_forecasts),
            'total_saved': saved_count
        }
    
    def _forecast_revenue_with_drivers(self, engine: ForecastEngine,
                                     drivers: Dict[str, Any],
                                     periods_ahead: int) -> Dict[date, float]:
        """Generate revenue forecast using SaaS drivers"""
        # Get latest MRR
        latest_mrr = engine._fetch_historical_data('mrr')
        if not latest_mrr:
            # Fallback to revenue
            latest_revenue = engine._fetch_historical_data('revenue')
            if latest_revenue:
                current_mrr = latest_revenue[-1].value / 12  # Approximate
            else:
                current_mrr = 100000  # Default
        else:
            current_mrr = latest_mrr[-1].value
        
        # Apply driver-based growth
        growth_rate = drivers.get('new_customer_growth', 0.02)  # 2% default
        churn_rate = drivers.get('churn_rate', 0.01)  # 1% default
        net_growth = growth_rate - churn_rate
        
        # Generate forecast
        forecast = {}
        last_date = date.today().replace(day=1)
        
        for i in range(periods_ahead):
            # Calculate next period
            future_date = last_date + relativedelta(months=i+1)
            period_date = normalize_period(future_date)
            
            # Apply growth
            current_mrr = current_mrr * (1 + net_growth)
            
            # Convert to revenue (monthly)
            forecast[period_date] = current_mrr
        
        return forecast
    
    def _forecast_opex_with_drivers(self, engine: ForecastEngine,
                                  drivers: Dict[str, Any],
                                  periods_ahead: int) -> Dict[date, float]:
        """Generate OpEx forecast using headcount drivers"""
        # Get latest OpEx
        latest_opex = engine._fetch_historical_data('opex')
        if latest_opex:
            current_opex = latest_opex[-1].value
        else:
            current_opex = 500000  # Default
        
        # Apply drivers
        hc_growth = drivers.get('headcount_growth', 0.01)  # 1% monthly default
        salary_inflation = drivers.get('salary_inflation', 0.03) / 12  # 3% annual
        
        # Generate forecast
        forecast = {}
        last_date = date.today().replace(day=1)
        
        for i in range(periods_ahead):
            # Calculate next period
            future_date = last_date + relativedelta(months=i+1)
            period_date = normalize_period(future_date)
            
            # Apply growth and inflation
            current_opex = current_opex * (1 + hc_growth + salary_inflation)
            
            forecast[period_date] = current_opex
        
        return forecast
    
    def trigger_variance_refresh(self, workspace_id: str) -> Dict[str, Any]:
        """Trigger variance check after forecast update"""
        watcher = VarianceWatcher()
        alerts = watcher.check_workspace(workspace_id)
        
        return {
            'alerts_generated': len(alerts),
            'alerts': [{'metric': a.metric_id, 'severity': a.severity} for a in alerts]
        }
    
    def save_populated_file(self, output_path: Optional[str] = None) -> str:
        """Save the workbook with driver data"""
        if output_path is None:
            populated_dir = Path('populated')
            populated_dir.mkdir(exist_ok=True)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = populated_dir / f"forecast_drivers_{timestamp}.xlsx"
        
        # Add named ranges for drivers
        driver_mappings = {}
        for sheet_drivers in self.drivers.values():
            for driver_id, value in sheet_drivers.items():
                driver_mappings[f"driver_{driver_id}"] = f"'{list(self.drivers.keys())[0]}'!B2"  # Example cell
        
        if driver_mappings:
            add_metric_named_ranges(self.wb, driver_mappings)
        
        # Save
        self.wb.save(output_path)
        logger.info(f"Saved populated workbook: {output_path}")
        
        return str(output_path)


def main():
    parser = argparse.ArgumentParser(description='Populate forecast drivers and update projections')
    parser.add_argument('--template', default='assets/templates/registered/Rolling 12-Month Forecast Template.xlsx',
                        help='Path to forecast template')
    parser.add_argument('--workspace', required=True, help='Workspace ID')
    parser.add_argument('--since', required=True, help='Start date (YYYY-MM-DD)')
    parser.add_argument('--until', help='End date (YYYY-MM-DD), defaults to +12 months')
    parser.add_argument('--scenario', help='Scenario name (base/upside/downside)')
    parser.add_argument('--output', help='Output filename')
    parser.add_argument('--debug', action='store_true', help='Enable debug logging')
    
    args = parser.parse_args()
    
    # Setup logging
    logging.basicConfig(
        level=logging.DEBUG if args.debug else logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # Default end date
    start_date = datetime.strptime(args.since, '%Y-%m-%d').date()
    if args.until:
        end_date = datetime.strptime(args.until, '%Y-%m-%d').date()
    else:
        end_date = start_date + relativedelta(months=12)
    
    # Initialize populator
    populator = ForecastDriverPopulator(args.template)
    
    try:
        # Load template
        populator.load_template()
        
        # Extract driver data
        logger.info("Extracting driver assumptions...")
        driver_sheets = populator.extract_driver_sheets()
        
        # Extract budget data
        logger.info("Extracting budget data...")
        budget_sheets = populator.extract_budget_sheets()
        
        # Extract scenarios
        logger.info("Extracting scenario assumptions...")
        scenarios = populator.extract_scenario_assumptions()
        
        # Combine all drivers
        all_drivers = {}
        for sheet_drivers in driver_sheets.values():
            all_drivers.update(sheet_drivers)
        
        # Apply scenario if specified
        if args.scenario and scenarios:
            for scenario_data in scenarios.values():
                if args.scenario.lower() in scenario_data:
                    logger.info(f"Applying {args.scenario} scenario")
                    all_drivers.update(scenario_data[args.scenario.lower()])
                    break
        
        if not all_drivers:
            logger.warning("No driver assumptions found")
        else:
            print(f"\n📊 Driver Assumptions:")
            for driver_id, value in sorted(all_drivers.items()):
                if 'percent' in driver_id or 'rate' in driver_id:
                    print(f"   {driver_id}: {value:.1f}%")
                else:
                    print(f"   {driver_id}: {value:,.0f}")
        
        # Store drivers in metric store
        logger.info("Storing driver metrics...")
        driver_results = populator.create_driver_metrics(
            args.workspace, all_drivers, start_date, end_date
        )
        
        # Update forecasts based on drivers
        logger.info("Updating forecasts with driver assumptions...")
        forecast_results = populator.update_forecasts_with_drivers(
            args.workspace, all_drivers
        )
        
        # Trigger variance refresh
        logger.info("Checking for new variances...")
        variance_results = populator.trigger_variance_refresh(args.workspace)
        
        # Save populated file
        output_path = populator.save_populated_file(args.output)
        
        # Print summary
        print(f"\n✅ Forecast Driver Update Complete")
        print(f"📊 Drivers extracted: {len(all_drivers)}")
        print(f"💾 Driver metrics created: {driver_results['created']}")
        print(f"📈 Forecasts updated: {forecast_results['total_saved']}")
        print(f"🚨 Variance alerts: {variance_results['alerts_generated']}")
        print(f"📁 Saved to: {output_path}")
        
        if variance_results['alerts_generated'] > 0:
            print(f"\n⚠️  New Alerts:")
            for alert in variance_results['alerts'][:5]:
                print(f"   - {alert['metric']} ({alert['severity']})")
        
    except Exception as e:
        logger.error(f"Error updating forecast drivers: {e}")
        raise


if __name__ == '__main__':
    main()