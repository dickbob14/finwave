#!/usr/bin/env python3
"""
ETL Script: QuickBooks to Excel Template (Enhanced Version)
Populates the FinWave Excel template DATA_GL sheet with real QuickBooks data
"""

import os
import sys
import json
import pandas as pd
import requests
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Tuple
import logging
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path

# Add parent directory to path
sys.path.append(str(Path(__file__).parent.parent))

# Import utilities
from templates.template_utils import (
    get_template_path, calculate_signed_amount, format_excel_date,
    get_prior_year_period, recalculate_workbook, validate_gl_totals,
    build_coa_mapping
)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class QuickBooksToExcelETL:
    def __init__(self, qb_server_url: str = "http://localhost:8000", template_path: str = None):
        self.qb_server_url = qb_server_url
        self.template_path = Path(template_path) if template_path else get_template_path() / "finwave_board_pack.xlsx"
        
    def fetch_qb_data(self, include_prior_year: bool = True) -> Dict[str, Any]:
        """Fetch data from QuickBooks server including prior year if requested"""
        try:
            logger.info("Fetching QuickBooks data from server...")
            
            # Get current period transactions
            response = requests.get(f"{self.qb_server_url}/real/transactions")
            response.raise_for_status()
            transactions = response.json()
            
            # Get company info
            company_response = requests.get(f"{self.qb_server_url}/real/company")
            company_response.raise_for_status()
            company = company_response.json()
            
            result = {
                "transactions": transactions,
                "company": company
            }
            
            # Fetch prior year data if requested
            if include_prior_year:
                period = transactions.get("period", {})
                if period.get("start") and period.get("end"):
                    py_start, py_end = get_prior_year_period(period["start"], period["end"])
                    
                    py_response = requests.get(
                        f"{self.qb_server_url}/real/transactions",
                        params={"start_date": py_start, "end_date": py_end}
                    )
                    py_response.raise_for_status()
                    result["transactions_prior_year"] = py_response.json()
                    logger.info(f"Fetched prior year data: {py_start} to {py_end}")
            
            logger.info(f"Fetched {len(transactions.get('invoices', []))} invoices and {len(transactions.get('expenses', []))} expenses")
            return result
            
        except requests.exceptions.RequestException as e:
            logger.error(f"Failed to fetch QuickBooks data: {e}")
            raise
    
    def transform_to_gl_format(self, qb_data: Dict[str, Any], is_prior_year: bool = False) -> pd.DataFrame:
        """Transform QuickBooks data into General Ledger format for Excel"""
        gl_entries = []
        
        transactions_key = "transactions_prior_year" if is_prior_year else "transactions"
        transactions = qb_data.get(transactions_key, {})
        
        # Process invoices (revenue)
        for invoice in transactions.get("invoices", []):
            try:
                txn_date = format_excel_date(invoice["TxnDate"])
                amount = float(invoice.get("TotalAmt", 0))
                
                # Revenue entry - use signed amount
                revenue_amount = calculate_signed_amount(0, amount, "Income")
                gl_entries.append({
                    "Date": txn_date,
                    "Account": "4000",
                    "Account_Name": "Revenue",
                    "Amount": revenue_amount,
                    "Description": f"Invoice #{invoice.get('DocNumber', 'N/A')} - {invoice.get('CustomerRef', {}).get('name', 'Unknown')}",
                    "Reference": invoice.get("DocNumber", ""),
                    "Customer": invoice.get("CustomerRef", {}).get("name", ""),
                    "Vendor": "",
                    "Class": "",
                    "Location": "",
                    "Memo": invoice.get("PrivateNote", ""),
                    "TxnID": invoice.get("Id", ""),
                    "Group": "Revenue",
                    "SubGroup": "Product Revenue"
                })
                
                # AR entry - use signed amount
                ar_amount = calculate_signed_amount(amount, 0, "Asset")
                gl_entries.append({
                    "Date": txn_date,
                    "Account": "1200",
                    "Account_Name": "Accounts Receivable",
                    "Amount": ar_amount,
                    "Description": f"A/R Invoice #{invoice.get('DocNumber', 'N/A')}",
                    "Reference": invoice.get("DocNumber", ""),
                    "Customer": invoice.get("CustomerRef", {}).get("name", ""),
                    "Vendor": "",
                    "Class": "",
                    "Location": "",
                    "Memo": invoice.get("PrivateNote", ""),
                    "TxnID": invoice.get("Id", ""),
                    "Group": "Current Assets",
                    "SubGroup": "Receivables"
                })
                
            except Exception as e:
                logger.warning(f"Error processing invoice {invoice.get('Id', 'unknown')}: {e}")
                continue
        
        # Process expenses
        for expense in transactions.get("expenses", []):
            try:
                txn_date = format_excel_date(expense["TxnDate"])
                amount = float(expense.get("TotalAmt", 0))
                account_name = expense.get("AccountRef", {}).get("name", "General Expenses")
                
                # Map expense categories
                account_code = "6000"
                group = "Operating Expenses"
                subgroup = "General & Administrative"
                
                if "travel" in account_name.lower():
                    account_code = "6100"
                    subgroup = "Travel & Entertainment"
                elif "office" in account_name.lower():
                    account_code = "6200"
                    subgroup = "Office Expenses"
                elif "marketing" in account_name.lower():
                    account_code = "6300"
                    subgroup = "Sales & Marketing"
                elif "payroll" in account_name.lower() or "salary" in account_name.lower():
                    account_code = "6400"
                    subgroup = "Compensation & Benefits"
                
                # Expense entry - use signed amount
                expense_amount = calculate_signed_amount(amount, 0, "Expense")
                gl_entries.append({
                    "Date": txn_date,
                    "Account": account_code,
                    "Account_Name": account_name,
                    "Amount": expense_amount,
                    "Description": f"Expense - {expense.get('PrivateNote', 'General expense')}",
                    "Reference": expense.get("DocNumber", ""),
                    "Customer": "",
                    "Vendor": expense.get("EntityRef", {}).get("name", ""),
                    "Class": "",
                    "Location": "",
                    "Memo": expense.get("PrivateNote", ""),
                    "TxnID": expense.get("Id", ""),
                    "Group": group,
                    "SubGroup": subgroup
                })
                
                # Cash/AP entry - use signed amount
                payment_type = expense.get("PaymentType", "")
                if payment_type == "Cash":
                    cash_amount = calculate_signed_amount(0, amount, "Asset")
                    gl_entries.append({
                        "Date": txn_date,
                        "Account": "1000",
                        "Account_Name": "Cash",
                        "Amount": cash_amount,
                        "Description": f"Cash payment for expense",
                        "Reference": expense.get("DocNumber", ""),
                        "Customer": "",
                        "Vendor": expense.get("EntityRef", {}).get("name", ""),
                        "Class": "",
                        "Location": "",
                        "Memo": expense.get("PrivateNote", ""),
                        "TxnID": expense.get("Id", ""),
                        "Group": "Current Assets",
                        "SubGroup": "Cash & Equivalents"
                    })
                else:
                    ap_amount = calculate_signed_amount(0, amount, "Liability")
                    gl_entries.append({
                        "Date": txn_date,
                        "Account": "2000",
                        "Account_Name": "Accounts Payable",
                        "Amount": ap_amount,
                        "Description": f"AP for expense",
                        "Reference": expense.get("DocNumber", ""),
                        "Customer": "",
                        "Vendor": expense.get("EntityRef", {}).get("name", ""),
                        "Class": "",
                        "Location": "",
                        "Memo": expense.get("PrivateNote", ""),
                        "TxnID": expense.get("Id", ""),
                        "Group": "Current Liabilities",
                        "SubGroup": "Payables"
                    })
                
            except Exception as e:
                logger.warning(f"Error processing expense {expense.get('Id', 'unknown')}: {e}")
                continue
        
        # Convert to DataFrame and sort by date
        df = pd.DataFrame(gl_entries)
        if not df.empty:
            df['Date'] = pd.to_datetime(df['Date'])
            df = df.sort_values('Date')
            # Keep dates in ISO format for Excel
            df['Date'] = df['Date'].dt.strftime('%Y-%m-%d')
        
        logger.info(f"Transformed {len(df)} general ledger entries{' (prior year)' if is_prior_year else ''}")
        return df
    
    def load_template(self) -> Any:
        """Load the Excel template"""
        try:
            if not self.template_path.exists():
                raise FileNotFoundError(f"Template not found: {self.template_path}")
            
            logger.info(f"Loading Excel template: {self.template_path}")
            workbook = load_workbook(self.template_path)
            return workbook
            
        except Exception as e:
            logger.error(f"Failed to load template: {e}")
            raise
    
    def populate_data_gl_sheet(self, workbook: Any, gl_df: pd.DataFrame, sheet_name: str = "DATA_GL") -> None:
        """Populate a GL data sheet with transformed data"""
        try:
            if sheet_name not in workbook.sheetnames:
                logger.error(f"{sheet_name} sheet not found in template")
                raise ValueError(f"{sheet_name} sheet not found")
            
            ws = workbook[sheet_name]
            logger.info(f"Populating {sheet_name} sheet...")
            
            # Clear existing data (keep headers)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.value = None
            
            # Expected columns for tblGL
            columns = [
                "Date", "Account", "Account_Name", "Amount", "Description",
                "Reference", "Customer", "Vendor", "Class", "Location",
                "Memo", "TxnID", "Group", "SubGroup"
            ]
            
            # Add data starting from row 2
            for r_idx, (index, row) in enumerate(gl_df.iterrows(), start=2):
                for c_idx, col in enumerate(columns, start=1):
                    value = row.get(col, "")
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Format date column
                    if col == "Date" and value:
                        cell.number_format = 'yyyy-mm-dd'
                    # Format amount column
                    elif col == "Amount" and value:
                        cell.number_format = '#,##0.00;[Red](#,##0.00)'
            
            logger.info(f"Populated {len(gl_df)} rows in {sheet_name} sheet")
            
        except Exception as e:
            logger.error(f"Failed to populate {sheet_name} sheet: {e}")
            raise
    
    def update_settings_sheet(self, workbook: Any, qb_data: Dict[str, Any]) -> None:
        """Update the SETTINGS sheet with company info and refresh date"""
        try:
            if "SETTINGS" not in workbook.sheetnames:
                logger.warning("SETTINGS sheet not found, skipping update")
                return
            
            ws = workbook["SETTINGS"]
            company_info = qb_data.get("company", {}).get("QueryResponse", {}).get("CompanyInfo", [{}])[0]
            
            # Update company name (B2)
            if company_info.get("CompanyName"):
                ws["B2"] = company_info["CompanyName"]
            
            # Update period dates
            transactions = qb_data.get("transactions", {})
            period = transactions.get("period", {})
            
            if period.get("start"):
                ws["B3"] = period["start"]  # Fiscal Year Start
            if period.get("end"):
                ws["B4"] = period["end"]    # Fiscal Year End
            
            # Update refresh timestamp (B6)
            ws["B6"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            logger.info("Updated SETTINGS sheet with company info and refresh time")
            
        except Exception as e:
            logger.warning(f"Failed to update SETTINGS sheet: {e}")
    
    def save_workbook(self, workbook: Any, output_path: str = None) -> str:
        """Save the populated workbook"""
        try:
            if not output_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = f"finwave_board_pack_populated_{timestamp}.xlsx"
            
            workbook.save(output_path)
            logger.info(f"Saved populated workbook: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to save workbook: {e}")
            raise
    
    def run_etl(self, output_path: str = None) -> str:
        """Run the complete ETL process"""
        try:
            logger.info("Starting QuickBooks to Excel ETL process...")
            
            # Step 1: Fetch QuickBooks data (including prior year)
            qb_data = self.fetch_qb_data(include_prior_year=True)
            
            # Step 2: Transform current year to GL format
            gl_df = self.transform_to_gl_format(qb_data, is_prior_year=False)
            
            # Step 3: Transform prior year to GL format (if available)
            gl_df_py = None
            if "transactions_prior_year" in qb_data:
                gl_df_py = self.transform_to_gl_format(qb_data, is_prior_year=True)
            
            if gl_df.empty:
                logger.warning("No data to populate - GL DataFrame is empty")
                return None
            
            # Step 4: Load template
            workbook = self.load_template()
            
            # Step 5: Populate DATA_GL sheet
            self.populate_data_gl_sheet(workbook, gl_df, "DATA_GL")
            
            # Step 6: Populate DATA_GL_PY sheet if prior year data exists
            if gl_df_py is not None and not gl_df_py.empty:
                if "DATA_GL_PY" in workbook.sheetnames:
                    self.populate_data_gl_sheet(workbook, gl_df_py, "DATA_GL_PY")
                    logger.info(f"Populated prior year data with {len(gl_df_py)} entries")
                else:
                    logger.warning("DATA_GL_PY sheet not found - skipping prior year data")
            
            # Step 7: Update settings
            self.update_settings_sheet(workbook, qb_data)
            
            # Step 8: Save populated workbook
            output_file = self.save_workbook(workbook, output_path)
            
            # Step 9: Recalculate and validate (if xlwings available)
            calc_values = recalculate_workbook(output_file)
            if calc_values:
                is_valid = validate_gl_totals(gl_df, calc_values)
                if not is_valid:
                    logger.warning("Validation failed - GL totals don't match calculated values")
            
            logger.info("ETL process completed successfully!")
            logger.info(f"✅ Populated Excel template: {output_file}")
            logger.info(f"📊 Total GL entries: {len(gl_df)}")
            logger.info(f"💰 Revenue entries: {len(gl_df[gl_df['Group'] == 'Revenue'])}")
            logger.info(f"💸 Expense entries: {len(gl_df[gl_df['Group'].str.contains('Expense', na=False)])}")
            
            return output_file
            
        except Exception as e:
            logger.error(f"ETL process failed: {e}")
            raise

def main():
    """Main entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(description="Populate FinWave Excel template with QuickBooks data")
    parser.add_argument("--template", help="Path to Excel template")
    parser.add_argument("--output", help="Output file path")
    parser.add_argument("--qb-server", default="http://localhost:8000", help="QuickBooks server URL")
    
    args = parser.parse_args()
    
    try:
        etl = QuickBooksToExcelETL(
            qb_server_url=args.qb_server,
            template_path=args.template
        )
        
        output_file = etl.run_etl(args.output)
        
        if output_file:
            print(f"\n🎉 SUCCESS! Populated Excel template saved as: {output_file}")
            print("\n📋 Next steps:")
            print("1. Open the file in Excel to verify formulas are calculating correctly")
            print("2. Review the financial reports on REPORT_P&L, REPORT_BS, and REPORT_CF sheets")
            print("3. Check the executive dashboard on DASH_KPI and DASH_TRENDS sheets")
            print("4. Use the template to generate PDF board pack reports")
        else:
            print("❌ No data was available to populate the template")
            
    except Exception as e:
        print(f"❌ ETL failed: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()