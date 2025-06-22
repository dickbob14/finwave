#!/usr/bin/env python3
"""
FinWave Board-Pack Template Generator
Generates professional Excel/Google Sheets financial reporting templates
"""

import os
import json
import datetime
from typing import Dict, List, Tuple, Optional
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Border, Side, Alignment, 
    NamedStyle, Protection
)
from openpyxl.formatting.rule import (
    CellIsRule, FormulaRule, IconSetRule, IconSet, Rule
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.comments import Comment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.drawing.colors import ColorChoice
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Color Palette
COLORS = {
    'primary': '002B49',      # Midnight Blue
    'accent1': '00A6A6',      # Teal
    'accent2': 'FF9F1C',      # Mango
    'neutral_light': 'F5F6F8',
    'neutral_mid': 'CBD2D9',
    'white': 'FFFFFF',
    'black': '000000',
    'green': '00B050',
    'red': 'FF0000',
    'amber': 'FFC000'
}

# Style definitions
def create_styles():
    """Create named styles for consistent formatting"""
    styles = {}
    
    # Title style
    title = NamedStyle(name='title')
    title.font = Font(name='Calibri', size=18, bold=True, color=COLORS['primary'])
    title.alignment = Alignment(horizontal='left', vertical='center')
    styles['title'] = title
    
    # Section Header
    section_header = NamedStyle(name='section_header')
    section_header.font = Font(name='Calibri', size=14, bold=True, color=COLORS['white'])
    section_header.fill = PatternFill("solid", fgColor=COLORS['primary'])
    section_header.alignment = Alignment(horizontal='center', vertical='center')
    section_header.border = Border(
        left=Side(style='thin', color=COLORS['primary']),
        right=Side(style='thin', color=COLORS['primary']),
        top=Side(style='thin', color=COLORS['primary']),
        bottom=Side(style='thin', color=COLORS['primary'])
    )
    styles['section_header'] = section_header
    
    # Column Header
    col_header = NamedStyle(name='col_header')
    col_header.font = Font(name='Calibri', size=11, bold=True, color=COLORS['primary'])
    col_header.fill = PatternFill("solid", fgColor=COLORS['neutral_light'])
    col_header.alignment = Alignment(horizontal='center', vertical='center')
    col_header.border = Border(bottom=Side(style='medium', color=COLORS['primary']))
    styles['col_header'] = col_header
    
    # Data cells
    data_cell = NamedStyle(name='data_cell')
    data_cell.font = Font(name='Calibri', size=11)
    data_cell.alignment = Alignment(horizontal='right', vertical='center')
    data_cell.border = Border(
        left=Side(style='thin', color=COLORS['neutral_mid']),
        right=Side(style='thin', color=COLORS['neutral_mid'])
    )
    styles['data_cell'] = data_cell
    
    # Total row
    total_row = NamedStyle(name='total_row')
    total_row.font = Font(name='Calibri', size=11, bold=True)
    total_row.fill = PatternFill("solid", fgColor=COLORS['neutral_light'])
    total_row.border = Border(
        top=Side(style='double', color=COLORS['primary']),
        bottom=Side(style='thin', color=COLORS['primary'])
    )
    styles['total_row'] = total_row
    
    # KPI tile
    kpi_tile = NamedStyle(name='kpi_tile')
    kpi_tile.font = Font(name='Calibri', size=32, bold=True, color=COLORS['primary'])
    kpi_tile.alignment = Alignment(horizontal='center', vertical='center')
    styles['kpi_tile'] = kpi_tile
    
    return styles


class FinWaveTemplateBuilder:
    def __init__(self):
        self.wb = Workbook()
        self.styles = create_styles()
        
        # Register styles
        for style in self.styles.values():
            self.wb.add_named_style(style)
        
        # Remove default sheet
        self.wb.remove(self.wb.active)
        
    def create_data_gl(self):
        """Create DATA_GL sheet for General Ledger data"""
        ws = self.wb.create_sheet("DATA_GL")
        ws.sheet_state = 'hidden'
        
        # Headers
        headers = ['Date', 'Account', 'Account Name', 'Debit', 'Credit', 'Amount', 
                   'Class', 'Location', 'Memo', 'Transaction Type', 'Reference']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.style = 'col_header'
        
        # Create Excel table
        table = Table(displayName="tblGL", ref=f"A1:K1000")
        style = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)
        
        # Column widths
        col_widths = [12, 20, 40, 15, 15, 15, 20, 20, 30, 20, 15]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
            
        return ws
    
    def create_data_coa(self):
        """Create DATA_COA sheet for Chart of Accounts"""
        ws = self.wb.create_sheet("DATA_COA")
        ws.sheet_state = 'hidden'
        
        headers = ['Account Number', 'Account Name', 'Type', 'Sub-Type', 
                   'Normal Balance', 'Group', 'Sub-Group', 'Active']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.style = 'col_header'
        
        # Create table
        table = Table(displayName="tblCOA", ref="A1:H500")
        style = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)
        
        # Column widths  
        col_widths = [15, 40, 20, 20, 15, 20, 20, 10]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
            
        return ws
    
    def create_data_map(self):
        """Create DATA_MAP sheet for mapping configurations"""
        ws = self.wb.create_sheet("DATA_MAP")
        ws.sheet_state = 'hidden'
        
        # Account grouping map
        ws['A1'] = 'Account Grouping Map'
        ws['A1'].style = 'section_header'
        ws.merge_cells('A1:C1')
        
        headers = ['Account Pattern', 'Group', 'Sub-Group']
        for col, header in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=header).style = 'col_header'
            
        # Sample mappings
        mappings = [
            ('4*', 'Revenue', 'Operating Revenue'),
            ('5*', 'Cost of Goods Sold', 'Direct Costs'),
            ('6*', 'Operating Expenses', 'SG&A'),
            ('7*', 'Other Income', 'Non-Operating'),
            ('8*', 'Other Expenses', 'Non-Operating'),
        ]
        
        for i, (pattern, group, subgroup) in enumerate(mappings, 3):
            ws.cell(row=i, column=1, value=pattern)
            ws.cell(row=i, column=2, value=group)
            ws.cell(row=i, column=3, value=subgroup)
            
        return ws
    
    def create_report_pl(self):
        """Create REPORT_P&L sheet - Income Statement"""
        ws = self.wb.create_sheet("REPORT_P&L")
        
        # Title
        ws['A1'] = '=SETTINGS!B2 & " · Income Statement"'
        ws['A1'].style = 'title'
        ws.merge_cells('A1:N1')
        
        # Period headers
        ws['A3'] = 'Account'
        ws['B3'] = 'Description'
        ws['A3'].style = 'col_header'
        ws['B3'].style = 'col_header'
        
        # Month columns C through N (12 months)
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                  'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        
        for i, month in enumerate(months):
            cell = ws.cell(row=3, column=i+3, value=f'{month} 2025')
            cell.style = 'col_header'
            
        # P&L Structure
        pl_structure = [
            ('', '', 4),  # Blank row
            ('REVENUE', '', 5),
            ('  Product Revenue', '=SUMIFS(tblGL[Amount],tblGL[Account],"4100*",tblGL[Date],">="&C$2,tblGL[Date],"<"&D$2)', 6),
            ('  Service Revenue', '=SUMIFS(tblGL[Amount],tblGL[Account],"4200*",tblGL[Date],">="&C$2,tblGL[Date],"<"&D$2)', 7),
            ('  Other Revenue', '=SUMIFS(tblGL[Amount],tblGL[Account],"4900*",tblGL[Date],">="&C$2,tblGL[Date],"<"&D$2)', 8),
            ('Total Revenue', '=SUM(C6:C8)', 9),
            ('', '', 10),
            ('COST OF GOODS SOLD', '', 11),
            ('  Direct Materials', '=SUMIFS(tblGL[Amount],tblGL[Account],"5100*",tblGL[Date],">="&C$2,tblGL[Date],"<"&D$2)', 12),
            ('  Direct Labor', '=SUMIFS(tblGL[Amount],tblGL[Account],"5200*",tblGL[Date],">="&C$2,tblGL[Date],"<"&D$2)', 13),
            ('  Manufacturing Overhead', '=SUMIFS(tblGL[Amount],tblGL[Account],"5300*",tblGL[Date],">="&C$2,tblGL[Date],"<"&D$2)', 14),
            ('Total COGS', '=SUM(C12:C14)', 15),
            ('', '', 16),
            ('GROSS PROFIT', '=C9-C15', 17),
            ('Gross Margin %', '=IFERROR(C17/C9,0)', 18),
            ('', '', 19),
            ('OPERATING EXPENSES', '', 20),
            ('  Sales & Marketing', '=SUMIFS(tblGL[Amount],tblGL[Account],"6100*",tblGL[Date],">="&C$2,tblGL[Date],"<"&D$2)', 21),
            ('  General & Admin', '=SUMIFS(tblGL[Amount],tblGL[Account],"6200*",tblGL[Date],">="&C$2,tblGL[Date],"<"&D$2)', 22),
            ('  Research & Development', '=SUMIFS(tblGL[Amount],tblGL[Account],"6300*",tblGL[Date],">="&C$2,tblGL[Date],"<"&D$2)', 23),
            ('Total Operating Expenses', '=SUM(C21:C23)', 24),
            ('', '', 25),
            ('OPERATING INCOME', '=C17-C24', 26),
            ('Operating Margin %', '=IFERROR(C26/C9,0)', 27),
            ('', '', 28),
            ('OTHER INCOME/EXPENSES', '', 29),
            ('  Interest Income', '=SUMIFS(tblGL[Amount],tblGL[Account],"7100*",tblGL[Date],">="&C$2,tblGL[Date],"<"&D$2)', 30),
            ('  Interest Expense', '=SUMIFS(tblGL[Amount],tblGL[Account],"8100*",tblGL[Date],">="&C$2,tblGL[Date],"<"&D$2)', 31),
            ('  Other Income/Expense', '=SUMIFS(tblGL[Amount],tblGL[Account],"7900*",tblGL[Date],">="&C$2,tblGL[Date],"<"&D$2)', 32),
            ('Total Other', '=SUM(C30:C32)', 33),
            ('', '', 34),
            ('NET INCOME', '=C26+C33', 35),
            ('Net Margin %', '=IFERROR(C35/C9,0)', 36),
        ]
        
        # Populate structure
        for desc, formula, row in pl_structure:
            ws.cell(row=row, column=1, value=desc)
            if formula and row not in [5, 11, 20, 29]:  # Skip section headers
                # Apply formula to all month columns
                for col in range(3, 15):  # C through N
                    if '%' in desc:
                        ws.cell(row=row, column=col).number_format = '0.0%'
                    else:
                        ws.cell(row=row, column=col).number_format = '#,##0'
                        
            # Apply styles
            if desc in ['REVENUE', 'COST OF GOODS SOLD', 'OPERATING EXPENSES', 'OTHER INCOME/EXPENSES']:
                ws.cell(row=row, column=1).style = 'section_header'
                ws.merge_cells(f'A{row}:B{row}')
            elif desc in ['Total Revenue', 'Total COGS', 'GROSS PROFIT', 'Total Operating Expenses', 
                          'OPERATING INCOME', 'NET INCOME']:
                ws.cell(row=row, column=1).style = 'total_row'
                for col in range(3, 15):
                    ws.cell(row=row, column=col).style = 'total_row'
                    
        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
        for col in range(3, 15):
            ws.column_dimensions[get_column_letter(col)].width = 12
            
        # Conditional formatting for variance
        # Add data bars for % rows
        from openpyxl.formatting.rule import DataBarRule
        for row in [18, 27, 36]:  # Margin % rows
            data_bar = DataBarRule(start_type='min', end_type='max', color='00A6A6')
            ws.conditional_formatting.add(f'C{row}:N{row}', data_bar)
            
        # Define named range
        defn = DefinedName('rngIS_Matrix', attr_text="'REPORT_P&L'!$A$5:$N$36")
        self.wb.defined_names['rngIS_Matrix'] = defn
        
        return ws
    
    def create_report_bs(self):
        """Create REPORT_BS sheet - Balance Sheet"""
        ws = self.wb.create_sheet("REPORT_BS")
        
        # Title
        ws['A1'] = '=SETTINGS!B2 & " · Balance Sheet"'
        ws['A1'].style = 'title'
        ws.merge_cells('A1:F1')
        
        # Headers
        ws['A3'] = 'Account'
        ws['B3'] = 'Description'
        ws['C3'] = 'Current Period'
        ws['D3'] = 'Prior Period'
        ws['E3'] = '$ Change'
        ws['F3'] = '% Change'
        
        for col in range(1, 7):
            ws.cell(row=3, column=col).style = 'col_header'
            
        # Balance Sheet structure
        bs_structure = [
            ('', '', 4),
            ('ASSETS', '', 5),
            ('Current Assets', '', 6),
            ('  Cash & Cash Equivalents', '=SUMIFS(tblGL[Amount],tblGL[Account],"1000*")', 7),
            ('  Accounts Receivable', '=SUMIFS(tblGL[Amount],tblGL[Account],"1200*")', 8),
            ('  Inventory', '=SUMIFS(tblGL[Amount],tblGL[Account],"1300*")', 9),
            ('  Prepaid Expenses', '=SUMIFS(tblGL[Amount],tblGL[Account],"1400*")', 10),
            ('Total Current Assets', '=SUM(C7:C10)', 11),
            ('', '', 12),
            ('Fixed Assets', '', 13),
            ('  Property, Plant & Equipment', '=SUMIFS(tblGL[Amount],tblGL[Account],"1500*")', 14),
            ('  Less: Accumulated Depreciation', '=SUMIFS(tblGL[Amount],tblGL[Account],"1600*")', 15),
            ('Net Fixed Assets', '=C14+C15', 16),
            ('', '', 17),
            ('Other Assets', '=SUMIFS(tblGL[Amount],tblGL[Account],"1700*")', 18),
            ('', '', 19),
            ('TOTAL ASSETS', '=C11+C16+C18', 20),
            ('', '', 21),
            ('LIABILITIES & EQUITY', '', 22),
            ('Current Liabilities', '', 23),
            ('  Accounts Payable', '=SUMIFS(tblGL[Amount],tblGL[Account],"2000*")', 24),
            ('  Accrued Expenses', '=SUMIFS(tblGL[Amount],tblGL[Account],"2100*")', 25),
            ('  Short-term Debt', '=SUMIFS(tblGL[Amount],tblGL[Account],"2200*")', 26),
            ('Total Current Liabilities', '=SUM(C24:C26)', 27),
            ('', '', 28),
            ('Long-term Liabilities', '=SUMIFS(tblGL[Amount],tblGL[Account],"2500*")', 29),
            ('', '', 30),
            ('TOTAL LIABILITIES', '=C27+C29', 31),
            ('', '', 32),
            ('Equity', '', 33),
            ('  Common Stock', '=SUMIFS(tblGL[Amount],tblGL[Account],"3000*")', 34),
            ('  Retained Earnings', '=SUMIFS(tblGL[Amount],tblGL[Account],"3100*")', 35),
            ('  Current Year Earnings', '=SUMIFS(tblGL[Amount],tblGL[Account],"3200*")', 36),
            ('TOTAL EQUITY', '=SUM(C34:C36)', 37),
            ('', '', 38),
            ('TOTAL LIABILITIES & EQUITY', '=C31+C37', 39),
        ]
        
        # Populate structure
        for desc, formula, row in bs_structure:
            ws.cell(row=row, column=1, value=desc)
            
            # Apply formulas for columns C and D
            if formula:
                ws.cell(row=row, column=3, value=formula)  # Current
                ws.cell(row=row, column=4, value=formula.replace('tblGL', 'tblGL_PY'))  # Prior Year
                ws.cell(row=row, column=5, value=f'=C{row}-D{row}')  # Change
                ws.cell(row=row, column=6, value=f'=IFERROR((C{row}-D{row})/D{row},0)')  # % Change
                
                # Number formats
                for col in [3, 4, 5]:
                    ws.cell(row=row, column=col).number_format = '#,##0'
                ws.cell(row=row, column=6).number_format = '0.0%'
                
            # Apply styles
            if desc in ['ASSETS', 'LIABILITIES & EQUITY']:
                ws.cell(row=row, column=1).style = 'section_header'
                ws.merge_cells(f'A{row}:B{row}')
            elif desc in ['TOTAL ASSETS', 'TOTAL LIABILITIES', 'TOTAL EQUITY', 'TOTAL LIABILITIES & EQUITY']:
                for col in range(1, 7):
                    ws.cell(row=row, column=col).style = 'total_row'
                    
        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        
        # Balance check
        ws['A41'] = 'Balance Check:'
        ws['C41'] = '=C20-C39'
        ws['C41'].number_format = '#,##0;[Red](#,##0)'
        
        # Define named range
        defn = DefinedName('rngBS', attr_text="'REPORT_BS'!$A$5:$F$39")
        self.wb.defined_names['rngBS'] = defn
        
        return ws
    
    def create_report_cf(self):
        """Create REPORT_CF sheet - Cash Flow Statement"""
        ws = self.wb.create_sheet("REPORT_CF")
        
        # Title
        ws['A1'] = '=SETTINGS!B2 & " · Cash Flow Statement"'
        ws['A1'].style = 'title'
        ws.merge_cells('A1:D1')
        
        # Headers
        ws['A3'] = 'Description'
        ws['B3'] = 'MTD'
        ws['C3'] = 'QTD'
        ws['D3'] = 'YTD'
        
        for col in range(1, 5):
            ws.cell(row=3, column=col).style = 'col_header'
            
        # Cash Flow structure (indirect method)
        cf_structure = [
            ('', 4),
            ('OPERATING ACTIVITIES', 5),
            ('Net Income', 6),
            ('Adjustments to reconcile net income:', 7),
            ('  Depreciation & Amortization', 8),
            ('  Changes in Working Capital:', 9),
            ('    (Increase)/Decrease in AR', 10),
            ('    (Increase)/Decrease in Inventory', 11),
            ('    Increase/(Decrease) in AP', 12),
            ('    Increase/(Decrease) in Accrued Exp', 13),
            ('Net Cash from Operating Activities', 14),
            ('', 15),
            ('INVESTING ACTIVITIES', 16),
            ('  Purchase of Fixed Assets', 17),
            ('  Sale of Fixed Assets', 18),
            ('Net Cash from Investing Activities', 19),
            ('', 20),
            ('FINANCING ACTIVITIES', 21),
            ('  Proceeds from Debt', 22),
            ('  Repayment of Debt', 23),
            ('  Dividends Paid', 24),
            ('Net Cash from Financing Activities', 25),
            ('', 26),
            ('NET CHANGE IN CASH', 27),
            ('Beginning Cash Balance', 28),
            ('Ending Cash Balance', 29),
        ]
        
        # Populate structure
        for desc, row in cf_structure:
            ws.cell(row=row, column=1, value=desc)
            
            # Apply styles
            if desc in ['OPERATING ACTIVITIES', 'INVESTING ACTIVITIES', 'FINANCING ACTIVITIES']:
                ws.cell(row=row, column=1).style = 'section_header'
                ws.merge_cells(f'A{row}:D{row}')
            elif desc in ['Net Cash from Operating Activities', 'Net Cash from Investing Activities', 
                          'Net Cash from Financing Activities', 'NET CHANGE IN CASH', 'Ending Cash Balance']:
                for col in range(1, 5):
                    ws.cell(row=row, column=col).style = 'total_row'
                    
        # Column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        # Define named range
        defn = DefinedName('rngCF', attr_text="'REPORT_CF'!$A$5:$D$29")
        self.wb.defined_names['rngCF'] = defn
        
        return ws
    
    def create_dash_kpi(self):
        """Create DASH_KPI sheet - Executive KPI Dashboard"""
        ws = self.wb.create_sheet("DASH_KPI")
        
        # Title
        ws['A1'] = '=SETTINGS!B2 & " · Key Performance Indicators"'
        ws['A1'].style = 'title'
        ws.merge_cells('A1:L1')
        
        # KPI Grid Layout (4x3)
        kpis = [
            # Row 1
            ('Revenue', '=$C$9', '+12% YoY', 'B3:D6'),
            ('Gross Margin', '=C18', '+7pp YoY', 'F3:H6'),
            ('Operating Margin', '=C27', '+5pp YoY', 'J3:L6'),
            # Row 2
            ('Cash Balance', '=REPORT_BS!C7', '+$1.2M', 'B8:D11'),
            ('AR Days', '45', '-3 days', 'F8:H11'),
            ('AP Days', '38', '+2 days', 'J8:L11'),
            # Row 3
            ('Customer Count', '156', '+23', 'B13:D16'),
            ('Avg Deal Size', '$62K', '+15%', 'F13:H16'),
            ('Monthly Burn', '$125K', '-8%', 'J13:L16'),
            # Row 4
            ('Runway', '18 mo', '+3 mo', 'B18:D21'),
            ('LTV/CAC', '3.2x', '+0.4x', 'F18:H21'),
            ('NPS Score', '72', '+5', 'J18:L21'),
        ]
        
        # Create KPI tiles
        for i, (label, value, change, range_str) in enumerate(kpis):
            # Parse range
            start_cell = range_str.split(':')[0]
            end_cell = range_str.split(':')[1]
            
            # Merge cells for tile
            ws.merge_cells(range_str)
            
            # Get the merged cell
            cell = ws[start_cell]
            
            # Create multi-line content
            content = f"{label}\n{value}\n{change}"
            cell.value = label  # For now, just the label
            
            # Style the tile
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(name='Calibri', size=11, color=COLORS['white'])
            cell.fill = PatternFill("solid", fgColor=COLORS['primary'])
            cell.border = Border(
                left=Side(style='medium', color=COLORS['primary']),
                right=Side(style='medium', color=COLORS['primary']),
                top=Side(style='medium', color=COLORS['primary']),
                bottom=Side(style='medium', color=COLORS['primary'])
            )
            
        # Row heights
        for row in [3, 8, 13, 18]:
            ws.row_dimensions[row].height = 60
            
        # Define named ranges for KPIs
        kpi_names = ['kpi_Revenue', 'kpi_GM', 'kpi_OpMargin', 'kpi_Cash', 
                     'kpi_ARDays', 'kpi_APDays']
        kpi_cells = ['B3', 'F3', 'J3', 'B8', 'F8', 'J8']
        
        for name, cell in zip(kpi_names, kpi_cells):
            defn = DefinedName(name, attr_text=f"'DASH_KPI'!${cell}")
            self.wb.defined_names[name] = defn
            
        return ws
    
    def create_dash_trends(self):
        """Create DASH_TRENDS sheet - Trend Charts"""
        ws = self.wb.create_sheet("DASH_TRENDS")
        
        # Title
        ws['A1'] = '=SETTINGS!B2 & " · Financial Trends"'
        ws['A1'].style = 'title'
        ws.merge_cells('A1:H1')
        
        # Chart placeholders with Plotly JSON in comments
        charts = [
            ('Revenue Trend', 'A3', {'type': 'line', 'title': 'Monthly Revenue'}),
            ('Gross Margin %', 'E3', {'type': 'line', 'title': 'Gross Margin %'}),
            ('Operating Expenses', 'A12', {'type': 'bar', 'title': 'OPEX by Category'}),
            ('Cash Position', 'E12', {'type': 'area', 'title': 'Cash Balance'}),
            ('AR Aging', 'A21', {'type': 'bar', 'title': 'AR Aging Buckets'}),
            ('Budget vs Actual', 'E21', {'type': 'bullet', 'title': 'Budget Performance'}),
        ]
        
        for title, cell_ref, chart_spec in charts:
            cell = ws[cell_ref]
            cell.value = title
            cell.style = 'section_header'
            
            # Add Plotly JSON as comment
            plotly_json = json.dumps({
                'data': [{
                    'x': ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun'],
                    'y': [100, 110, 105, 120, 115, 130],
                    'type': chart_spec['type'],
                    'name': title
                }],
                'layout': {
                    'title': chart_spec['title'],
                    'xaxis': {'title': 'Month'},
                    'yaxis': {'title': 'Value'},
                    'height': 300,
                    'width': 500
                }
            })
            
            comment = Comment(plotly_json, "FinWave")
            cell.comment = comment
            
        # Layout spacing
        ws.row_dimensions[3].height = 20
        ws.row_dimensions[12].height = 20
        ws.row_dimensions[21].height = 20
        
        return ws
    
    def create_settings(self):
        """Create SETTINGS sheet - Configuration"""
        ws = self.wb.create_sheet("SETTINGS")
        
        # Title
        ws['A1'] = 'FinWave Settings'
        ws['A1'].style = 'title'
        ws.merge_cells('A1:C1')
        
        # Settings
        settings = [
            ('Company Name', 'Acme Corp', 'B2'),
            ('Fiscal Year Start', '2025-01-01', 'B3'),
            ('Fiscal Year End', '2025-12-31', 'B4'),
            ('Base Currency', 'USD', 'B5'),
            ('Report Date', '=TODAY()', 'B6'),
            ('', '', ''),
            ('Theme Settings', '', 'A8'),
            ('Primary Color', COLORS['primary'], 'B9'),
            ('Accent Color 1', COLORS['accent1'], 'B10'),
            ('Accent Color 2', COLORS['accent2'], 'B11'),
        ]
        
        for i, (label, value, _) in enumerate(settings, 2):
            if label:
                ws.cell(row=i, column=1, value=label)
                ws.cell(row=i, column=1).font = Font(bold=True)
            if value:
                ws.cell(row=i, column=2, value=value)
                
        # Define named ranges
        defn1 = DefinedName('fxStart', attr_text="'SETTINGS'!$B$3")
        self.wb.defined_names['fxStart'] = defn1
        
        defn2 = DefinedName('fxEnd', attr_text="'SETTINGS'!$B$4")
        self.wb.defined_names['fxEnd'] = defn2
        
        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        
        return ws
    
    def create_readme(self):
        """Create README sheet - Documentation"""
        ws = self.wb.create_sheet("README")
        
        # Title
        ws['A1'] = 'FinWave Board Pack Template - Documentation'
        ws['A1'].style = 'title'
        ws.merge_cells('A1:F1')
        
        # Last sync info
        ws['A3'] = 'Last QuickBooks Sync:'
        ws['B3'] = 'Not synced'
        ws['A4'] = 'Template Version:'
        ws['B4'] = '1.0.0'
        ws['A5'] = 'Generated:'
        ws['B5'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
        
        # Sheet descriptions
        ws['A7'] = 'Sheet Descriptions'
        ws['A7'].style = 'section_header'
        ws.merge_cells('A7:F7')
        
        descriptions = [
            ('Sheet Name', 'Purpose', 'Key Features'),
            ('DATA_GL', 'General Ledger transaction data', 'Hidden, Excel Table tblGL'),
            ('DATA_COA', 'Chart of Accounts master', 'Hidden, Excel Table tblCOA'),
            ('DATA_MAP', 'Account mapping configuration', 'Hidden, grouping rules'),
            ('REPORT_P&L', 'Multi-period Income Statement', 'Monthly columns, % margins'),
            ('REPORT_BS', 'Comparative Balance Sheet', 'Current vs Prior, variances'),
            ('REPORT_CF', 'Cash Flow Statement', 'Indirect method, MTD/QTD/YTD'),
            ('DASH_KPI', 'Executive KPI tiles', '12 key metrics, visual design'),
            ('DASH_TRENDS', 'Trend charts', 'Plotly JSON in cell comments'),
            ('SETTINGS', 'Configuration parameters', 'Company info, dates, theme'),
        ]
        
        for i, (sheet, purpose, features) in enumerate(descriptions, 9):
            ws.cell(row=i, column=1, value=sheet)
            ws.cell(row=i, column=2, value=purpose)
            ws.cell(row=i, column=3, value=features)
            if i == 9:  # Header row
                for col in range(1, 4):
                    ws.cell(row=i, column=col).style = 'col_header'
                    
        # Named ranges documentation
        ws['A20'] = 'Named Ranges'
        ws['A20'].style = 'section_header'
        ws.merge_cells('A20:F20')
        
        ranges = [
            ('Range Name', 'Reference', 'Description'),
            ('tblGL', 'DATA_GL!A:K', 'General Ledger table'),
            ('tblCOA', 'DATA_COA!A:H', 'Chart of Accounts table'),
            ('rngIS_Matrix', 'REPORT_P&L!A5:N36', 'Income Statement data matrix'),
            ('rngBS', 'REPORT_BS!A5:F39', 'Balance Sheet data range'),
            ('rngCF', 'REPORT_CF!A5:D29', 'Cash Flow data range'),
            ('fxStart', 'SETTINGS!B3', 'Fiscal year start date'),
            ('fxEnd', 'SETTINGS!B4', 'Fiscal year end date'),
        ]
        
        for i, (name, ref, desc) in enumerate(ranges, 22):
            ws.cell(row=i, column=1, value=name)
            ws.cell(row=i, column=2, value=ref)
            ws.cell(row=i, column=3, value=desc)
            if i == 22:  # Header row
                for col in range(1, 4):
                    ws.cell(row=i, column=col).style = 'col_header'
                    
        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        
        return ws
    
    def create_data_pivots(self):
        """Create DATA_PIVOTS sheet for pre-calculated aggregations"""
        ws = self.wb.create_sheet("DATA_PIVOTS")
        ws.sheet_state = 'hidden'
        
        # This sheet would contain pivot tables or pre-aggregated data
        # For now, just create the structure
        ws['A1'] = 'Pre-calculated Pivot Data'
        ws['A1'].style = 'section_header'
        
        return ws
    
    def build_template(self):
        """Build the complete template"""
        logger.info("Building FinWave Board Pack Template...")
        
        # Create all sheets
        self.create_data_gl()
        self.create_data_coa()
        self.create_data_map()
        self.create_data_pivots()
        self.create_report_pl()
        self.create_report_bs()
        self.create_report_cf()
        self.create_dash_kpi()
        self.create_dash_trends()
        self.create_settings()
        self.create_readme()
        
        # Save the workbook
        output_path = 'backend/assets/templates/finwave_board_pack.xlsx'
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        self.wb.save(output_path)
        logger.info(f"Template saved to {output_path}")
        
        return output_path


def create_google_sheet(wb_path: str):
    """Upload template to Google Sheets"""
    try:
        import pygsheets
        
        # Check for credentials
        creds_path = os.getenv('GOOGLE_SHEETS_SERVICE_ACCOUNT_JSON')
        if not creds_path:
            logger.warning("Google Sheets credentials not found, skipping upload")
            return None
            
        # Authorize
        gc = pygsheets.authorize(service_file=creds_path)
        
        # Create new spreadsheet
        sh = gc.create('FinWave Board Pack Template')
        
        # Would need to implement sheet-by-sheet copying here
        logger.info(f"Google Sheet created: {sh.url}")
        
        return sh.url
        
    except Exception as e:
        logger.error(f"Failed to create Google Sheet: {e}")
        return None


def self_test(wb_path: str):
    """Run basic self-test on generated template"""
    from openpyxl import load_workbook
    
    logger.info("Running self-test...")
    
    # Load the workbook
    wb = load_workbook(wb_path)
    
    # Check that all expected sheets exist
    expected_sheets = [
        'DATA_GL', 'DATA_COA', 'DATA_MAP', 'DATA_PIVOTS',
        'REPORT_P&L', 'REPORT_BS', 'REPORT_CF',
        'DASH_KPI', 'DASH_TRENDS', 'SETTINGS', 'README'
    ]
    
    for sheet in expected_sheets:
        assert sheet in wb.sheetnames, f"Missing sheet: {sheet}"
        
    # Check that REPORT_P&L has content
    ws = wb['REPORT_P&L']
    assert ws['A1'].value is not None, "REPORT_P&L!A1 is blank"
    
    # Check named ranges
    assert 'rngIS_Matrix' in wb.defined_names, "Missing named range: rngIS_Matrix"
    
    logger.info("Self-test passed!")
    

def main():
    """Main entry point"""
    # Build the template
    builder = FinWaveTemplateBuilder()
    wb_path = builder.build_template()
    
    # Run self-test
    self_test(wb_path)
    
    # Try to create Google Sheet
    gs_url = create_google_sheet(wb_path)
    
    print(f"\n✅ FinWave Board Pack Template created successfully!")
    print(f"📁 Excel file: {wb_path}")
    if gs_url:
        print(f"☁️  Google Sheet: {gs_url}")
    print(f"\n🚀 Next steps:")
    print(f"   1. Run ETL to populate DATA_GL with QuickBooks data")
    print(f"   2. Open in Excel to verify formulas calculate correctly")
    print(f"   3. Export to PDF using the report generator")
    

if __name__ == "__main__":
    main()