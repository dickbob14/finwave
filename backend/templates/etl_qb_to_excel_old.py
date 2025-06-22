#!/usr/bin/env python3
"""
ETL Script: QuickBooks to Excel Template
Populates the FinWave Excel template DATA_GL sheet with real QuickBooks data
"""

import os
import sys
import json
import pandas as pd
import requests
from datetime import datetime, timedelta
from typing import Dict, List, Any
import logging
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Add parent directory to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class QuickBooksToExcelETL:
    def __init__(self, qb_server_url: str = "http://localhost:8000", template_path: str = None):
        self.qb_server_url = qb_server_url
        self.template_path = template_path or "finwave_board_pack.xlsx"
        
    def fetch_qb_data(self) -> Dict[str, Any]:
        """Fetch data from QuickBooks server"""
        try:
            logger.info("Fetching QuickBooks data from server...")
            
            # Get transactions
            response = requests.get(f"{self.qb_server_url}/real/transactions")
            response.raise_for_status()
            transactions = response.json()
            
            # Get company info
            company_response = requests.get(f"{self.qb_server_url}/real/company")
            company_response.raise_for_status()
            company = company_response.json()
            
            logger.info(f"Fetched {len(transactions.get('invoices', []))} invoices and {len(transactions.get('expenses', []))} expenses")
            return {
                "transactions": transactions,
                "company": company
            }
            
        except requests.exceptions.RequestException as e:
            logger.error(f"Failed to fetch QuickBooks data: {e}")
            raise
    
    def transform_to_gl_format(self, qb_data: Dict[str, Any]) -> pd.DataFrame:
        """Transform QuickBooks data into General Ledger format for Excel"""
        gl_entries = []
        
        transactions = qb_data["transactions"]
        
        # Process invoices (revenue)
        for invoice in transactions.get("invoices", []):
            try:
                # Revenue entry (credit)
                gl_entries.append({
                    "Date": datetime.strptime(invoice["TxnDate"], "%Y-%m-%d").strftime("%m/%d/%Y"),
                    "Account": "4000 - Revenue",
                    "Account_Code": "4000",
                    "Account_Type": "Income",
                    "Description": f"Invoice #{invoice.get('DocNumber', 'N/A')} - {invoice.get('CustomerRef', {}).get('name', 'Unknown Customer')}",
                    "Reference": invoice.get("DocNumber", ""),
                    "Debit": 0,
                    "Credit": float(invoice.get("TotalAmt", 0)),
                    "Net_Amount": float(invoice.get("TotalAmt", 0)),
                    "Running_Balance": 0,  # Will calculate later
                    "Customer": invoice.get("CustomerRef", {}).get("name", ""),
                    "Vendor": "",
                    "Class": "",
                    "Location": "",
                    "Memo": invoice.get("PrivateNote", ""),
                    "TxnID": invoice.get("Id", ""),
                    "Source": "QuickBooks Invoice"
                })
                
                # Accounts Receivable entry (debit) 
                gl_entries.append({
                    "Date": datetime.strptime(invoice["TxnDate"], "%Y-%m-%d").strftime("%m/%d/%Y"),
                    "Account": "1200 - Accounts Receivable",
                    "Account_Code": "1200", 
                    "Account_Type": "Asset",
                    "Description": f"A/R Invoice #{invoice.get('DocNumber', 'N/A')} - {invoice.get('CustomerRef', {}).get('name', 'Unknown Customer')}",
                    "Reference": invoice.get("DocNumber", ""),
                    "Debit": float(invoice.get("TotalAmt", 0)),
                    "Credit": 0,
                    "Net_Amount": float(invoice.get("TotalAmt", 0)),
                    "Running_Balance": 0,
                    "Customer": invoice.get("CustomerRef", {}).get("name", ""),
                    "Vendor": "",
                    "Class": "",
                    "Location": "",
                    "Memo": invoice.get("PrivateNote", ""),
                    "TxnID": invoice.get("Id", ""),
                    "Source": "QuickBooks Invoice"
                })
                
            except Exception as e:
                logger.warning(f"Error processing invoice {invoice.get('Id', 'unknown')}: {e}")
                continue
        
        # Process expenses
        for expense in transactions.get("expenses", []):
            try:
                account_name = expense.get("AccountRef", {}).get("name", "6000 - General Expenses")
                account_code = "6000"  # Default expense code
                
                # Map common expense accounts
                if "travel" in account_name.lower():
                    account_code = "6100"
                elif "office" in account_name.lower():
                    account_code = "6200"
                elif "marketing" in account_name.lower():
                    account_code = "6300"
                
                # Expense entry (debit)
                gl_entries.append({
                    "Date": datetime.strptime(expense["TxnDate"], "%Y-%m-%d").strftime("%m/%d/%Y"),
                    "Account": f"{account_code} - {account_name}",
                    "Account_Code": account_code,
                    "Account_Type": "Expense",
                    "Description": f"Expense - {expense.get('PrivateNote', 'General expense')}",
                    "Reference": expense.get("DocNumber", ""),
                    "Debit": float(expense.get("TotalAmt", 0)),
                    "Credit": 0,
                    "Net_Amount": -float(expense.get("TotalAmt", 0)),
                    "Running_Balance": 0,
                    "Customer": "",
                    "Vendor": expense.get("EntityRef", {}).get("name", ""),
                    "Class": "",
                    "Location": "",
                    "Memo": expense.get("PrivateNote", ""),
                    "TxnID": expense.get("Id", ""),
                    "Source": "QuickBooks Expense"
                })
                
                # Cash/AP entry (credit)
                payment_account = "1000 - Cash" if expense.get("PaymentType") == "Cash" else "2000 - Accounts Payable"
                payment_code = "1000" if expense.get("PaymentType") == "Cash" else "2000"
                
                gl_entries.append({
                    "Date": datetime.strptime(expense["TxnDate"], "%Y-%m-%d").strftime("%m/%d/%Y"),
                    "Account": payment_account,
                    "Account_Code": payment_code,
                    "Account_Type": "Asset" if payment_code == "1000" else "Liability",
                    "Description": f"Payment for expense - {expense.get('PrivateNote', 'General expense')}",
                    "Reference": expense.get("DocNumber", ""),
                    "Debit": 0,
                    "Credit": float(expense.get("TotalAmt", 0)),
                    "Net_Amount": -float(expense.get("TotalAmt", 0)),
                    "Running_Balance": 0,
                    "Customer": "",
                    "Vendor": expense.get("EntityRef", {}).get("name", ""),
                    "Class": "",
                    "Location": "",
                    "Memo": expense.get("PrivateNote", ""),
                    "TxnID": expense.get("Id", ""),
                    "Source": "QuickBooks Expense"
                })
                
            except Exception as e:
                logger.warning(f"Error processing expense {expense.get('Id', 'unknown')}: {e}")
                continue
        
        # Convert to DataFrame and sort by date
        df = pd.DataFrame(gl_entries)
        if not df.empty:
            df['Date'] = pd.to_datetime(df['Date'], format='%m/%d/%Y')
            df = df.sort_values('Date')
            df['Date'] = df['Date'].dt.strftime('%m/%d/%Y')
            
            # Calculate running balance by account
            df['Running_Balance'] = df.groupby('Account_Code')['Net_Amount'].cumsum()
        
        logger.info(f"Transformed {len(df)} general ledger entries")
        return df
    
    def load_template(self) -> Any:
        """Load the Excel template"""
        try:
            if not os.path.exists(self.template_path):
                raise FileNotFoundError(f"Template not found: {self.template_path}")
            
            logger.info(f"Loading Excel template: {self.template_path}")
            workbook = load_workbook(self.template_path)
            return workbook
            
        except Exception as e:
            logger.error(f"Failed to load template: {e}")
            raise
    
    def populate_data_gl_sheet(self, workbook: Any, gl_df: pd.DataFrame) -> None:
        """Populate the DATA_GL sheet with transformed data"""
        try:
            if "DATA_GL" not in workbook.sheetnames:
                logger.error("DATA_GL sheet not found in template")
                raise ValueError("DATA_GL sheet not found")
            
            ws = workbook["DATA_GL"]
            logger.info("Populating DATA_GL sheet...")
            
            # Clear existing data (keep headers)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.value = None
            
            # Add data starting from row 2
            for r_idx, (index, row) in enumerate(gl_df.iterrows(), start=2):
                ws[f"A{r_idx}"] = row["Date"]
                ws[f"B{r_idx}"] = row["Account"]
                ws[f"C{r_idx}"] = row["Account_Code"]
                ws[f"D{r_idx}"] = row["Account_Type"]
                ws[f"E{r_idx}"] = row["Description"]
                ws[f"F{r_idx}"] = row["Reference"]
                ws[f"G{r_idx}"] = row["Debit"]
                ws[f"H{r_idx}"] = row["Credit"]
                ws[f"I{r_idx}"] = row["Net_Amount"]
                ws[f"J{r_idx}"] = row["Running_Balance"]
                ws[f"K{r_idx}"] = row["Customer"]
                ws[f"L{r_idx}"] = row["Vendor"]
                ws[f"M{r_idx}"] = row["Class"]
                ws[f"N{r_idx}"] = row["Location"]
                ws[f"O{r_idx}"] = row["Memo"]
                ws[f"P{r_idx}"] = row["TxnID"]
                ws[f"Q{r_idx}"] = row["Source"]
            
            logger.info(f"Populated {len(gl_df)} rows in DATA_GL sheet")
            
        except Exception as e:
            logger.error(f"Failed to populate DATA_GL sheet: {e}")
            raise
    
    def update_settings_sheet(self, workbook: Any, qb_data: Dict[str, Any]) -> None:
        """Update the SETTINGS sheet with company info and refresh date"""
        try:
            if "SETTINGS" not in workbook.sheetnames:
                logger.warning("SETTINGS sheet not found, skipping update")
                return
            
            ws = workbook["SETTINGS"]
            company_info = qb_data.get("company", {}).get("QueryResponse", {}).get("CompanyInfo", [{}])[0]
            
            # Update company name (B2)
            if company_info.get("CompanyName"):
                ws["B2"] = company_info["CompanyName"]
            
            # Update data refresh timestamp (B4)
            ws["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Update period end date (B6) 
            transactions = qb_data.get("transactions", {})
            if transactions.get("period", {}).get("end"):
                ws["B6"] = transactions["period"]["end"]
            
            logger.info("Updated SETTINGS sheet with company info and refresh time")
            
        except Exception as e:
            logger.warning(f"Failed to update SETTINGS sheet: {e}")
    
    def save_workbook(self, workbook: Any, output_path: str = None) -> str:
        """Save the populated workbook"""
        try:
            if not output_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = f"finwave_board_pack_populated_{timestamp}.xlsx"
            
            workbook.save(output_path)
            logger.info(f"Saved populated workbook: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to save workbook: {e}")
            raise
    
    def run_etl(self, output_path: str = None) -> str:
        """Run the complete ETL process"""
        try:
            logger.info("Starting QuickBooks to Excel ETL process...")
            
            # Step 1: Fetch QuickBooks data
            qb_data = self.fetch_qb_data()
            
            # Step 2: Transform to GL format
            gl_df = self.transform_to_gl_format(qb_data)
            
            if gl_df.empty:
                logger.warning("No data to populate - GL DataFrame is empty")
                return None
            
            # Step 3: Load template
            workbook = self.load_template()
            
            # Step 4: Populate DATA_GL sheet
            self.populate_data_gl_sheet(workbook, gl_df)
            
            # Step 5: Update settings
            self.update_settings_sheet(workbook, qb_data)
            
            # Step 6: Save populated workbook
            output_file = self.save_workbook(workbook, output_path)
            
            logger.info("ETL process completed successfully!")
            logger.info(f"✅ Populated Excel template: {output_file}")
            logger.info(f"📊 Total GL entries: {len(gl_df)}")
            logger.info(f"💰 Revenue entries: {len(gl_df[gl_df['Account_Type'] == 'Income'])}")
            logger.info(f"💸 Expense entries: {len(gl_df[gl_df['Account_Type'] == 'Expense'])}")
            
            return output_file
            
        except Exception as e:
            logger.error(f"ETL process failed: {e}")
            raise

def main():
    """Main entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(description="Populate FinWave Excel template with QuickBooks data")
    parser.add_argument("--template", default="finwave_board_pack.xlsx", help="Path to Excel template")
    parser.add_argument("--output", help="Output file path")
    parser.add_argument("--qb-server", default="http://localhost:8000", help="QuickBooks server URL")
    
    args = parser.parse_args()
    
    try:
        etl = QuickBooksToExcelETL(
            qb_server_url=args.qb_server,
            template_path=args.template
        )
        
        output_file = etl.run_etl(args.output)
        
        if output_file:
            print(f"\n🎉 SUCCESS! Populated Excel template saved as: {output_file}")
            print("\n📋 Next steps:")
            print("1. Open the file in Excel to verify formulas are calculating correctly")
            print("2. Review the financial reports on REPORT_P&L, REPORT_BS, and REPORT_CF sheets")
            print("3. Check the executive dashboard on DASH_KPI and DASH_TRENDS sheets")
            print("4. Use the template to generate PDF board pack reports")
        else:
            print("❌ No data was available to populate the template")
            
    except Exception as e:
        print(f"❌ ETL failed: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()