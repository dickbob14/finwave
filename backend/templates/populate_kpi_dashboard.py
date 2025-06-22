#!/usr/bin/env python3
"""
Populate Cube KPI Dashboard template with multi-dimensional business data
Usage: python populate_kpi_dashboard.py --since 2024-01-01 --until 2024-12-31 [--sheet-id SHEET_ID]
"""

import argparse
import json
import logging
import os
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Any

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Optional Google Sheets support
try:
    import pygsheets
    GSHEETS_AVAILABLE = True
except ImportError:
    GSHEETS_AVAILABLE = False

logger = logging.getLogger(__name__)

# Schema for KPI Dashboard data
KPI_DASHBOARD_SCHEMA = {
    "Drivers": {
        "columns": [
            {"name": "Entity", "type": "string"},
            {"name": "Department", "type": "string"},
            {"name": "Product", "type": "string"},
            {"name": "Market", "type": "string"},
            {"name": "Date", "type": "date"},
            {"name": "New Customers", "type": "number"},
            {"name": "Revenue", "type": "number"},
            {"name": "Orders", "type": "number"},
            {"name": "COGS", "type": "number"},
            {"name": "OpEx", "type": "number"},
            {"name": "Headcount", "type": "number"},
            {"name": "Marketing Spend", "type": "number"},
            {"name": "Sales Spend", "type": "number"},
            {"name": "R&D Spend", "type": "number"},
            {"name": "G&A Spend", "type": "number"}
        ]
    },
    "Filters": {
        "Department": ["Sales", "Marketing", "Engineering", "G&A", "All"],
        "Entity": ["US Corp", "UK Ltd", "DE GmbH", "All"],
        "Type": ["Actuals", "Budget", "Forecast"],
        "Month": ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    }
}

class KPIDashboardPopulator:
    """Populates Cube KPI Dashboard with multi-dimensional business data"""
    
    def __init__(self, template_path: str):
        self.template_path = Path(template_path)
        self.wb = None
        
    def load_template(self) -> None:
        """Load the Excel template preserving formulas and formatting"""
        logger.info(f"Loading template: {self.template_path}")
        self.wb = load_workbook(self.template_path, keep_vba=True, data_only=False)
        
    def fetch_business_metrics(self, start_date: str, end_date: str) -> pd.DataFrame:
        """
        Fetch multi-dimensional business metrics
        In production, this would aggregate from QuickBooks, CRM, HRIS, etc.
        """
        # Generate date range
        start = pd.to_datetime(start_date)
        end = pd.to_datetime(end_date)
        date_range = pd.date_range(start, end, freq='D')
        
        # Dimensions
        entities = ['US Corp', 'UK Ltd', 'DE GmbH']
        departments = ['Sales', 'Marketing', 'Engineering', 'G&A']
        products = ['Product A', 'Product B', 'Product C']
        markets = ['North America', 'Europe', 'Asia Pacific']
        
        # Generate sample data
        data_rows = []
        
        for date in date_range:
            # Daily data for each entity/department combination
            for entity in entities:
                for dept in departments:
                    # Vary metrics by department type
                    base_revenue = 10000 if dept == 'Sales' else 0
                    base_orders = 50 if dept == 'Sales' else 0
                    base_customers = 5 if dept in ['Sales', 'Marketing'] else 0
                    
                    # Department-specific costs
                    dept_costs = {
                        'Sales': {'Sales Spend': 5000, 'OpEx': 3000},
                        'Marketing': {'Marketing Spend': 8000, 'OpEx': 2000},
                        'Engineering': {'R&D Spend': 15000, 'OpEx': 5000},
                        'G&A': {'G&A Spend': 4000, 'OpEx': 1000}
                    }
                    
                    # Aggregate by product/market
                    for product in products:
                        for market in markets:
                            # Add some randomness and trends
                            trend_factor = 1 + (date.month - 1) * 0.05
                            
                            row = {
                                'Entity': entity,
                                'Department': dept,
                                'Product': product,
                                'Market': market,
                                'Date': date,
                                'New Customers': int(base_customers * trend_factor),
                                'Revenue': base_revenue * trend_factor,
                                'Orders': int(base_orders * trend_factor),
                                'COGS': base_revenue * 0.4 if base_revenue > 0 else 0,
                                'OpEx': dept_costs[dept]['OpEx'],
                                'Headcount': {'Sales': 10, 'Marketing': 8, 
                                             'Engineering': 25, 'G&A': 5}[dept],
                                'Marketing Spend': dept_costs[dept].get('Marketing Spend', 0),
                                'Sales Spend': dept_costs[dept].get('Sales Spend', 0),
                                'R&D Spend': dept_costs[dept].get('R&D Spend', 0),
                                'G&A Spend': dept_costs[dept].get('G&A Spend', 0)
                            }
                            data_rows.append(row)
        
        df = pd.DataFrame(data_rows)
        
        # Aggregate to monthly for dashboard
        df['Month'] = df['Date'].dt.to_period('M')
        
        return df
    
    def populate_drivers_sheet(self, metrics_df: pd.DataFrame) -> None:
        """Populate the Drivers data sheet"""
        ws = self.wb['Drivers']
        
        # Clear existing data (keep headers in row 1)
        ws.delete_rows(2, ws.max_row)
        
        # Aggregate data to monthly level for the dashboard
        monthly_agg = metrics_df.groupby(
            ['Entity', 'Department', 'Product', 'Market', 'Month']
        ).agg({
            'New Customers': 'sum',
            'Revenue': 'sum',
            'Orders': 'sum',
            'COGS': 'sum',
            'OpEx': 'sum',
            'Headcount': 'mean',  # Average headcount
            'Marketing Spend': 'sum',
            'Sales Spend': 'sum',
            'R&D Spend': 'sum',
            'G&A Spend': 'sum'
        }).reset_index()
        
        # Convert Month period to date for Excel
        monthly_agg['Date'] = monthly_agg['Month'].dt.to_timestamp()
        monthly_agg = monthly_agg.drop('Month', axis=1)
        
        # Reorder columns to match template
        column_order = [
            'Entity', 'Department', 'Product', 'Market', 'Date',
            'New Customers', 'Revenue', 'Orders', 'COGS', 'OpEx',
            'Headcount', 'Marketing Spend', 'Sales Spend', 'R&D Spend', 'G&A Spend'
        ]
        monthly_agg = monthly_agg[column_order]
        
        # Write data starting from row 2
        for r_idx, row in enumerate(dataframe_to_rows(monthly_agg, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        logger.info(f"Populated Drivers sheet with {len(monthly_agg)} rows")
    
    def update_filter_lists(self) -> None:
        """Update dropdown filter lists based on actual data"""
        # The template already has these configured, but we could update them
        # dynamically based on the data if needed
        pass
    
    def refresh_pivot_tables(self) -> None:
        """Refresh any pivot tables in the workbook"""
        # Note: This requires xlwings or win32com on Windows
        # For now, we'll rely on Excel to refresh on open
        logger.info("Pivot tables will refresh when workbook is opened in Excel")
    
    def save_populated_file(self, output_path: Optional[str] = None) -> str:
        """Save the populated workbook"""
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f"kpi_dashboard_populated_{timestamp}.xlsx"
        
        self.wb.save(output_path)
        logger.info(f"Saved populated workbook: {output_path}")
        return output_path
    
    def upload_to_google_sheets(self, sheet_id: str, metrics_df: pd.DataFrame) -> str:
        """Upload populated data to Google Sheets"""
        if not GSHEETS_AVAILABLE:
            raise ImportError("pygsheets not installed. Run: pip install pygsheets")
        
        credentials_path = os.getenv('GOOGLE_SHEETS_JSON')
        if not credentials_path:
            raise ValueError("GOOGLE_SHEETS_JSON environment variable not set")
        
        gc = pygsheets.authorize(service_file=credentials_path)
        sh = gc.open_by_key(sheet_id)
        
        # Upload Drivers data
        try:
            drivers_ws = sh.worksheet_by_title('Drivers')
        except:
            drivers_ws = sh.add_worksheet('Drivers')
        
        # Prepare monthly aggregated data
        monthly_agg = metrics_df.groupby(
            ['Entity', 'Department', 'Product', 'Market', metrics_df['Date'].dt.to_period('M')]
        ).agg({
            'New Customers': 'sum',
            'Revenue': 'sum',
            'Orders': 'sum',
            'COGS': 'sum',
            'OpEx': 'sum',
            'Headcount': 'mean',
            'Marketing Spend': 'sum',
            'Sales Spend': 'sum',
            'R&D Spend': 'sum',
            'G&A Spend': 'sum'
        }).reset_index()
        
        # Convert for upload
        monthly_agg['Date'] = monthly_agg['Date'].astype(str)
        
        # Clear and upload
        drivers_ws.clear()
        drivers_ws.set_dataframe(monthly_agg, (1, 1))
        
        # Note: Dashboard sheets with formulas would need special handling
        # For full functionality, you'd recreate the dashboard formulas in Sheets
        
        return f"https://docs.google.com/spreadsheets/d/{sheet_id}"


def main():
    parser = argparse.ArgumentParser(description='Populate KPI Dashboard with business metrics')
    parser.add_argument('--template', default='templates/Cube - KPI Dashboard.xlsx',
                        help='Path to template file')
    parser.add_argument('--since', required=True, help='Start date (YYYY-MM-DD)')
    parser.add_argument('--until', help='End date (YYYY-MM-DD), defaults to today')
    parser.add_argument('--sheet-id', help='Google Sheets ID for cloud upload')
    parser.add_argument('--output', help='Output filename')
    parser.add_argument('--debug', action='store_true', help='Enable debug logging')
    
    args = parser.parse_args()
    
    # Setup logging
    logging.basicConfig(
        level=logging.DEBUG if args.debug else logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # Default end date to today
    if not args.until:
        args.until = datetime.now().strftime('%Y-%m-%d')
    
    # Initialize populator
    populator = KPIDashboardPopulator(args.template)
    
    try:
        # Load template
        populator.load_template()
        
        # Fetch data
        logger.info(f"Fetching metrics from {args.since} to {args.until}")
        metrics_df = populator.fetch_business_metrics(args.since, args.until)
        
        # Populate sheets
        populator.populate_drivers_sheet(metrics_df)
        populator.update_filter_lists()
        populator.refresh_pivot_tables()
        
        # Save file
        output_path = populator.save_populated_file(args.output)
        
        # Upload to Google Sheets if requested
        if args.sheet_id:
            logger.info(f"Uploading to Google Sheets: {args.sheet_id}")
            sheet_url = populator.upload_to_google_sheets(args.sheet_id, metrics_df)
            logger.info(f"Google Sheets URL: {sheet_url}")
        
        print(f"✅ Successfully populated KPI Dashboard: {output_path}")
        print(f"📊 Total data points: {len(metrics_df)}")
        print(f"📅 Period: {args.since} to {args.until}")
        
    except Exception as e:
        logger.error(f"Error populating template: {e}")
        raise


if __name__ == '__main__':
    main()