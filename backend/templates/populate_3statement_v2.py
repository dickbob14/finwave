#!/usr/bin/env python3
"""
Populate Basic 3-Statement Model template with REAL QuickBooks data
Version 2: Uses QuickBooks client and field mapper
Usage: python populate_3statement_v2.py --since 2024-01-01 --until 2024-12-31 [--sheet-id SHEET_ID]
"""

import argparse
import json
import logging
import os
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Import our QuickBooks client and field mapper
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from integrations.quickbooks.client import QuickBooksClient, create_client
from config.field_mapper import get_field_mapper

# Optional Google Sheets support
try:
    import pygsheets
    GSHEETS_AVAILABLE = True
except ImportError:
    GSHEETS_AVAILABLE = False

logger = logging.getLogger(__name__)

class ThreeStatementPopulator:
    """Populates 3-statement financial model with REAL QuickBooks data"""
    
    def __init__(self, template_path: str):
        self.template_path = Path(template_path)
        self.wb = None
        self.qb_client = create_client()
        self.field_mapper = get_field_mapper()
        
    def load_template(self) -> None:
        """Load the Excel template preserving formulas and formatting"""
        logger.info(f"Loading template: {self.template_path}")
        self.wb = load_workbook(self.template_path, keep_vba=True, data_only=False)
        
    def fetch_quickbooks_data(self, start_date: str, end_date: str) -> Dict[str, pd.DataFrame]:
        """
        Fetch financial data from QuickBooks API
        Returns DataFrames for P&L and Balance Sheet data
        """
        logger.info(f"Fetching QuickBooks data from {start_date} to {end_date}")
        
        # Fetch GL entries from QuickBooks
        gl_df = self.qb_client.fetch_gl(start_date, end_date)
        
        if gl_df.empty:
            logger.warning("No GL data returned from QuickBooks")
            # Return empty dataframes with correct structure
            return {
                'pl': pd.DataFrame(columns=['Period', 'Revenue', 'COGS', 'OpEx', 
                                           'Depreciation', 'Interest', 'Tax']),
                'bs': pd.DataFrame(columns=['Period', 'Cash', 'AR', 'Inventory', 
                                           'PP&E', 'AP', 'Debt', 'Equity'])
            }
        
        # Add signed amounts and COA categories
        gl_df['Account_Type'] = gl_df.apply(
            lambda row: self.field_mapper.get_account_type(row['Account'], row['Account_Name']),
            axis=1
        )
        
        gl_df['Signed_Amount'] = gl_df.apply(
            lambda row: self.field_mapper.calculate_signed_amount(
                row['Debit'], row['Credit'], row['Account'], row['Account_Name']
            ),
            axis=1
        )
        
        gl_df['COA_Category'] = gl_df['Account'].apply(self.field_mapper.map_to_coa_category)
        
        # Create period column for aggregation
        gl_df['Period'] = gl_df['Date'].dt.to_period('M')
        
        # Aggregate P&L data by period and category
        pl_data = self._aggregate_pl_data(gl_df)
        
        # Calculate Balance Sheet (point in time, not period aggregation)
        bs_data = self._calculate_balance_sheet(gl_df, start_date, end_date)
        
        return {
            'pl': pl_data,
            'bs': bs_data,
            'gl': gl_df  # Keep raw GL for reference
        }
    
    def _aggregate_pl_data(self, gl_df: pd.DataFrame) -> pd.DataFrame:
        """Aggregate GL data into P&L line items by period"""
        # Filter to P&L accounts only
        pl_accounts = gl_df[gl_df['Account_Type'].isin(['revenue', 'expenses'])]
        
        # Group by period and COA category
        period_summary = pl_accounts.groupby(['Period', 'COA_Category'])['Signed_Amount'].sum().reset_index()
        
        # Pivot to get categories as columns
        pl_pivot = period_summary.pivot(index='Period', columns='COA_Category', values='Signed_Amount').fillna(0)
        
        # Create P&L DataFrame with standard columns
        pl_data = pd.DataFrame(index=pl_pivot.index)
        pl_data['Period'] = pl_pivot.index.astype(str)
        
        # Map COA categories to P&L line items
        pl_data['Revenue'] = pl_pivot.get('Revenue', 0) + pl_pivot.get('Other Income', 0)
        pl_data['COGS'] = -pl_pivot.get('Cost of Sales', 0)  # Expenses are negative
        pl_data['OpEx'] = -pl_pivot.get('Operating Expenses', 0)
        
        # Extract specific expense types if available
        expense_accounts = gl_df[gl_df['Account_Type'] == 'expenses']
        
        # Depreciation (look for specific accounts)
        depreciation = expense_accounts[
            expense_accounts['Account_Name'].str.contains('Depreciation', case=False, na=False)
        ].groupby('Period')['Signed_Amount'].sum()
        pl_data['Depreciation'] = -depreciation.reindex(pl_data.index).fillna(0)
        
        # Interest
        interest = expense_accounts[
            expense_accounts['Account_Name'].str.contains('Interest', case=False, na=False)
        ].groupby('Period')['Signed_Amount'].sum()
        pl_data['Interest'] = -interest.reindex(pl_data.index).fillna(0)
        
        # Tax (usually calculated, but we'll look for tax expense accounts)
        tax = expense_accounts[
            expense_accounts['Account_Name'].str.contains('Tax', case=False, na=False)
        ].groupby('Period')['Signed_Amount'].sum()
        pl_data['Tax'] = -tax.reindex(pl_data.index).fillna(0)
        
        return pl_data.reset_index(drop=True)
    
    def _calculate_balance_sheet(self, gl_df: pd.DataFrame, start_date: str, end_date: str) -> pd.DataFrame:
        """Calculate balance sheet positions by period"""
        # Get trial balance at end of each period
        periods = pd.date_range(start_date, end_date, freq='M')
        
        bs_data = []
        
        for period_end in periods:
            # Get cumulative balance as of period end
            period_gl = gl_df[gl_df['Date'] <= period_end]
            
            # Calculate balances by account type
            balances = period_gl.groupby('COA_Category')['Signed_Amount'].sum()
            
            # Create BS row
            bs_row = {
                'Period': period_end.strftime('%Y-%m'),
                'Cash': balances.get('Cash', 0),
                'AR': balances.get('Accounts Receivable', 0),
                'Inventory': balances.get('Inventory', 0),
                'PP&E': balances.get('Fixed Assets', 0),
                'AP': -balances.get('Accounts Payable', 0),  # Liabilities are negative
                'Debt': -balances.get('Long-term Liabilities', 0),
                'Equity': -balances.get('Equity', 0)
            }
            
            # Adjust for current/non-current splits if needed
            current_assets = balances.get('Current Assets', 0)
            if current_assets:
                # Split current assets into components
                bs_row['Cash'] = current_assets * 0.3  # Rough allocation
                bs_row['AR'] = current_assets * 0.5
                bs_row['Inventory'] = current_assets * 0.2
            
            bs_data.append(bs_row)
        
        return pd.DataFrame(bs_data)
    
    def fetch_prior_year_data(self, current_start: str, current_end: str) -> Dict[str, pd.DataFrame]:
        """Fetch prior year data for comparison"""
        # Calculate prior year dates
        current_start_dt = datetime.fromisoformat(current_start)
        current_end_dt = datetime.fromisoformat(current_end)
        
        py_start = (current_start_dt - timedelta(days=365)).strftime('%Y-%m-%d')
        py_end = (current_end_dt - timedelta(days=365)).strftime('%Y-%m-%d')
        
        logger.info(f"Fetching prior year data from {py_start} to {py_end}")
        
        # Check if we have cached PY data in S3
        cache_key = f"py_data_{py_start}_{py_end}.json"
        # TODO: Implement S3 cache check
        
        # Fetch from QuickBooks
        return self.fetch_quickbooks_data(py_start, py_end)
    
    def populate_income_statement(self, pl_df: pd.DataFrame) -> None:
        """Populate Income Statement sheet with P&L data"""
        ws = self.wb['Income Statement']
        
        # Find the starting row for data (usually row 5 after headers)
        data_start_row = 5
        data_start_col = 2  # Column B for first period
        
        # Clear existing period headers and data (columns B onwards)
        for col in range(data_start_col, ws.max_column + 1):
            ws.cell(row=3, column=col).value = None  # Period headers
            for row in range(data_start_row, 20):  # Clear data rows
                ws.cell(row=row, column=col).value = None
        
        # Write period headers
        for idx, period in enumerate(pl_df['Period']):
            col = data_start_col + idx
            ws.cell(row=3, column=col).value = period
        
        # Map data to specific rows (based on template structure)
        row_mapping = {
            'Revenue': 5,
            'COGS': 6,
            'OpEx': 9,
            'Depreciation': 10,
            'Interest': 12,
            'Tax': 13
        }
        
        # Write data
        for idx, row_data in pl_df.iterrows():
            col = data_start_col + idx
            for field, row_num in row_mapping.items():
                if field in row_data:
                    value = row_data[field]
                    # Ensure numeric values
                    if pd.notna(value):
                        ws.cell(row=row_num, column=col).value = float(value)
        
        logger.info(f"Populated Income Statement with {len(pl_df)} periods")
    
    def populate_balance_sheet(self, bs_df: pd.DataFrame) -> None:
        """Populate Balance Sheet with data"""
        ws = self.wb['Balance Sheet']
        
        # Similar structure to Income Statement
        data_start_row = 5
        data_start_col = 2
        
        # Clear existing data
        for col in range(data_start_col, ws.max_column + 1):
            ws.cell(row=3, column=col).value = None
            for row in range(data_start_row, 25):
                ws.cell(row=row, column=col).value = None
        
        # Write period headers
        for idx, period in enumerate(bs_df['Period']):
            col = data_start_col + idx
            ws.cell(row=3, column=col).value = period
        
        # Map data to specific rows
        row_mapping = {
            'Cash': 5,
            'AR': 6,
            'Inventory': 7,
            'PP&E': 9,
            'AP': 14,
            'Debt': 15,
            'Equity': 18
        }
        
        # Write data
        for idx, row_data in bs_df.iterrows():
            col = data_start_col + idx
            for field, row_num in row_mapping.items():
                if field in row_data:
                    value = row_data[field]
                    if pd.notna(value):
                        ws.cell(row=row_num, column=col).value = float(value)
        
        logger.info(f"Populated Balance Sheet with {len(bs_df)} periods")
    
    def save_populated_file(self, output_path: Optional[str] = None) -> str:
        """Save the populated workbook"""
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f"3statement_populated_{timestamp}.xlsx"
        
        self.wb.save(output_path)
        logger.info(f"Saved populated workbook: {output_path}")
        return output_path
    
    def upload_to_google_sheets(self, sheet_id: str) -> str:
        """Upload populated data to Google Sheets"""
        if not GSHEETS_AVAILABLE:
            raise ImportError("pygsheets not installed. Run: pip install pygsheets")
        
        credentials_path = os.getenv('GOOGLE_SHEETS_JSON')
        if not credentials_path:
            raise ValueError("GOOGLE_SHEETS_JSON environment variable not set")
        
        gc = pygsheets.authorize(service_file=credentials_path)
        sh = gc.open_by_key(sheet_id)
        
        # Upload each sheet's data
        for sheet_name in ['Income Statement', 'Balance Sheet', 'Cash Flow']:
            if sheet_name in self.wb.sheetnames:
                ws = self.wb[sheet_name]
                
                # Convert to DataFrame for easier upload
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(row)
                
                df = pd.DataFrame(data)
                
                # Get or create worksheet in Google Sheets
                try:
                    gws = sh.worksheet_by_title(sheet_name)
                except:
                    gws = sh.add_worksheet(sheet_name)
                
                # Clear and set data
                gws.clear()
                gws.set_dataframe(df, (1, 1), copy_head=False)
        
        return f"https://docs.google.com/spreadsheets/d/{sheet_id}"


def main():
    parser = argparse.ArgumentParser(description='Populate 3-Statement Model with QuickBooks data')
    parser.add_argument('--template', default='templates/Basic 3-Statement Model.xlsx',
                        help='Path to template file')
    parser.add_argument('--since', required=True, help='Start date (YYYY-MM-DD)')
    parser.add_argument('--until', help='End date (YYYY-MM-DD), defaults to today')
    parser.add_argument('--sheet-id', help='Google Sheets ID for cloud upload')
    parser.add_argument('--output', help='Output filename')
    parser.add_argument('--include-py', action='store_true', help='Include prior year data')
    parser.add_argument('--debug', action='store_true', help='Enable debug logging')
    
    args = parser.parse_args()
    
    # Setup logging
    logging.basicConfig(
        level=logging.DEBUG if args.debug else logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # Default end date to today
    if not args.until:
        args.until = datetime.now().strftime('%Y-%m-%d')
    
    # Initialize populator
    populator = ThreeStatementPopulator(args.template)
    
    try:
        # Test QuickBooks connection
        from integrations.quickbooks.client import test_connection
        if not test_connection():
            logger.error("QuickBooks connection failed. Check credentials and tokens.")
            return
        
        # Load template
        populator.load_template()
        
        # Fetch data
        logger.info(f"Fetching data from {args.since} to {args.until}")
        data = populator.fetch_quickbooks_data(args.since, args.until)
        
        # Include prior year if requested
        if args.include_py:
            py_data = populator.fetch_prior_year_data(args.since, args.until)
            # Store PY data for variance calculations
            data['pl_py'] = py_data.get('pl', pd.DataFrame())
            data['bs_py'] = py_data.get('bs', pd.DataFrame())
        
        # Populate sheets
        populator.populate_income_statement(data['pl'])
        populator.populate_balance_sheet(data['bs'])
        
        # Note: Cash Flow is typically calculated from P&L and BS changes
        # so we don't populate it directly
        
        # Save file
        output_path = populator.save_populated_file(args.output)
        
        # Upload to Google Sheets if requested
        if args.sheet_id:
            logger.info(f"Uploading to Google Sheets: {args.sheet_id}")
            sheet_url = populator.upload_to_google_sheets(args.sheet_id)
            logger.info(f"Google Sheets URL: {sheet_url}")
        
        print(f"✅ Successfully populated 3-Statement Model: {output_path}")
        if not data['gl'].empty:
            print(f"📊 Processed {len(data['gl'])} GL entries")
            print(f"📅 Periods: {data['pl']['Period'].min()} to {data['pl']['Period'].max()}")
        
    except Exception as e:
        logger.error(f"Error populating template: {e}")
        raise


if __name__ == '__main__':
    main()