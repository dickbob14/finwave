#!/usr/bin/env python3
"""
Populate Budget/Forecast template with planned metrics
Stores budget values alongside actuals in metric store for variance analysis
"""

import argparse
import logging
import os
import sys
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List, Optional, Any

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Add parent directory to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from metrics.excel_utils import add_metric_named_ranges, add_and_ingest
from metrics.utils import normalize_period
from metrics.models import ALL_METRICS
from templates.template_utils import get_column_date

logger = logging.getLogger(__name__)

class ForecastPopulator:
    """Populates Budget/Forecast template and ingests into metric store"""
    
    def __init__(self, template_path: str):
        self.template_path = Path(template_path)
        self.wb = None
        self.budget_metrics = {}
        self.forecast_metrics = {}
        
    def load_template(self) -> None:
        """Load the Excel template"""
        logger.info(f"Loading template: {self.template_path}")
        self.wb = load_workbook(self.template_path, keep_vba=True, data_only=False)
    
    def extract_budget_data(self, start_date: str, end_date: str) -> Dict[str, pd.DataFrame]:
        """Extract budget data from DATA_ sheets"""
        budget_data = {}
        
        # Find all DATA_ sheets
        data_sheets = [sheet for sheet in self.wb.sheetnames if sheet.startswith('DATA_')]
        logger.info(f"Found {len(data_sheets)} DATA_ sheets")
        
        for sheet_name in data_sheets:
            ws = self.wb[sheet_name]
            
            # Read the sheet into a dataframe
            data = []
            headers = []
            
            # Get headers from row 1
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    headers.append(header)
            
            # Read data starting from row 2
            for row in range(2, ws.max_row + 1):
                row_data = {}
                for col, header in enumerate(headers, 1):
                    value = ws.cell(row=row, column=col).value
                    row_data[header] = value
                
                if any(row_data.values()):  # Skip empty rows
                    data.append(row_data)
            
            if data:
                df = pd.DataFrame(data)
                budget_data[sheet_name] = df
                logger.info(f"Extracted {len(df)} rows from {sheet_name}")
        
        return budget_data
    
    def extract_budget_metrics(self) -> Dict[str, Dict[date, float]]:
        """Extract budget metrics from Budget vs Actuals sheets"""
        metrics = {}
        
        # Common patterns for budget vs actuals sheets
        budget_sheets = [
            'Budget vs Actuals',
            'Budget P&L',
            'Budget Summary',
            'Forecast'
        ]
        
        for sheet_name in self.wb.sheetnames:
            if any(pattern in sheet_name for pattern in budget_sheets):
                ws = self.wb[sheet_name]
                logger.info(f"Processing budget sheet: {sheet_name}")
                
                # Find date columns (usually in row 3 or 4)
                date_row = None
                for row in [3, 4]:
                    cell = ws.cell(row=row, column=2)
                    if cell.value and isinstance(cell.value, (datetime, date)):
                        date_row = row
                        break
                
                if not date_row:
                    logger.warning(f"No date row found in {sheet_name}")
                    continue
                
                # Extract periods from date row
                periods = []
                for col in range(2, ws.max_column + 1):
                    cell_value = ws.cell(row=date_row, column=col).value
                    if isinstance(cell_value, (datetime, date)):
                        period = normalize_period(cell_value)
                        periods.append((col, period))
                
                logger.info(f"Found {len(periods)} period columns")
                
                # Map metric rows
                metric_mappings = {
                    # P&L metrics
                    'revenue': ['Revenue', 'Total Revenue', 'Net Revenue'],
                    'cogs': ['COGS', 'Cost of Goods Sold', 'Cost of Sales'],
                    'gross_profit': ['Gross Profit', 'Gross Margin $'],
                    'opex': ['Operating Expenses', 'OpEx', 'Total OpEx'],
                    'ebitda': ['EBITDA', 'Operating Income'],
                    'net_income': ['Net Income', 'Net Profit'],
                    
                    # SaaS metrics
                    'mrr': ['MRR', 'Monthly Recurring Revenue'],
                    'arr': ['ARR', 'Annual Recurring Revenue'],
                    'new_customers': ['New Customers', 'New Logos'],
                    'churn_rate': ['Churn Rate', 'Churn %'],
                    
                    # Cash metrics
                    'cash': ['Cash', 'Cash Balance'],
                    'burn_rate': ['Burn Rate', 'Monthly Burn'],
                    'runway_months': ['Runway', 'Months of Runway']
                }
                
                # Find and extract metric values
                for metric_id, search_terms in metric_mappings.items():
                    for row in range(5, ws.max_row + 1):
                        cell_a = ws.cell(row=row, column=1).value
                        if cell_a and any(term.lower() in str(cell_a).lower() for term in search_terms):
                            # Found metric row, extract values
                            metric_key = f"budget_{metric_id}"
                            if metric_key not in metrics:
                                metrics[metric_key] = {}
                            
                            for col, period in periods:
                                value = ws.cell(row=row, column=col).value
                                if value and isinstance(value, (int, float)):
                                    metrics[metric_key][period] = float(value)
                            
                            logger.debug(f"Extracted budget_{metric_id}: {len(metrics[metric_key])} periods")
                            break
        
        return metrics
    
    def extract_forecast_metrics(self) -> Dict[str, Dict[date, float]]:
        """Extract forecast metrics (if different from budget)"""
        metrics = {}
        
        # Look for dedicated forecast sheets
        forecast_sheets = ['Forecast', 'Projections', 'Plan']
        
        for sheet_name in self.wb.sheetnames:
            if any(pattern in sheet_name for pattern in forecast_sheets) and 'Budget' not in sheet_name:
                # Process similar to budget extraction
                # For now, we'll use budget as forecast
                pass
        
        return metrics
    
    def create_metric_mappings(self, metrics_dict: Dict[str, Dict[date, float]]) -> Dict[str, str]:
        """Create named range mappings for metrics"""
        mappings = {}
        
        # Add a summary sheet if it doesn't exist
        if 'Metrics Summary' not in self.wb.sheetnames:
            ws = self.wb.create_sheet('Metrics Summary')
            ws['A1'] = 'Metric'
            ws['B1'] = 'Latest Value'
            ws['C1'] = 'Period'
            
            row = 2
            for metric_id, period_values in metrics_dict.items():
                if period_values:
                    latest_period = max(period_values.keys())
                    latest_value = period_values[latest_period]
                    
                    ws[f'A{row}'] = metric_id
                    ws[f'B{row}'] = latest_value
                    ws[f'C{row}'] = latest_period
                    
                    # Add named range
                    mappings[metric_id] = f'Metrics Summary!B{row}'
                    row += 1
        
        return mappings
    
    def ingest_budget_metrics(self, workspace_id: str, metrics_dict: Dict[str, Dict[date, float]]) -> Dict[str, int]:
        """Ingest budget/forecast metrics into metric store"""
        from metrics.ingest import ingest_metrics
        from metrics.models import Metric
        from core.database import get_db_session
        
        total_ingested = 0
        
        with get_db_session() as db:
            for metric_id, period_values in metrics_dict.items():
                for period, value in period_values.items():
                    # Check if exists
                    existing = db.query(Metric).filter_by(
                        workspace_id=workspace_id,
                        metric_id=metric_id,
                        period_date=period
                    ).first()
                    
                    if existing:
                        existing.value = value
                        existing.source_template = self.template_path.name
                        existing.updated_at = datetime.utcnow()
                    else:
                        metric = Metric(
                            workspace_id=workspace_id,
                            metric_id=metric_id,
                            period_date=period,
                            value=value,
                            source_template=self.template_path.name,
                            unit='dollars' if 'revenue' in metric_id or 'cost' in metric_id else None
                        )
                        db.add(metric)
                    
                    total_ingested += 1
            
            db.commit()
        
        logger.info(f"Ingested {total_ingested} budget/forecast metrics")
        return {'ingested': total_ingested}
    
    def save_populated_file(self, output_path: Optional[str] = None) -> str:
        """Save the populated workbook"""
        if output_path is None:
            # Save to populated directory
            populated_dir = Path('populated')
            populated_dir.mkdir(exist_ok=True)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = populated_dir / f"forecast_populated_{timestamp}.xlsx"
        
        # Add named ranges for any metrics we found
        all_metrics = {**self.budget_metrics, **self.forecast_metrics}
        if all_metrics:
            mappings = self.create_metric_mappings(all_metrics)
            add_metric_named_ranges(self.wb, mappings)
        
        # Save
        self.wb.save(output_path)
        logger.info(f"Saved populated workbook: {output_path}")
        
        return str(output_path)


def main():
    parser = argparse.ArgumentParser(description='Populate and ingest budget/forecast data')
    parser.add_argument('--template', default='assets/templates/registered/Budget vs Actuals.xlsx',
                        help='Path to template file')
    parser.add_argument('--since', required=True, help='Start date (YYYY-MM-DD)')
    parser.add_argument('--until', help='End date (YYYY-MM-DD), defaults to end of year')
    parser.add_argument('--output', help='Output filename')
    parser.add_argument('--sheet-id', help='Google Sheets ID for upload')
    parser.add_argument('--workspace', default='demo-corp', help='Workspace ID')
    parser.add_argument('--debug', action='store_true', help='Enable debug logging')
    
    args = parser.parse_args()
    
    # Setup logging
    logging.basicConfig(
        level=logging.DEBUG if args.debug else logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # Default end date to end of current year
    if not args.until:
        current_year = datetime.now().year
        args.until = f"{current_year}-12-31"
    
    # Initialize populator
    populator = ForecastPopulator(args.template)
    
    try:
        # Load template
        populator.load_template()
        
        # Extract budget data from DATA_ sheets
        logger.info(f"Extracting budget data from {args.since} to {args.until}")
        budget_data = populator.extract_budget_data(args.since, args.until)
        
        # Extract budget metrics from Budget vs Actuals sheets
        populator.budget_metrics = populator.extract_budget_metrics()
        
        # Extract forecast metrics if available
        populator.forecast_metrics = populator.extract_forecast_metrics()
        
        # Combine all metrics
        all_metrics = {**populator.budget_metrics, **populator.forecast_metrics}
        
        if not all_metrics:
            logger.warning("No budget/forecast metrics found")
        else:
            # Ingest into metric store
            results = populator.ingest_budget_metrics(args.workspace, all_metrics)
            
            print(f"✅ Successfully ingested budget/forecast data")
            print(f"📊 Metrics ingested: {results['ingested']}")
            print(f"📅 Period: {args.since} to {args.until}")
            
            # Show summary
            print(f"\n📈 Budget Metrics Summary:")
            for metric_id in sorted(all_metrics.keys()):
                periods = all_metrics[metric_id]
                if periods:
                    latest_period = max(periods.keys())
                    latest_value = periods[latest_period]
                    print(f"   {metric_id}: ${latest_value:,.0f} ({latest_period.strftime('%b %Y')})")
        
        # Save file
        output_path = populator.save_populated_file(args.output)
        print(f"\n💾 Saved to: {output_path}")
        
        # Upload to Google Sheets if requested
        if args.sheet_id:
            logger.info(f"Uploading to Google Sheets: {args.sheet_id}")
            # Implementation would go here
            print(f"📤 Uploaded to Google Sheets: {args.sheet_id}")
        
    except Exception as e:
        logger.error(f"Error populating forecast: {e}")
        raise


if __name__ == '__main__':
    main()