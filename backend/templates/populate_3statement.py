#!/usr/bin/env python3
"""
Populate Basic 3-Statement Model template with QuickBooks data
Usage: python populate_3statement.py --since 2024-01-01 --until 2024-12-31 [--sheet-id SHEET_ID]
"""

import argparse
import json
import logging
import os
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Optional Google Sheets support
try:
    import pygsheets
    GSHEETS_AVAILABLE = True
except ImportError:
    GSHEETS_AVAILABLE = False

logger = logging.getLogger(__name__)

# Schema for 3-Statement Model data requirements
FINANCIAL_DATA_SCHEMA = {
    "Income Statement": {
        "columns": [
            {"name": "Period", "type": "date", "format": "YYYY-MM"},
            {"name": "Revenue", "type": "number"},
            {"name": "COGS", "type": "number"},
            {"name": "OpEx", "type": "number"},
            {"name": "Depreciation", "type": "number"},
            {"name": "Interest", "type": "number"},
            {"name": "Tax", "type": "number"}
        ]
    },
    "Balance Sheet": {
        "columns": [
            {"name": "Period", "type": "date", "format": "YYYY-MM"},
            {"name": "Cash", "type": "number"},
            {"name": "AR", "type": "number"},
            {"name": "Inventory", "type": "number"},
            {"name": "PP&E", "type": "number"},
            {"name": "AP", "type": "number"},
            {"name": "Debt", "type": "number"},
            {"name": "Equity", "type": "number"}
        ]
    }
}

class ThreeStatementPopulator:
    """Populates 3-statement financial model with QuickBooks data"""
    
    def __init__(self, template_path: str):
        self.template_path = Path(template_path)
        self.wb = None
        
    def load_template(self) -> None:
        """Load the Excel template preserving formulas and formatting"""
        logger.info(f"Loading template: {self.template_path}")
        self.wb = load_workbook(self.template_path, keep_vba=True, data_only=False)
        
    def fetch_quickbooks_data(self, start_date: str, end_date: str) -> Dict[str, pd.DataFrame]:
        """
        Fetch financial data from QuickBooks API
        Returns DataFrames for P&L and Balance Sheet data
        """
        # TODO: Replace with actual QuickBooks API calls
        # For now, using sample data structure
        
        # Generate monthly periods
        start = pd.to_datetime(start_date)
        end = pd.to_datetime(end_date)
        periods = pd.date_range(start, end, freq='M')
        
        # Sample P&L data
        pl_data = []
        for period in periods:
            pl_data.append({
                'Period': period.strftime('%Y-%m'),
                'Revenue': 1000000 + (period.month * 50000),
                'COGS': 400000 + (period.month * 20000),
                'OpEx': 300000 + (period.month * 10000),
                'Depreciation': 10000,
                'Interest': 5000,
                'Tax': 0  # Calculated by formula
            })
        
        # Sample Balance Sheet data
        bs_data = []
        cash_balance = 500000
        for period in periods:
            cash_balance += 100000  # Simple cash growth
            bs_data.append({
                'Period': period.strftime('%Y-%m'),
                'Cash': cash_balance,
                'AR': 200000 + (period.month * 10000),
                'Inventory': 150000,
                'PP&E': 1000000 - (period.month * 10000),  # Depreciation
                'AP': 100000,
                'Debt': 500000,
                'Equity': 1000000
            })
        
        return {
            'pl': pd.DataFrame(pl_data),
            'bs': pd.DataFrame(bs_data)
        }
    
    def populate_income_statement(self, pl_df: pd.DataFrame) -> None:
        """Populate Income Statement sheet with P&L data"""
        ws = self.wb['Income Statement']
        
        # Find the starting row for data (usually row 5 after headers)
        data_start_row = 5
        data_start_col = 2  # Column B for first period
        
        # Clear existing period headers and data (columns B onwards)
        for col in range(data_start_col, ws.max_column + 1):
            ws.cell(row=3, column=col).value = None  # Period headers
            for row in range(data_start_row, 20):  # Clear data rows
                ws.cell(row=row, column=col).value = None
        
        # Write period headers
        for idx, period in enumerate(pl_df['Period']):
            col = data_start_col + idx
            ws.cell(row=3, column=col).value = period
        
        # Map data to specific rows (based on template structure)
        row_mapping = {
            'Revenue': 5,
            'COGS': 6,
            'OpEx': 9,
            'Depreciation': 10,
            'Interest': 12
        }
        
        # Write data
        for idx, row_data in pl_df.iterrows():
            col = data_start_col + idx
            for field, row_num in row_mapping.items():
                if field in row_data:
                    ws.cell(row=row_num, column=col).value = row_data[field]
        
        logger.info(f"Populated Income Statement with {len(pl_df)} periods")
    
    def populate_balance_sheet(self, bs_df: pd.DataFrame) -> None:
        """Populate Balance Sheet with data"""
        ws = self.wb['Balance Sheet']
        
        # Similar structure to Income Statement
        data_start_row = 5
        data_start_col = 2
        
        # Clear existing data
        for col in range(data_start_col, ws.max_column + 1):
            ws.cell(row=3, column=col).value = None
            for row in range(data_start_row, 25):
                ws.cell(row=row, column=col).value = None
        
        # Write period headers
        for idx, period in enumerate(bs_df['Period']):
            col = data_start_col + idx
            ws.cell(row=3, column=col).value = period
        
        # Map data to specific rows
        row_mapping = {
            'Cash': 5,
            'AR': 6,
            'Inventory': 7,
            'PP&E': 9,
            'AP': 14,
            'Debt': 15,
            'Equity': 18
        }
        
        # Write data
        for idx, row_data in bs_df.iterrows():
            col = data_start_col + idx
            for field, row_num in row_mapping.items():
                if field in row_data:
                    ws.cell(row=row_num, column=col).value = row_data[field]
        
        logger.info(f"Populated Balance Sheet with {len(bs_df)} periods")
    
    def save_populated_file(self, output_path: Optional[str] = None) -> str:
        """Save the populated workbook"""
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f"3statement_populated_{timestamp}.xlsx"
        
        self.wb.save(output_path)
        logger.info(f"Saved populated workbook: {output_path}")
        return output_path
    
    def upload_to_google_sheets(self, sheet_id: str) -> str:
        """Upload populated data to Google Sheets"""
        if not GSHEETS_AVAILABLE:
            raise ImportError("pygsheets not installed. Run: pip install pygsheets")
        
        credentials_path = os.getenv('GOOGLE_SHEETS_JSON')
        if not credentials_path:
            raise ValueError("GOOGLE_SHEETS_JSON environment variable not set")
        
        gc = pygsheets.authorize(service_file=credentials_path)
        sh = gc.open_by_key(sheet_id)
        
        # Upload each sheet's data
        # Note: This is simplified - in production you'd handle formulas differently
        for sheet_name in ['Income Statement', 'Balance Sheet', 'Cash Flow']:
            if sheet_name in self.wb.sheetnames:
                ws = self.wb[sheet_name]
                
                # Convert to DataFrame for easier upload
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(row)
                
                df = pd.DataFrame(data)
                
                # Get or create worksheet in Google Sheets
                try:
                    gws = sh.worksheet_by_title(sheet_name)
                except:
                    gws = sh.add_worksheet(sheet_name)
                
                # Clear and set data
                gws.clear()
                gws.set_dataframe(df, (1, 1), copy_head=False)
        
        return f"https://docs.google.com/spreadsheets/d/{sheet_id}"


def main():
    parser = argparse.ArgumentParser(description='Populate 3-Statement Model with QuickBooks data')
    parser.add_argument('--template', default='templates/Basic 3-Statement Model.xlsx',
                        help='Path to template file')
    parser.add_argument('--since', required=True, help='Start date (YYYY-MM-DD)')
    parser.add_argument('--until', help='End date (YYYY-MM-DD), defaults to today')
    parser.add_argument('--sheet-id', help='Google Sheets ID for cloud upload')
    parser.add_argument('--output', help='Output filename')
    parser.add_argument('--debug', action='store_true', help='Enable debug logging')
    
    args = parser.parse_args()
    
    # Setup logging
    logging.basicConfig(
        level=logging.DEBUG if args.debug else logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # Default end date to today
    if not args.until:
        args.until = datetime.now().strftime('%Y-%m-%d')
    
    # Initialize populator
    populator = ThreeStatementPopulator(args.template)
    
    try:
        # Load template
        populator.load_template()
        
        # Fetch data
        logger.info(f"Fetching data from {args.since} to {args.until}")
        data = populator.fetch_quickbooks_data(args.since, args.until)
        
        # Populate sheets
        populator.populate_income_statement(data['pl'])
        populator.populate_balance_sheet(data['bs'])
        
        # Note: Cash Flow is typically calculated from P&L and BS changes
        # so we don't populate it directly
        
        # Save file
        output_path = populator.save_populated_file(args.output)
        
        # Upload to Google Sheets if requested
        if args.sheet_id:
            logger.info(f"Uploading to Google Sheets: {args.sheet_id}")
            sheet_url = populator.upload_to_google_sheets(args.sheet_id)
            logger.info(f"Google Sheets URL: {sheet_url}")
        
        print(f"✅ Successfully populated 3-Statement Model: {output_path}")
        
    except Exception as e:
        logger.error(f"Error populating template: {e}")
        raise


if __name__ == '__main__':
    main()