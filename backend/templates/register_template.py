#!/usr/bin/env python3
"""
Template Registration Tool
Validates and registers new Excel templates with the Template Manager
"""

import argparse
import json
import logging
from pathlib import Path
from typing import Dict, List, Any
import shutil

import openpyxl
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

class TemplateRegistrar:
    """Validates and registers financial templates"""
    
    def __init__(self):
        self.templates_dir = Path(__file__).parent / 'files'
        self.registry_file = Path(__file__).parent / 'template_registry.json'
        
    def validate_template(self, template_path: Path) -> Dict[str, Any]:
        """
        Validate template structure and identify DATA sheets
        Returns validation results and template metadata
        """
        results = {
            'valid': True,
            'errors': [],
            'warnings': [],
            'metadata': {},
            'data_sheets': [],
            'report_sheets': [],
            'settings_sheets': []
        }
        
        try:
            # Load workbook
            wb = load_workbook(template_path, data_only=False)
            
            # Identify sheet types
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                if sheet_name.startswith('DATA_'):
                    # Analyze data sheet structure
                    data_info = self._analyze_data_sheet(ws)
                    results['data_sheets'].append({
                        'name': sheet_name,
                        'columns': data_info['columns'],
                        'row_count': data_info['row_count']
                    })
                    
                elif sheet_name.startswith('REPORT_') or sheet_name.startswith('DASH_'):
                    results['report_sheets'].append(sheet_name)
                    
                elif sheet_name in ['SETTINGS', 'CONFIG', 'PARAMETERS']:
                    results['settings_sheets'].append(sheet_name)
                    
                else:
                    results['warnings'].append(f"Unknown sheet type: {sheet_name}")
            
            # Validation checks
            if not results['data_sheets']:
                results['errors'].append("No DATA_ sheets found")
                results['valid'] = False
            
            # Extract metadata
            results['metadata'] = {
                'filename': template_path.name,
                'sheet_count': len(wb.sheetnames),
                'has_formulas': self._check_for_formulas(wb),
                'has_named_ranges': bool(wb.defined_names),
                'file_size': template_path.stat().st_size
            }
            
            # Check for required elements
            if not results['report_sheets']:
                results['warnings'].append("No REPORT_ or DASH_ sheets found")
            
        except Exception as e:
            results['valid'] = False
            results['errors'].append(f"Error loading template: {str(e)}")
        
        return results
    
    def _analyze_data_sheet(self, worksheet) -> Dict[str, Any]:
        """Analyze a data sheet structure"""
        # Get column headers (assume row 1)
        columns = []
        for cell in worksheet[1]:
            if cell.value:
                columns.append({
                    'name': str(cell.value),
                    'column': cell.column_letter,
                    'data_type': self._infer_column_type(worksheet, cell.column)
                })
        
        # Count data rows
        row_count = worksheet.max_row - 1  # Exclude header
        
        return {
            'columns': columns,
            'row_count': row_count
        }
    
    def _infer_column_type(self, worksheet, column: int) -> str:
        """Infer data type for a column"""
        # Sample first 10 data rows
        sample_values = []
        for row in range(2, min(12, worksheet.max_row + 1)):
            value = worksheet.cell(row=row, column=column).value
            if value is not None:
                sample_values.append(value)
        
        if not sample_values:
            return 'empty'
        
        # Check types
        if all(isinstance(v, (int, float)) for v in sample_values):
            return 'number'
        elif all(isinstance(v, str) and v.replace('-', '').replace('/', '').isdigit() 
                for v in sample_values):
            return 'date'
        else:
            return 'string'
    
    def _check_for_formulas(self, workbook) -> bool:
        """Check if workbook contains formulas"""
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':  # Formula
                        return True
        return False
    
    def register_template(self, template_path: Path, template_name: str = None,
                         version: str = None) -> bool:
        """
        Register a validated template
        Copies to templates directory and updates registry
        """
        # Validate first
        validation = self.validate_template(template_path)
        
        if not validation['valid']:
            logger.error(f"Template validation failed: {validation['errors']}")
            return False
        
        # Generate template name if not provided
        if not template_name:
            base_name = template_path.stem.lower()
            template_name = base_name.replace(' ', '_').replace('-', '_')
        
        # Generate version if not provided
        if not version:
            from datetime import datetime
            version = datetime.now().strftime('%Y.%m')
        
        # Copy to templates directory
        dest_path = self.templates_dir / template_path.name
        self.templates_dir.mkdir(exist_ok=True)
        shutil.copy2(template_path, dest_path)
        
        # Update registry
        registry = self._load_registry()
        
        registry[template_name] = {
            'name': template_name,
            'filename': template_path.name,
            'version': version,
            'registered_date': datetime.now().isoformat(),
            'validation': validation,
            'data_sheets': validation['data_sheets'],
            'report_sheets': validation['report_sheets']
        }
        
        self._save_registry(registry)
        
        logger.info(f"Registered template: {template_name} v{version}")
        return True
    
    def _load_registry(self) -> Dict[str, Any]:
        """Load template registry"""
        if self.registry_file.exists():
            with open(self.registry_file, 'r') as f:
                return json.load(f)
        return {}
    
    def _save_registry(self, registry: Dict[str, Any]):
        """Save template registry"""
        with open(self.registry_file, 'w') as f:
            json.dump(registry, f, indent=2)
    
    def list_registered_templates(self) -> List[Dict[str, Any]]:
        """List all registered templates"""
        registry = self._load_registry()
        templates = []
        
        for name, info in registry.items():
            templates.append({
                'name': name,
                'filename': info['filename'],
                'version': info['version'],
                'registered': info['registered_date'],
                'data_sheets': len(info['data_sheets']),
                'report_sheets': len(info['report_sheets'])
            })
        
        return templates
    
    def generate_populator_prompt(self, template_name: str) -> str:
        """Generate Claude prompt for creating a populator script"""
        registry = self._load_registry()
        
        if template_name not in registry:
            raise ValueError(f"Template not found: {template_name}")
        
        info = registry[template_name]
        
        prompt = f"""Here is a financial workbook: {info['filename']}

Identify all sheets whose name starts with 'DATA_':
{json.dumps(info['data_sheets'], indent=2)}

Generate a Python 3.11 script that:
1. Loads QuickBooks sample data from CSVs or API
2. Overwrites those DATA tables from pandas DataFrames with the same columns
3. Does not touch formulas or formatting in other sheets
4. Saves as populate_{template_name}.py

Follow the project's Template Manager pattern used in populate_3statement.py:
- Use QuickBooksClient for data fetching
- Use FieldMapper for column mapping
- Include Google Sheets upload support
- Add proper logging and error handling

The script should have this structure:
```python
class {info['filename'].replace(' ', '').replace('-', '')}Populator:
    def load_template(self)
    def fetch_quickbooks_data(self, start_date, end_date)
    def populate_{sheet['name'].lower()}(self, df) for each DATA sheet
    def save_populated_file(self, output_path)
    def upload_to_google_sheets(self, sheet_id)
```

Return the complete populator script ready to run."""
        
        return prompt


def main():
    parser = argparse.ArgumentParser(description='Register financial templates')
    parser.add_argument('command', choices=['validate', 'register', 'list', 'prompt'],
                        help='Command to execute')
    parser.add_argument('--file', help='Template file path')
    parser.add_argument('--name', help='Template name for registration')
    parser.add_argument('--version', help='Template version')
    parser.add_argument('--debug', action='store_true', help='Enable debug logging')
    
    args = parser.parse_args()
    
    # Setup logging
    logging.basicConfig(
        level=logging.DEBUG if args.debug else logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    registrar = TemplateRegistrar()
    
    if args.command == 'validate':
        if not args.file:
            print("Error: --file required for validate command")
            return
        
        template_path = Path(args.file)
        if not template_path.exists():
            print(f"Error: File not found: {args.file}")
            return
        
        results = registrar.validate_template(template_path)
        
        print(f"\n📋 Template Validation: {template_path.name}")
        print("=" * 60)
        print(f"Valid: {'✅' if results['valid'] else '❌'}")
        
        if results['errors']:
            print("\n❌ Errors:")
            for error in results['errors']:
                print(f"   - {error}")
        
        if results['warnings']:
            print("\n⚠️  Warnings:")
            for warning in results['warnings']:
                print(f"   - {warning}")
        
        print(f"\n📊 Data Sheets: {len(results['data_sheets'])}")
        for sheet in results['data_sheets']:
            print(f"   - {sheet['name']}: {len(sheet['columns'])} columns")
        
        print(f"\n📈 Report Sheets: {len(results['report_sheets'])}")
        for sheet in results['report_sheets']:
            print(f"   - {sheet}")
    
    elif args.command == 'register':
        if not args.file:
            print("Error: --file required for register command")
            return
        
        template_path = Path(args.file)
        success = registrar.register_template(template_path, args.name, args.version)
        
        if success:
            print(f"✅ Successfully registered template!")
            if args.name:
                print(f"📋 Name: {args.name}")
        else:
            print("❌ Registration failed. Check logs for details.")
    
    elif args.command == 'list':
        templates = registrar.list_registered_templates()
        
        print("\n📚 Registered Templates:")
        print("=" * 80)
        for tmpl in templates:
            print(f"\n{tmpl['name']} (v{tmpl['version']})")
            print(f"  File: {tmpl['filename']}")
            print(f"  Data sheets: {tmpl['data_sheets']}")
            print(f"  Report sheets: {tmpl['report_sheets']}")
            print(f"  Registered: {tmpl['registered']}")
    
    elif args.command == 'prompt':
        if not args.name:
            print("Error: --name required for prompt command")
            return
        
        try:
            prompt = registrar.generate_populator_prompt(args.name)
            print(prompt)
        except ValueError as e:
            print(f"Error: {e}")


if __name__ == '__main__':
    main()