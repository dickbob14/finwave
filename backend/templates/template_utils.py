"""
Template utility functions for FinWave board pack generation
"""
import os
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import pandas as pd

logger = logging.getLogger(__name__)

# Check for optional dependencies
try:
    import xlwings as xw
    XLWINGS_AVAILABLE = True
except ImportError:
    XLWINGS_AVAILABLE = False
    logger.warning("xlwings not available - formula recalculation disabled")

try:
    from weasyprint import HTML
    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False
    logger.warning("WeasyPrint not available - PDF generation disabled")

def get_template_path() -> Path:
    """Get the path to templates directory relative to repo root"""
    return Path(__file__).parent.parent / "assets" / "templates"

def get_month_columns(start_date: str, end_date: str) -> List[Tuple[str, datetime]]:
    """
    Generate list of month columns between start and end dates
    Returns list of (display_name, date) tuples
    """
    start = pd.to_datetime(start_date)
    end = pd.to_datetime(end_date)
    
    months = []
    current = start.replace(day=1)  # Start of month
    
    while current <= end:
        display_name = current.strftime("%b %Y")
        months.append((display_name, current))
        current = current + relativedelta(months=1)
    
    return months

def build_coa_mapping(coa_df: pd.DataFrame) -> Tuple[Dict[str, str], Dict[str, str]]:
    """
    Build account to group/subgroup mappings from COA data
    Returns: (acct_to_group, acct_to_subgroup) dictionaries
    """
    acct_to_group = {}
    acct_to_subgroup = {}
    
    for _, row in coa_df.iterrows():
        account_code = str(row.get('Account_Code', ''))
        group = str(row.get('Group', ''))
        subgroup = str(row.get('SubGroup', ''))
        
        if account_code:
            acct_to_group[account_code] = group
            acct_to_subgroup[account_code] = subgroup
    
    return acct_to_group, acct_to_subgroup

def calculate_signed_amount(debit: float, credit: float, account_type: str) -> float:
    """
    Calculate signed amount based on account type and debit/credit
    For GL display:
    - Assets: Debit positive, Credit negative
    - Expenses: Debit negative (to show as outflow)
    - Liabilities/Equity/Income: Credit positive, Debit negative
    """
    debit = float(debit or 0)
    credit = float(credit or 0)
    
    if account_type.lower() == 'asset':
        return debit - credit
    elif account_type.lower() == 'expense':
        # Expenses shown as negative (outflows)
        return -(debit - credit)
    else:  # liability, equity, income
        return credit - debit

def format_excel_date(date_str: str) -> str:
    """Convert date string to Excel-friendly ISO format"""
    try:
        dt = pd.to_datetime(date_str)
        return dt.strftime('%Y-%m-%d')
    except:
        return date_str

def get_prior_year_period(start_date: str, end_date: str) -> Tuple[str, str]:
    """Get the prior year period dates"""
    start = pd.to_datetime(start_date)
    end = pd.to_datetime(end_date)
    
    py_start = start - relativedelta(years=1)
    py_end = end - relativedelta(years=1)
    
    return py_start.strftime('%Y-%m-%d'), py_end.strftime('%Y-%m-%d')

def recalculate_workbook(wb_path: str) -> Dict[str, Any]:
    """
    Recalculate Excel workbook using xlwings if available
    Returns dict of key values for validation
    """
    results = {}
    
    if not XLWINGS_AVAILABLE:
        logger.warning("xlwings not available - skipping recalculation")
        return results
    
    try:
        # Open workbook invisibly
        app = xw.App(visible=False)
        wb = app.books.open(wb_path)
        
        # Force recalculation
        app.calculate()
        
        # Extract key values for validation
        if 'REPORT_P&L' in [s.name for s in wb.sheets]:
            pl_sheet = wb.sheets['REPORT_P&L']
            # Assuming revenue is in row 9
            results['total_revenue'] = pl_sheet.range('C9').value
            results['total_expenses'] = pl_sheet.range('C20').value  # Adjust row as needed
            results['net_income'] = pl_sheet.range('C25').value  # Adjust row as needed
        
        wb.save()
        wb.close()
        app.quit()
        
        logger.info(f"Workbook recalculated: {results}")
        
    except Exception as e:
        logger.error(f"Failed to recalculate workbook: {e}")
    
    return results

def validate_gl_totals(gl_df: pd.DataFrame, calculated_values: Dict[str, Any]) -> bool:
    """
    Validate that GL totals match calculated values
    """
    # Group by account type
    revenue_total = gl_df[gl_df['Account_Type'] == 'Income']['Net_Amount'].sum()
    expense_total = abs(gl_df[gl_df['Account_Type'] == 'Expense']['Net_Amount'].sum())
    
    # Compare with calculated values
    calc_revenue = calculated_values.get('total_revenue', 0) or 0
    calc_expenses = calculated_values.get('total_expenses', 0) or 0
    
    revenue_match = abs(revenue_total - calc_revenue) < 0.01
    expense_match = abs(expense_total - calc_expenses) < 0.01
    
    if not revenue_match:
        logger.error(f"Revenue mismatch: GL={revenue_total}, Calc={calc_revenue}")
    if not expense_match:
        logger.error(f"Expense mismatch: GL={expense_total}, Calc={calc_expenses}")
    
    return revenue_match and expense_match

def create_icon_set_rule(ws, range_address: str, reverse: bool = False):
    """
    Create 3-icon set rule for percentage cells
    Falls back to data bars if icon sets not supported
    """
    from openpyxl.formatting.rule import IconSetRule, DataBarRule
    
    try:
        # Try to create icon set rule
        icon_set = IconSetRule(
            icon_style='3Arrows',  # or '3TrafficLights1'
            type='percent',
            values=[0, 33, 67],
            reverse=reverse
        )
        ws.conditional_formatting.add(range_address, icon_set)
        logger.info(f"Added icon set rule to {range_address}")
        
    except Exception as e:
        logger.warning(f"Icon set failed, using data bars: {e}")
        # Fallback to data bars
        data_bar = DataBarRule(
            start_type='num',
            start_value=0,
            end_type='num', 
            end_value=100,
            color="638EC6"  # Blue color
        )
        ws.conditional_formatting.add(range_address, data_bar)

def copy_to_google_sheets(excel_path: str, sheet_name: str) -> Optional[str]:
    """
    Copy Excel workbook to Google Sheets if credentials available
    Returns the Google Sheet URL or None
    """
    try:
        import pygsheets
        from openpyxl import load_workbook
        
        # Check for credentials
        creds_path = os.getenv('GOOGLE_SHEETS_SERVICE_ACCOUNT_JSON')
        if not creds_path or not os.path.exists(creds_path):
            logger.info("Google Sheets credentials not found - skipping upload")
            return None
        
        # Authorize
        gc = pygsheets.authorize(service_file=creds_path)
        
        # Create new spreadsheet
        sh = gc.create(sheet_name)
        logger.info(f"Created Google Sheet: {sh.title}")
        
        # Load Excel workbook
        wb = load_workbook(excel_path, data_only=True)
        
        # Copy each worksheet
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            
            # Get all values
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(list(row))
            
            # Create or get worksheet
            if ws_name == wb.sheetnames[0]:
                # First sheet already exists
                wks = sh.sheet1
                wks.title = ws_name
            else:
                wks = sh.add_worksheet(ws_name)
            
            # Update values
            if data:
                wks.update_values('A1', data)
            
            logger.info(f"Copied worksheet: {ws_name}")
        
        # Return URL
        return sh.url
        
    except ImportError:
        logger.warning("pygsheets not installed - Google Sheets upload disabled")
        return None
    except Exception as e:
        logger.error(f"Failed to upload to Google Sheets: {e}")
        return None