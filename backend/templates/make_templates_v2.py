#!/usr/bin/env python3
"""
FinWave Board-Pack Template Generator (Enhanced Version)
Generates professional Excel/Google Sheets financial reporting templates
"""

import os
import json
import datetime
from typing import Dict, List, Tuple, Optional
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Border, Side, Alignment, 
    NamedStyle, Protection
)
from openpyxl.formatting.rule import (
    CellIsRule, FormulaRule, IconSetRule, IconSet, Rule, DataBarRule
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.comments import Comment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.drawing.colors import ColorChoice
import logging
from pathlib import Path

# Import utilities
from template_utils import (
    get_template_path, get_month_columns, create_icon_set_rule,
    copy_to_google_sheets, WEASYPRINT_AVAILABLE
)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Color Palette
COLORS = {
    'primary': '002B49',      # Midnight Blue
    'accent1': '00A6A6',      # Teal
    'accent2': 'FF9F1C',      # Mango
    'neutral_light': 'F5F6F8',
    'neutral_mid': 'CBD2D9',
    'white': 'FFFFFF',
    'black': '000000',
    'green': '00B050',
    'red': 'FF0000',
    'amber': 'FFC000'
}

# Style definitions
def create_styles():
    """Create named styles for consistent formatting"""
    styles = {}
    
    # Title style
    title = NamedStyle(name='title')
    title.font = Font(name='Calibri', size=18, bold=True, color=COLORS['primary'])
    title.alignment = Alignment(horizontal='left', vertical='center')
    styles['title'] = title
    
    # Section Header
    section_header = NamedStyle(name='section_header')
    section_header.font = Font(name='Calibri', size=14, bold=True, color=COLORS['white'])
    section_header.fill = PatternFill("solid", fgColor=COLORS['primary'])
    section_header.alignment = Alignment(horizontal='center', vertical='center')
    section_header.border = Border(
        left=Side(style='thin', color=COLORS['primary']),
        right=Side(style='thin', color=COLORS['primary']),
        top=Side(style='thin', color=COLORS['primary']),
        bottom=Side(style='thin', color=COLORS['primary'])
    )
    styles['section_header'] = section_header
    
    # Column Header
    col_header = NamedStyle(name='col_header')
    col_header.font = Font(name='Calibri', size=11, bold=True, color=COLORS['primary'])
    col_header.fill = PatternFill("solid", fgColor=COLORS['neutral_light'])
    col_header.alignment = Alignment(horizontal='center', vertical='center')
    col_header.border = Border(bottom=Side(style='medium', color=COLORS['primary']))
    styles['col_header'] = col_header
    
    # Data cells
    data_cell = NamedStyle(name='data_cell')
    data_cell.font = Font(name='Calibri', size=11)
    data_cell.alignment = Alignment(horizontal='right', vertical='center')
    data_cell.border = Border(
        left=Side(style='thin', color=COLORS['neutral_mid']),
        right=Side(style='thin', color=COLORS['neutral_mid'])
    )
    styles['data_cell'] = data_cell
    
    # Total row
    total_row = NamedStyle(name='total_row')
    total_row.font = Font(name='Calibri', size=11, bold=True)
    total_row.fill = PatternFill("solid", fgColor=COLORS['neutral_light'])
    total_row.border = Border(
        top=Side(style='double', color=COLORS['primary']),
        bottom=Side(style='thin', color=COLORS['primary'])
    )
    styles['total_row'] = total_row
    
    return styles

class FinWaveTemplateBuilder:
    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        
        # Add styles
        styles = create_styles()
        for style in styles.values():
            self.wb.add_named_style(style)
    
    def create_data_gl(self):
        """Create DATA_GL sheet - General Ledger data with signed amounts"""
        ws = self.wb.create_sheet("DATA_GL")
        ws.sheet_properties.tabColor = COLORS['neutral_mid']
        
        # Headers for tblGL
        headers = [
            'Date', 'Account', 'Account_Name', 'Amount', 'Description',
            'Reference', 'Customer', 'Vendor', 'Class', 'Location',
            'Memo', 'TxnID', 'Group', 'SubGroup'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.style = 'col_header'
            
        # Column widths
        col_widths = {
            'A': 12, 'B': 12, 'C': 30, 'D': 15, 'E': 40,
            'F': 15, 'G': 20, 'H': 20, 'I': 15, 'J': 15,
            'K': 30, 'L': 20, 'M': 15, 'N': 15
        }
        
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width
        
        # Create table
        table = Table(displayName="tblGL", ref="A1:N10000")
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)
        
        return ws
    
    def create_data_gl_py(self):
        """Create DATA_GL_PY sheet - Prior Year General Ledger data"""
        ws = self.wb.create_sheet("DATA_GL_PY")
        ws.sheet_properties.tabColor = COLORS['neutral_mid']
        
        # Same structure as DATA_GL
        headers = [
            'Date', 'Account', 'Account_Name', 'Amount', 'Description',
            'Reference', 'Customer', 'Vendor', 'Class', 'Location',
            'Memo', 'TxnID', 'Group', 'SubGroup'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.style = 'col_header'
            
        # Column widths
        col_widths = {
            'A': 12, 'B': 12, 'C': 30, 'D': 15, 'E': 40,
            'F': 15, 'G': 20, 'H': 20, 'I': 15, 'J': 15,
            'K': 30, 'L': 20, 'M': 15, 'N': 15
        }
        
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width
        
        # Create table
        table = Table(displayName="tblGL_PY", ref="A1:N10000")
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)
        
        return ws
    
    def create_data_coa(self):
        """Create DATA_COA sheet - Chart of Accounts with mapping"""
        ws = self.wb.create_sheet("DATA_COA")
        ws.sheet_properties.tabColor = COLORS['neutral_mid']
        
        # Headers
        headers = ['Account_Code', 'Account_Name', 'Account_Type', 'Group', 'SubGroup', 'Normal_Balance']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.style = 'col_header'
            
        # Sample COA data with proper grouping
        coa_data = [
            # Assets
            ('1000', 'Cash', 'Asset', 'Current Assets', 'Cash & Equivalents', 'Debit'),
            ('1100', 'Petty Cash', 'Asset', 'Current Assets', 'Cash & Equivalents', 'Debit'),
            ('1200', 'Accounts Receivable', 'Asset', 'Current Assets', 'Receivables', 'Debit'),
            ('1300', 'Inventory', 'Asset', 'Current Assets', 'Inventory', 'Debit'),
            ('1400', 'Prepaid Expenses', 'Asset', 'Current Assets', 'Other Current Assets', 'Debit'),
            ('1500', 'Property, Plant & Equipment', 'Asset', 'Fixed Assets', 'PPE', 'Debit'),
            ('1600', 'Accumulated Depreciation', 'Asset', 'Fixed Assets', 'PPE', 'Credit'),
            
            # Liabilities
            ('2000', 'Accounts Payable', 'Liability', 'Current Liabilities', 'Payables', 'Credit'),
            ('2100', 'Accrued Expenses', 'Liability', 'Current Liabilities', 'Accruals', 'Credit'),
            ('2200', 'Short-term Debt', 'Liability', 'Current Liabilities', 'Debt', 'Credit'),
            ('2500', 'Long-term Debt', 'Liability', 'Long-term Liabilities', 'Debt', 'Credit'),
            
            # Equity
            ('3000', 'Common Stock', 'Equity', 'Equity', 'Contributed Capital', 'Credit'),
            ('3100', 'Retained Earnings', 'Equity', 'Equity', 'Retained Earnings', 'Credit'),
            ('3200', 'Current Year Earnings', 'Equity', 'Equity', 'Current Earnings', 'Credit'),
            
            # Revenue
            ('4000', 'Product Revenue', 'Income', 'Revenue', 'Product Revenue', 'Credit'),
            ('4100', 'Service Revenue', 'Income', 'Revenue', 'Service Revenue', 'Credit'),
            ('4200', 'Other Revenue', 'Income', 'Revenue', 'Other Revenue', 'Credit'),
            
            # COGS
            ('5000', 'Cost of Goods Sold', 'Expense', 'Cost of Goods Sold', 'Direct Costs', 'Debit'),
            ('5100', 'Direct Labor', 'Expense', 'Cost of Goods Sold', 'Direct Labor', 'Debit'),
            ('5200', 'Materials', 'Expense', 'Cost of Goods Sold', 'Materials', 'Debit'),
            
            # Operating Expenses
            ('6000', 'General & Administrative', 'Expense', 'Operating Expenses', 'General & Administrative', 'Debit'),
            ('6100', 'Travel & Entertainment', 'Expense', 'Operating Expenses', 'Travel & Entertainment', 'Debit'),
            ('6200', 'Office Expenses', 'Expense', 'Operating Expenses', 'Office Expenses', 'Debit'),
            ('6300', 'Sales & Marketing', 'Expense', 'Operating Expenses', 'Sales & Marketing', 'Debit'),
            ('6400', 'Compensation & Benefits', 'Expense', 'Operating Expenses', 'Compensation & Benefits', 'Debit'),
            ('6500', 'Professional Fees', 'Expense', 'Operating Expenses', 'Professional Services', 'Debit'),
            
            # Other
            ('7000', 'Interest Income', 'Income', 'Other Income', 'Interest', 'Credit'),
            ('7100', 'Interest Expense', 'Expense', 'Other Expenses', 'Interest', 'Debit'),
            ('8000', 'Income Tax Expense', 'Expense', 'Tax Expense', 'Income Tax', 'Debit'),
        ]
        
        # Populate data
        for row_idx, data in enumerate(coa_data, 2):
            for col_idx, value in enumerate(data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
                
        # Create table
        table = Table(displayName="tblCOA", ref=f"A1:F{len(coa_data)+1}")
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)
        
        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 15
        
        return ws
    
    def create_data_map(self):
        """Create DATA_MAP sheet - Account mapping rules"""
        ws = self.wb.create_sheet("DATA_MAP")
        ws.sheet_properties.tabColor = COLORS['neutral_mid']
        
        # Headers
        headers = ['Line_Item', 'Group', 'SubGroup', 'Account_Pattern', 'Sign_Convention']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.style = 'col_header'
            
        # Mapping data for financial statements
        mapping_data = [
            # P&L mappings
            ('Product Revenue', 'Revenue', 'Product Revenue', '4000', '+'),
            ('Service Revenue', 'Revenue', 'Service Revenue', '4100', '+'),
            ('Other Revenue', 'Revenue', 'Other Revenue', '4200', '+'),
            ('COGS', 'Cost of Goods Sold', 'All', '5*', '-'),
            ('G&A Expenses', 'Operating Expenses', 'General & Administrative', '6000', '-'),
            ('T&E Expenses', 'Operating Expenses', 'Travel & Entertainment', '6100', '-'),
            ('Office Expenses', 'Operating Expenses', 'Office Expenses', '6200', '-'),
            ('S&M Expenses', 'Operating Expenses', 'Sales & Marketing', '6300', '-'),
            ('Compensation', 'Operating Expenses', 'Compensation & Benefits', '6400', '-'),
            ('Professional Fees', 'Operating Expenses', 'Professional Services', '6500', '-'),
            ('Interest Income', 'Other Income', 'Interest', '7000', '+'),
            ('Interest Expense', 'Other Expenses', 'Interest', '7100', '-'),
            ('Income Tax', 'Tax Expense', 'Income Tax', '8000', '-'),
            
            # Balance Sheet mappings
            ('Cash', 'Current Assets', 'Cash & Equivalents', '10*,11*', '+'),
            ('Accounts Receivable', 'Current Assets', 'Receivables', '1200', '+'),
            ('Inventory', 'Current Assets', 'Inventory', '1300', '+'),
            ('Prepaid Expenses', 'Current Assets', 'Other Current Assets', '1400', '+'),
            ('Fixed Assets', 'Fixed Assets', 'PPE', '1500', '+'),
            ('Accumulated Depreciation', 'Fixed Assets', 'PPE', '1600', '-'),
            ('Accounts Payable', 'Current Liabilities', 'Payables', '2000', '+'),
            ('Accrued Expenses', 'Current Liabilities', 'Accruals', '2100', '+'),
            ('Short-term Debt', 'Current Liabilities', 'Debt', '2200', '+'),
            ('Long-term Debt', 'Long-term Liabilities', 'Debt', '2500', '+'),
            ('Common Stock', 'Equity', 'Contributed Capital', '3000', '+'),
            ('Retained Earnings', 'Equity', 'Retained Earnings', '3100', '+'),
        ]
        
        # Populate data
        for row_idx, data in enumerate(mapping_data, 2):
            for col_idx, value in enumerate(data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
                
        # Create table
        table = Table(displayName="tblMap", ref=f"A1:E{len(mapping_data)+1}")
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        
        return ws
    
    def create_report_pl(self):
        """Create REPORT_P&L sheet - Income Statement with dynamic months"""
        ws = self.wb.create_sheet("REPORT_P&L")
        
        # Title
        ws['A1'] = '=SETTINGS!B2 & " · Income Statement"'
        ws['A1'].style = 'title'
        ws.merge_cells('A1:N1')
        
        # Generate dynamic month columns
        # This will be populated based on fxStart and fxEnd from SETTINGS
        ws['A3'] = 'Account'
        ws['B3'] = 'Description'
        ws['A3'].style = 'col_header'
        ws['B3'].style = 'col_header'
        
        # Month headers - dynamic formula based on SETTINGS
        month_formulas = []
        for i in range(12):  # Up to 12 months
            col_letter = get_column_letter(i + 3)
            month_formula = f'=IF(EOMONTH(fxStart,{i})<=fxEnd,TEXT(EOMONTH(fxStart,{i}),"MMM YYYY"),"")'
            ws[f'{col_letter}3'] = month_formula
            ws[f'{col_letter}3'].style = 'col_header'
            month_formulas.append((col_letter, i))
        
        # P&L structure with COA-based formulas
        pl_structure = [
            ('', '', 4),
            ('REVENUE', '', 5),
            ('4000', '  Product Revenue', 6),
            ('4100', '  Service Revenue', 7),
            ('4200', '  Other Revenue', 8),
            ('', 'Total Revenue', 9),
            ('', '', 10),
            ('COST OF GOODS SOLD', '', 11),
            ('5000', '  Direct Costs', 12),
            ('5100', '  Direct Labor', 13),
            ('5200', '  Materials', 14),
            ('', 'Total COGS', 15),
            ('', '', 16),
            ('', 'GROSS PROFIT', 17),
            ('', 'Gross Margin %', 18),
            ('', '', 19),
            ('OPERATING EXPENSES', '', 20),
            ('6000', '  General & Administrative', 21),
            ('6100', '  Travel & Entertainment', 22),
            ('6200', '  Office Expenses', 23),
            ('6300', '  Sales & Marketing', 24),
            ('6400', '  Compensation & Benefits', 25),
            ('6500', '  Professional Fees', 26),
            ('', 'Total Operating Expenses', 27),
            ('', '', 28),
            ('', 'OPERATING INCOME', 29),
            ('', 'Operating Margin %', 30),
            ('', '', 31),
            ('OTHER INCOME/EXPENSES', '', 32),
            ('7000', '  Interest Income', 33),
            ('7100', '  Interest Expense', 34),
            ('', 'Total Other', 35),
            ('', '', 36),
            ('', 'INCOME BEFORE TAX', 37),
            ('8000', '  Income Tax Expense', 38),
            ('', 'NET INCOME', 39),
            ('', 'Net Margin %', 40),
        ]
        
        # Populate structure with formulas
        for account, desc, row in pl_structure:
            ws.cell(row=row, column=1, value=account)
            ws.cell(row=row, column=2, value=desc)
            
            # Apply formulas to month columns
            if account and row not in [5, 11, 20, 32]:  # Skip section headers
                # Use SUMIFS with tblGL to sum by Group/SubGroup for the month
                for col_letter, month_idx in month_formulas:
                    # Build date range for the month
                    month_start = f'EOMONTH(fxStart,{month_idx-1})+1'
                    month_end = f'EOMONTH(fxStart,{month_idx})'
                    
                    # SUMIFS formula using Group from DATA_MAP
                    formula = (f'=SUMIFS(tblGL[Amount],'
                              f'tblGL[Account],"{account}*",'
                              f'tblGL[Date],">="&{month_start},'
                              f'tblGL[Date],"<="&{month_end})')
                    
                    cell = ws.cell(row=row, column=ord(col_letter) - ord('A') + 1)
                    cell.value = formula
                    cell.number_format = '#,##0'
            
            # Total rows
            elif desc == 'Total Revenue':
                for col_letter, _ in month_formulas:
                    col = ord(col_letter) - ord('A') + 1
                    formula = f'=SUM({col_letter}6:{col_letter}8)'
                    ws.cell(row=row, column=col, value=formula).number_format = '#,##0'
            elif desc == 'Total COGS':
                for col_letter, _ in month_formulas:
                    col = ord(col_letter) - ord('A') + 1
                    formula = f'=SUM({col_letter}12:{col_letter}14)'
                    ws.cell(row=row, column=col, value=formula).number_format = '#,##0'
            elif desc == 'GROSS PROFIT':
                for col_letter, _ in month_formulas:
                    col = ord(col_letter) - ord('A') + 1
                    formula = f'={col_letter}9-{col_letter}15'
                    ws.cell(row=row, column=col, value=formula).number_format = '#,##0'
            elif desc == 'Gross Margin %':
                for col_letter, _ in month_formulas:
                    col = ord(col_letter) - ord('A') + 1
                    formula = f'=IF({col_letter}9=0,0,{col_letter}17/{col_letter}9)'
                    ws.cell(row=row, column=col, value=formula).number_format = '0.0%'
            elif desc == 'Total Operating Expenses':
                for col_letter, _ in month_formulas:
                    col = ord(col_letter) - ord('A') + 1
                    formula = f'=SUM({col_letter}21:{col_letter}26)'
                    ws.cell(row=row, column=col, value=formula).number_format = '#,##0'
            elif desc == 'OPERATING INCOME':
                for col_letter, _ in month_formulas:
                    col = ord(col_letter) - ord('A') + 1
                    formula = f'={col_letter}17-{col_letter}27'
                    ws.cell(row=row, column=col, value=formula).number_format = '#,##0'
            elif desc == 'Operating Margin %':
                for col_letter, _ in month_formulas:
                    col = ord(col_letter) - ord('A') + 1
                    formula = f'=IF({col_letter}9=0,0,{col_letter}29/{col_letter}9)'
                    ws.cell(row=row, column=col, value=formula).number_format = '0.0%'
            elif desc == 'Total Other':
                for col_letter, _ in month_formulas:
                    col = ord(col_letter) - ord('A') + 1
                    formula = f'={col_letter}33-{col_letter}34'
                    ws.cell(row=row, column=col, value=formula).number_format = '#,##0'
            elif desc == 'INCOME BEFORE TAX':
                for col_letter, _ in month_formulas:
                    col = ord(col_letter) - ord('A') + 1
                    formula = f'={col_letter}29+{col_letter}35'
                    ws.cell(row=row, column=col, value=formula).number_format = '#,##0'
            elif desc == 'NET INCOME':
                for col_letter, _ in month_formulas:
                    col = ord(col_letter) - ord('A') + 1
                    formula = f'={col_letter}37-{col_letter}38'
                    ws.cell(row=row, column=col, value=formula).number_format = '#,##0'
            elif desc == 'Net Margin %':
                for col_letter, _ in month_formulas:
                    col = ord(col_letter) - ord('A') + 1
                    formula = f'=IF({col_letter}9=0,0,{col_letter}39/{col_letter}9)'
                    ws.cell(row=row, column=col, value=formula).number_format = '0.0%'
                    
            # Apply styles
            if desc in ['REVENUE', 'COST OF GOODS SOLD', 'OPERATING EXPENSES', 'OTHER INCOME/EXPENSES']:
                ws.cell(row=row, column=1).style = 'section_header'
                ws.merge_cells(f'A{row}:B{row}')
            elif desc in ['Total Revenue', 'Total COGS', 'GROSS PROFIT', 'Total Operating Expenses', 
                          'OPERATING INCOME', 'NET INCOME', 'INCOME BEFORE TAX']:
                ws.cell(row=row, column=1).style = 'total_row'
                ws.cell(row=row, column=2).style = 'total_row'
                for col_letter, _ in month_formulas:
                    col = ord(col_letter) - ord('A') + 1
                    ws.cell(row=row, column=col).style = 'total_row'
                    
        # Column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 35
        for col in range(3, 15):
            ws.column_dimensions[get_column_letter(col)].width = 12
            
        # Conditional formatting for margin % rows with icon sets
        for row in [18, 30, 40]:  # Margin % rows
            range_str = f'C{row}:N{row}'
            create_icon_set_rule(ws, range_str, reverse=False)
            
        # Define named range
        defn = DefinedName('rngIS_Matrix', attr_text="'REPORT_P&L'!$A$5:$N$40")
        self.wb.defined_names['rngIS_Matrix'] = defn
        
        return ws
    
    def create_report_bs(self):
        """Create REPORT_BS sheet - Balance Sheet with prior year comparison"""
        ws = self.wb.create_sheet("REPORT_BS")
        
        # Title
        ws['A1'] = '=SETTINGS!B2 & " · Balance Sheet"'
        ws['A1'].style = 'title'
        ws.merge_cells('A1:F1')
        
        # Headers
        ws['A3'] = 'Account'
        ws['B3'] = 'Description'
        ws['C3'] = 'Current Period'
        ws['D3'] = 'Prior Period'
        ws['E3'] = '$ Change'
        ws['F3'] = '% Change'
        
        for col in range(1, 7):
            ws.cell(row=3, column=col).style = 'col_header'
            
        # Balance Sheet structure with COA-based formulas
        bs_structure = [
            ('', '', 4),
            ('ASSETS', '', 5),
            ('Current Assets', '', 6),
            ('1000', '  Cash & Cash Equivalents', 7),
            ('1200', '  Accounts Receivable', 8),
            ('1300', '  Inventory', 9),
            ('1400', '  Prepaid Expenses', 10),
            ('', 'Total Current Assets', 11),
            ('', '', 12),
            ('Fixed Assets', '', 13),
            ('1500', '  Property, Plant & Equipment', 14),
            ('1600', '  Less: Accumulated Depreciation', 15),
            ('', 'Net Fixed Assets', 16),
            ('', '', 17),
            ('1700', 'Other Assets', 18),
            ('', '', 19),
            ('', 'TOTAL ASSETS', 20),
            ('', '', 21),
            ('LIABILITIES & EQUITY', '', 22),
            ('Current Liabilities', '', 23),
            ('2000', '  Accounts Payable', 24),
            ('2100', '  Accrued Expenses', 25),
            ('2200', '  Short-term Debt', 26),
            ('', 'Total Current Liabilities', 27),
            ('', '', 28),
            ('2500', 'Long-term Liabilities', 29),
            ('', '', 30),
            ('', 'TOTAL LIABILITIES', 31),
            ('', '', 32),
            ('Equity', '', 33),
            ('3000', '  Common Stock', 34),
            ('3100', '  Retained Earnings', 35),
            ('', '  Current Year Earnings', 36),
            ('', 'TOTAL EQUITY', 37),
            ('', '', 38),
            ('', 'TOTAL LIABILITIES & EQUITY', 39),
        ]
        
        # Populate structure
        for account, desc, row in bs_structure:
            ws.cell(row=row, column=1, value=account)
            ws.cell(row=row, column=2, value=desc)
            
            # Current period formulas
            if account and row not in [5, 6, 13, 22, 23, 33]:
                # Use SUMIFS with account code
                formula = f'=SUMIFS(tblGL[Amount],tblGL[Account],"{account}*")'
                ws.cell(row=row, column=3, value=formula).number_format = '#,##0'
                
                # Prior period formula (using tblGL_PY)
                formula_py = f'=SUMIFS(tblGL_PY[Amount],tblGL_PY[Account],"{account}*")'
                ws.cell(row=row, column=4, value=formula_py).number_format = '#,##0'
            
            # Calculate totals
            elif desc == 'Total Current Assets':
                ws.cell(row=row, column=3, value='=SUM(C7:C10)').number_format = '#,##0'
                ws.cell(row=row, column=4, value='=SUM(D7:D10)').number_format = '#,##0'
            elif desc == 'Net Fixed Assets':
                ws.cell(row=row, column=3, value='=C14+C15').number_format = '#,##0'
                ws.cell(row=row, column=4, value='=D14+D15').number_format = '#,##0'
            elif desc == 'TOTAL ASSETS':
                ws.cell(row=row, column=3, value='=C11+C16+C18').number_format = '#,##0'
                ws.cell(row=row, column=4, value='=D11+D16+D18').number_format = '#,##0'
            elif desc == 'Total Current Liabilities':
                ws.cell(row=row, column=3, value='=SUM(C24:C26)').number_format = '#,##0'
                ws.cell(row=row, column=4, value='=SUM(D24:D26)').number_format = '#,##0'
            elif desc == 'TOTAL LIABILITIES':
                ws.cell(row=row, column=3, value='=C27+C29').number_format = '#,##0'
                ws.cell(row=row, column=4, value='=D27+D29').number_format = '#,##0'
            elif desc == '  Current Year Earnings':
                # Link to P&L Net Income
                ws.cell(row=row, column=3, value="='REPORT_P&L'!N39").number_format = '#,##0'
                ws.cell(row=row, column=4, value="0").number_format = '#,##0'
            elif desc == 'TOTAL EQUITY':
                ws.cell(row=row, column=3, value='=SUM(C34:C36)').number_format = '#,##0'
                ws.cell(row=row, column=4, value='=SUM(D34:D36)').number_format = '#,##0'
            elif desc == 'TOTAL LIABILITIES & EQUITY':
                ws.cell(row=row, column=3, value='=C31+C37').number_format = '#,##0'
                ws.cell(row=row, column=4, value='=D31+D37').number_format = '#,##0'
                
            # Variance columns
            if row >= 7:
                # $ Change
                ws.cell(row=row, column=5, value=f'=C{row}-D{row}').number_format = '#,##0'
                # % Change
                ws.cell(row=row, column=6, value=f'=IF(D{row}=0,0,(C{row}-D{row})/ABS(D{row}))').number_format = '0.0%'
                
            # Apply styles
            if desc in ['ASSETS', 'LIABILITIES & EQUITY', 'Current Assets', 'Fixed Assets', 
                        'Current Liabilities', 'Equity']:
                ws.cell(row=row, column=2).style = 'section_header'
            elif desc in ['Total Current Assets', 'Net Fixed Assets', 'TOTAL ASSETS', 
                          'Total Current Liabilities', 'TOTAL LIABILITIES', 'TOTAL EQUITY', 
                          'TOTAL LIABILITIES & EQUITY']:
                for col in range(2, 7):
                    ws.cell(row=row, column=col).style = 'total_row'
                    
        # Column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 35
        for col in ['C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 15
            
        # Conditional formatting for variance
        # Green for positive, red for negative
        from openpyxl.formatting.rule import CellIsRule
        
        # $ Change column
        green_fill = PatternFill(start_color=COLORS['green'], end_color=COLORS['green'], fill_type='solid')
        red_fill = PatternFill(start_color=COLORS['red'], end_color=COLORS['red'], fill_type='solid')
        
        ws.conditional_formatting.add('E7:E39',
            CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill))
        ws.conditional_formatting.add('E7:E39',
            CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))
            
        # Define named range
        defn = DefinedName('rngBS_Matrix', attr_text="'REPORT_BS'!$A$5:$F$39")
        self.wb.defined_names['rngBS_Matrix'] = defn
        
        return ws
    
    def create_dash_kpi(self):
        """Create DASH_KPI sheet - Executive KPI Dashboard with rich formatting"""
        ws = self.wb.create_sheet("DASH_KPI")
        
        # Title
        ws['A1'] = '=SETTINGS!B2 & " · Key Performance Indicators"'
        ws['A1'].style = 'title'
        ws.merge_cells('A1:L1')
        
        # KPI Grid Layout (4x3)
        kpis = [
            # Row 1
            ('Revenue', "='REPORT_P&L'!N9", '=TEXT((N9-M9)/M9,"0.0%") & " MoM"', 'B3:D6', 'green'),
            ('Gross Margin', "='REPORT_P&L'!N18", '=TEXT(N18-M18,"0.0pp") & " MoM"', 'F3:H6', 'amber'),
            ('Operating Margin', "='REPORT_P&L'!N30", '=TEXT(N30-M30,"0.0pp") & " MoM"', 'J3:L6', 'amber'),
            # Row 2
            ('Cash Balance', "='REPORT_BS'!C7", '=TEXT(E7,"$#,##0") & " vs PY"', 'B8:D11', 'green'),
            ('AR Days', '=ROUND(REPORT_BS!C8/(REPORT_P&L!N9/30),0)', '=TEXT(C8-D8,"0") & " days"', 'F8:H11', 'red'),
            ('Working Capital', '=REPORT_BS!C11-REPORT_BS!C27', '=TEXT((C11-C27)-(D11-D27),"$#,##0")', 'J8:L11', 'green'),
            # Row 3
            ('Revenue/Employee', '=REPORT_P&L!N9/SETTINGS!B7', '=TEXT((N9/B7)-(M9/B7),"$#,##0")', 'B13:D16', 'green'),
            ('Burn Rate', '=-(REPORT_P&L!N39-REPORT_P&L!M39)', '=TEXT(-((N39-M39)-(M39-L39)),"$#,##0")', 'F13:H16', 'amber'),
            ('Runway (months)', '=ROUND(REPORT_BS!C7/F13,0)', '=TEXT(C7/F13-C7/E13,"0.0") & " mo"', 'J13:L16', 'green'),
            # Row 4
            ('Gross Profit', "='REPORT_P&L'!N17", '=TEXT((N17-M17)/M17,"0.0%") & " MoM"', 'B18:D21', 'green'),
            ('EBITDA', "='REPORT_P&L'!N29", '=TEXT((N29-M29),"$#,##0")', 'F18:H21', 'amber'),
            ('Quick Ratio', '=(REPORT_BS!C7+REPORT_BS!C8)/REPORT_BS!C27', '=TEXT(C7/C27-D7/D27,"0.00x")', 'J18:L21', 'green'),
        ]
        
        # Create KPI tiles with full formatting
        for i, (label, value_formula, change_formula, range_str, status) in enumerate(kpis):
            # Parse range
            start_cell = range_str.split(':')[0]
            end_cell = range_str.split(':')[1]
            start_row = int(start_cell[1:])
            start_col = ord(start_cell[0]) - ord('A') + 1
            
            # Merge cells for tile
            ws.merge_cells(range_str)
            
            # Create rich text content in the tile
            # We'll use multiple cells within the merged range for different elements
            
            # Label (top of tile)
            label_cell = ws.cell(row=start_row, column=start_col)
            label_cell.value = label
            label_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            label_cell.font = Font(name='Calibri', size=12, bold=True, color=COLORS['white'])
            
            # Value (middle of tile - one row down)
            value_cell = ws.cell(row=start_row + 1, column=start_col)
            value_cell.value = value_formula
            value_cell.alignment = Alignment(horizontal='center', vertical='center')
            value_cell.font = Font(name='Calibri', size=16, bold=True, color=COLORS['white'])
            
            # Change (bottom of tile - two rows down)
            change_cell = ws.cell(row=start_row + 2, column=start_col)
            change_cell.value = change_formula
            change_cell.alignment = Alignment(horizontal='right', vertical='bottom')
            change_cell.font = Font(name='Calibri', size=10, color=COLORS['neutral_light'])
            
            # Apply background color based on status
            tile_color = COLORS['green'] if status == 'green' else (COLORS['amber'] if status == 'amber' else COLORS['red'])
            fill = PatternFill("solid", fgColor=tile_color)
            
            # Apply fill to all cells in the merged range
            for row in range(start_row, start_row + 4):
                for col in range(start_col, start_col + 3):
                    ws.cell(row=row, column=col).fill = fill
            
            # Add border
            border = Border(
                left=Side(style='medium', color=COLORS['primary']),
                right=Side(style='medium', color=COLORS['primary']),
                top=Side(style='medium', color=COLORS['primary']),
                bottom=Side(style='medium', color=COLORS['primary'])
            )
            label_cell.border = border
            
        # Row heights
        for row in [3, 8, 13, 18]:
            ws.row_dimensions[row].height = 20
            ws.row_dimensions[row + 1].height = 25
            ws.row_dimensions[row + 2].height = 15
            ws.row_dimensions[row + 3].height = 10
            
        # Column widths
        for col in ['B', 'C', 'D', 'F', 'G', 'H', 'J', 'K', 'L']:
            ws.column_dimensions[col].width = 12
            
        return ws
    
    def create_settings(self):
        """Create SETTINGS sheet - Configuration"""
        ws = self.wb.create_sheet("SETTINGS")
        
        # Title
        ws['A1'] = 'FinWave Settings'
        ws['A1'].style = 'title'
        ws.merge_cells('A1:C1')
        
        # Settings with defaults
        settings = [
            ('Company Name', 'Acme Corp', 'B2'),
            ('Fiscal Year Start', '2025-01-01', 'B3'),
            ('Fiscal Year End', '2025-12-31', 'B4'),
            ('Base Currency', 'USD', 'B5'),
            ('Report Date', '=TODAY()', 'B6'),
            ('Employee Count', '50', 'B7'),
            ('', '', ''),
            ('Theme Settings', '', 'A9'),
            ('Primary Color', COLORS['primary'], 'B10'),
            ('Accent Color 1', COLORS['accent1'], 'B11'),
            ('Accent Color 2', COLORS['accent2'], 'B12'),
        ]
        
        for i, (label, value, _) in enumerate(settings, 2):
            if label:
                ws.cell(row=i, column=1, value=label)
                ws.cell(row=i, column=1).font = Font(bold=True)
            if value:
                ws.cell(row=i, column=2, value=value)
                
        # Define named ranges
        defn1 = DefinedName('fxStart', attr_text="'SETTINGS'!$B$3")
        self.wb.defined_names['fxStart'] = defn1
        
        defn2 = DefinedName('fxEnd', attr_text="'SETTINGS'!$B$4")
        self.wb.defined_names['fxEnd'] = defn2
        
        defn3 = DefinedName('CompanyName', attr_text="'SETTINGS'!$B$2")
        self.wb.defined_names['CompanyName'] = defn3
        
        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        
        return ws
    
    def create_readme(self):
        """Create README sheet - Documentation"""
        ws = self.wb.create_sheet("README")
        
        # Title
        ws['A1'] = 'FinWave Board Pack Template'
        ws['A1'].style = 'title'
        ws.merge_cells('A1:D1')
        
        # Documentation content
        content = [
            ('', '', 2),
            ('Overview', 'Professional financial reporting template with automated calculations', 3),
            ('Version', '2.0 - Enhanced with dynamic periods and COA mapping', 4),
            ('', '', 5),
            ('Sheet Guide', '', 6),
            ('DATA_GL', 'General ledger transactions (current period)', 7),
            ('DATA_GL_PY', 'General ledger transactions (prior year)', 8),
            ('DATA_COA', 'Chart of accounts with grouping', 9),
            ('DATA_MAP', 'Account mapping rules for reports', 10),
            ('REPORT_P&L', 'Income statement with monthly columns', 11),
            ('REPORT_BS', 'Balance sheet with prior year comparison', 12),
            ('DASH_KPI', 'Executive KPI dashboard', 13),
            ('SETTINGS', 'Configuration and parameters', 14),
            ('', '', 15),
            ('Instructions', '', 16),
            ('1. Data Import', 'Run ETL script to populate DATA_GL sheets', 17),
            ('2. Period Setup', 'Update fiscal year dates in SETTINGS', 18),
            ('3. Refresh', 'Press Ctrl+Alt+F9 to recalculate all formulas', 19),
            ('4. Export', 'Use File > Export to create PDF board pack', 20),
        ]
        
        for label, desc, row in content:
            if label:
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=1).font = Font(bold=True)
            if desc:
                ws.cell(row=row, column=2, value=desc)
                ws.merge_cells(f'B{row}:D{row}')
                
        # Section headers
        for row in [6, 16]:
            ws.cell(row=row, column=1).style = 'section_header'
            
        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 60
        
        return ws
    
    def build_template(self):
        """Build the complete template"""
        logger.info("Building FinWave Board Pack Template...")
        
        # Create sheets in order
        self.create_data_gl()
        self.create_data_gl_py()
        self.create_data_coa()
        self.create_data_map()
        self.create_report_pl()
        self.create_report_bs()
        self.create_dash_kpi()
        self.create_settings()
        self.create_readme()
        
        return self.wb
    
    def save_template(self, filename: str = None):
        """Save the template to file"""
        if not filename:
            output_dir = get_template_path()
            output_dir.mkdir(parents=True, exist_ok=True)
            filename = output_dir / "finwave_board_pack.xlsx"
        
        self.wb.save(filename)
        logger.info(f"Template saved to {filename}")
        return filename
    
    def self_test(self):
        """Run self-test to validate template"""
        logger.info("Running self-test...")
        
        # Check all expected sheets exist
        expected_sheets = [
            'DATA_GL', 'DATA_GL_PY', 'DATA_COA', 'DATA_MAP',
            'REPORT_P&L', 'REPORT_BS', 'DASH_KPI', 'SETTINGS', 'README'
        ]
        
        for sheet in expected_sheets:
            if sheet not in self.wb.sheetnames:
                raise ValueError(f"Missing sheet: {sheet}")
                
        # Check named ranges
        expected_ranges = ['fxStart', 'fxEnd', 'CompanyName', 'rngIS_Matrix', 'rngBS_Matrix']
        for name in expected_ranges:
            if name not in self.wb.defined_names:
                raise ValueError(f"Missing named range: {name}")
                
        logger.info("Self-test passed!")
        return True

def main():
    """Main entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(description="Generate FinWave Board Pack Excel template")
    parser.add_argument("--output", help="Output file path")
    parser.add_argument("--google-sheets", action="store_true", help="Also create Google Sheets version")
    
    args = parser.parse_args()
    
    try:
        # Build template
        builder = FinWaveTemplateBuilder()
        builder.build_template()
        
        # Save Excel file
        excel_path = builder.save_template(args.output)
        
        # Run self-test
        builder.self_test()
        
        print(f"✅ FinWave Board Pack Template created successfully!")
        print(f"📁 Excel file: {excel_path}")
        
        # Google Sheets upload if requested
        if args.google_sheets:
            sheet_url = copy_to_google_sheets(str(excel_path), "FinWave Board Pack")
            if sheet_url:
                print(f"📊 Google Sheet: {sheet_url}")
            else:
                print("⚠️  Google Sheets upload skipped (no credentials)")
        
        print("\n🚀 Next steps:")
        print("   1. Run ETL to populate DATA_GL with QuickBooks data")
        print("   2. Open in Excel to verify formulas calculate correctly")
        print("   3. Export to PDF using the report generator")
        
    except Exception as e:
        print(f"❌ Template generation failed: {e}")
        raise

if __name__ == "__main__":
    main()