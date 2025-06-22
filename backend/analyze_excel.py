#!/usr/bin/env python3

import sys
import json
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. Installing...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "--user", "openpyxl"])
    import openpyxl
    from openpyxl.utils import get_column_letter

def analyze_excel_file(file_path):
    """Analyze an Excel file and return its structure"""
    # Load in non-read-only mode to get full access to properties
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    analysis = {
        "file_name": Path(file_path).name,
        "sheets": {},
        "named_ranges": [],
        "tables": []
    }
    
    # Get named ranges
    try:
        if wb.defined_names:
            for name_def in wb.defined_names:
                if hasattr(name_def, 'name'):
                    analysis["named_ranges"].append({
                        "name": name_def.name,
                        "scope": name_def.localSheetId if hasattr(name_def, 'localSheetId') else None,
                        "formula": str(name_def.value) if hasattr(name_def, 'value') else str(name_def)
                    })
    except Exception as e:
        print(f"Warning: Could not read named ranges: {e}")
    
    # Analyze each sheet
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_info = {
            "name": sheet_name,
            "dimensions": sheet.calculate_dimensions() if hasattr(sheet, 'calculate_dimensions') else "Unknown",
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "is_data_sheet": sheet_name.startswith("DATA_") or "data" in sheet_name.lower(),
            "columns": [],
            "sample_data": []
        }
        
        # Get column headers (check first few rows for headers)
        if sheet.max_row > 0:
            headers = []
            header_row = 1
            
            # Try to find the header row (first row with multiple non-empty cells)
            for row_idx in range(1, min(10, sheet.max_row + 1)):
                non_empty_count = 0
                for col in range(1, min(sheet.max_column + 1, 20)):
                    if sheet.cell(row=row_idx, column=col).value:
                        non_empty_count += 1
                if non_empty_count > 3:  # Found likely header row
                    header_row = row_idx
                    break
            
            for col in range(1, min(sheet.max_column + 1, 50)):  # Limit to 50 columns
                cell = sheet.cell(row=header_row, column=col)
                if cell.value:
                    headers.append({
                        "column": get_column_letter(col),
                        "name": str(cell.value),
                        "index": col
                    })
            sheet_info["columns"] = headers
            sheet_info["header_row"] = header_row
            
            # Get sample data (first 5 rows after header)
            for row_idx in range(header_row + 1, min(header_row + 6, sheet.max_row + 1)):
                row_data = {}
                for col_info in headers:
                    cell = sheet.cell(row=row_idx, column=col_info["index"])
                    if cell.value is not None:
                        value = cell.value
                        # Determine type
                        value_type = type(value).__name__
                        if hasattr(cell, 'is_date') and cell.is_date:
                            value_type = "datetime"
                        row_data[col_info["name"]] = {
                            "value": str(value),
                            "type": value_type
                        }
                if row_data:
                    sheet_info["sample_data"].append(row_data)
        
        # Check for tables in the sheet
        if hasattr(sheet, 'tables'):
            for table in sheet.tables.values():
                analysis["tables"].append({
                    "name": table.name,
                    "sheet": sheet_name,
                    "ref": table.ref,
                    "displayName": table.displayName if hasattr(table, 'displayName') else table.name
                })
        
        analysis["sheets"][sheet_name] = sheet_info
    
    wb.close()
    return analysis

def generate_json_schema(sheet_info):
    """Generate JSON schema for a data sheet"""
    if not sheet_info["columns"] or not sheet_info["sample_data"]:
        return None
    
    schema = {
        "type": "object",
        "properties": {},
        "required": []
    }
    
    # Analyze data types from sample data
    column_types = {}
    for col in sheet_info["columns"]:
        col_name = col["name"]
        types_found = set()
        
        for row in sheet_info["sample_data"]:
            if col_name in row:
                types_found.add(row[col_name]["type"])
        
        # Determine the most appropriate type
        if "float" in types_found or "int" in types_found:
            json_type = "number"
        elif "datetime" in types_found:
            json_type = "string"
            format_type = "date-time"
        elif "bool" in types_found:
            json_type = "boolean"
        else:
            json_type = "string"
        
        prop_def = {"type": json_type}
        if json_type == "string" and "datetime" in types_found:
            prop_def["format"] = "date-time"
        
        schema["properties"][col_name] = prop_def
        schema["required"].append(col_name)
    
    return schema

def main():
    files = [
        "/Users/alexandermillar/Downloads/Basic 3-Statement Model-2.xlsx",
        "/Users/alexandermillar/Downloads/Cube - KPI Dashboard-1.xlsx"
    ]
    
    for file_path in files:
        print(f"\n{'='*60}")
        print(f"Analyzing: {Path(file_path).name}")
        print('='*60)
        
        try:
            analysis = analyze_excel_file(file_path)
            
            # Print sheet summary
            print("\nSheets found:")
            for sheet_name, sheet_info in analysis["sheets"].items():
                print(f"\n  {sheet_name}:")
                print(f"    - Dimensions: {sheet_info['dimensions']}")
                print(f"    - Rows: {sheet_info['max_row']}, Columns: {sheet_info['max_column']}")
                print(f"    - Is Data Sheet: {sheet_info['is_data_sheet']}")
                
                if sheet_info["columns"]:
                    print(f"    - Column Headers: {', '.join([c['name'] for c in sheet_info['columns'][:10]])}")
                    if len(sheet_info["columns"]) > 10:
                        print(f"      ... and {len(sheet_info['columns']) - 10} more columns")
            
            # Print named ranges
            if analysis["named_ranges"]:
                print(f"\nNamed Ranges: {len(analysis['named_ranges'])}")
                for nr in analysis["named_ranges"][:10]:
                    print(f"  - {nr['name']}: {nr['formula']}")
                if len(analysis['named_ranges']) > 10:
                    print(f"  ... and {len(analysis['named_ranges']) - 10} more")
            
            # Print tables
            if analysis["tables"]:
                print(f"\nTables: {len(analysis['tables'])}")
                for table in analysis["tables"]:
                    print(f"  - {table['name']} in {table['sheet']}: {table['ref']}")
            
            # Generate JSON schemas for data sheets
            print("\nJSON Schemas for Data Sheets:")
            data_sheets_found = False
            for sheet_name, sheet_info in analysis["sheets"].items():
                if sheet_info["is_data_sheet"] or (sheet_info["columns"] and len(sheet_info["columns"]) > 3):
                    data_sheets_found = True
                    schema = generate_json_schema(sheet_info)
                    if schema:
                        print(f"\n{sheet_name} Schema:")
                        print(json.dumps(schema, indent=2))
                        
                        # Show sample data
                        if sheet_info["sample_data"]:
                            print(f"\nSample data from {sheet_name}:")
                            for i, row in enumerate(sheet_info["sample_data"][:2]):
                                print(f"  Row {i+1}:")
                                for col, data in list(row.items())[:5]:
                                    print(f"    {col}: {data['value']} ({data['type']})")
            
            if not data_sheets_found:
                print("\nNo obvious data sheets found. Showing all sheets with data:")
                for sheet_name, sheet_info in analysis["sheets"].items():
                    if sheet_info["columns"]:
                        print(f"\n{sheet_name}:")
                        print(f"  Columns: {', '.join([c['name'] for c in sheet_info['columns'][:10]])}")
                        if len(sheet_info["columns"]) > 10:
                            print(f"  ... and {len(sheet_info['columns']) - 10} more")
            
            # Save full analysis
            output_file = f"{Path(file_path).stem}_analysis.json"
            with open(output_file, 'w') as f:
                json.dump(analysis, f, indent=2)
            print(f"\nFull analysis saved to: {output_file}")
            
        except Exception as e:
            print(f"Error analyzing {file_path}: {str(e)}")
            import traceback
            traceback.print_exc()

if __name__ == "__main__":
    main()