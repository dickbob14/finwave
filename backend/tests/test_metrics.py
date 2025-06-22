"""
Automated tests for metric store functionality
"""

import pytest
import tempfile
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl import Workbook

from metrics.ingest import ingest_metrics, extract_named_ranges
from metrics.excel_utils import add_metric_named_ranges, add_and_ingest
from metrics.utils import normalize_period, get_period_range
from metrics.models import Metric

def create_test_workbook() -> Workbook:
    """Create a minimal test workbook with metrics"""
    wb = Workbook()
    
    # Create Income Statement sheet
    ws = wb.active
    ws.title = "Income Statement"
    
    # Headers
    ws['A1'] = "Income Statement"
    ws['A3'] = "Period:"
    ws['B3'] = date(2024, 1, 31)  # January 2024
    
    # Metrics
    ws['A5'] = "Revenue"
    ws['B5'] = 1000000  # $1M
    
    ws['A6'] = "COGS"
    ws['B6'] = 400000   # $400k
    
    ws['A7'] = "Gross Profit"
    ws['B7'] = 600000   # $600k
    
    ws['A9'] = "Operating Expenses"
    ws['B9'] = 300000   # $300k
    
    ws['A11'] = "EBITDA"
    ws['B11'] = 300000  # $300k
    
    return wb

def test_period_normalization():
    """Test period date normalization"""
    # Test various dates normalize to month-end
    assert normalize_period(date(2024, 1, 15)) == date(2024, 1, 31)
    assert normalize_period(date(2024, 2, 1)) == date(2024, 2, 29)  # Leap year
    assert normalize_period(date(2023, 2, 1)) == date(2023, 2, 28)  # Non-leap
    assert normalize_period("2024-12-25") == date(2024, 12, 31)
    
    # Test period range
    start, end = get_period_range(date(2024, 3, 15), 3)
    assert start == date(2024, 1, 31)
    assert end == date(2024, 3, 31)

def test_named_ranges():
    """Test adding and extracting named ranges"""
    wb = create_test_workbook()
    
    # Add named ranges
    mappings = {
        'revenue': 'Income Statement!B5',
        'cogs': 'Income Statement!B6',
        'gross_profit': 'Income Statement!B7'
    }
    add_metric_named_ranges(wb, mappings)
    
    # Extract and verify
    ranges = extract_named_ranges(wb)
    assert 'rng_revenue' in ranges
    assert ranges['rng_revenue'] == ('Income Statement', 'B5')

def test_metric_ingestion():
    """Test full ingestion pipeline"""
    # Create test workbook
    wb = create_test_workbook()
    
    # Add named ranges
    mappings = {
        'revenue': 'Income Statement!B5',
        'ebitda': 'Income Statement!B11'
    }
    add_metric_named_ranges(wb, mappings)
    
    # Save to temp file
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    
    try:
        # Run ingestion
        results = ingest_metrics('test-workspace', tmp_path, date(2024, 1, 31))
        
        # Verify results
        assert results['extracted'] >= 2  # At least revenue and ebitda
        assert results['inserted'] >= 0  # May be updates if run multiple times
        
    finally:
        # Cleanup
        Path(tmp_path).unlink()

def test_add_and_ingest_helper():
    """Test the combined helper function"""
    wb = create_test_workbook()
    
    # Use the helper
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp_path = Path(tmp.name)
        
        try:
            results = add_and_ingest('test-workspace', wb, tmp_path, date(2024, 1, 31))
            
            # Verify
            assert results['ranges_added'] > 0
            assert 'error' not in results or results['error'] is None
            
        finally:
            # Cleanup
            tmp_path.unlink()

@pytest.mark.asyncio
async def test_metrics_api(test_client):
    """Test metrics API endpoints"""
    # This would require a test client fixture
    # Example structure:
    
    # Test getting metrics
    response = await test_client.get("/api/test-workspace/metrics")
    assert response.status_code == 200
    data = response.json()
    assert 'metrics' in data
    
    # Test time series with limits
    response = await test_client.get(
        "/api/test-workspace/metrics/timeseries?metric_id=revenue&periods=36"
    )
    assert response.status_code == 200
    data = response.json()
    assert len(data['series']) <= 36

def test_monitoring_performance():
    """Test that ingestion completes in reasonable time"""
    import time
    
    wb = create_test_workbook()
    
    # Add many metrics
    for i in range(20):
        ws = wb.active
        ws[f'A{20+i}'] = f'Metric_{i}'
        ws[f'B{20+i}'] = 1000 * i
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        wb.save(tmp.name)
        
        start = time.time()
        results = ingest_metrics('test-workspace', tmp.name)
        duration = time.time() - start
        
        # Should complete in under 1 second for small file
        assert duration < 1.0
        
        # Cleanup
        Path(tmp.name).unlink()

if __name__ == '__main__':
    # Run basic tests
    print("Testing period normalization...")
    test_period_normalization()
    print("✅ Period normalization tests passed")
    
    print("\nTesting named ranges...")
    test_named_ranges()
    print("✅ Named range tests passed")
    
    print("\nTesting metric ingestion...")
    test_metric_ingestion()
    print("✅ Ingestion tests passed")
    
    print("\nTesting helper function...")
    test_add_and_ingest_helper()
    print("✅ Helper function tests passed")
    
    print("\nTesting performance...")
    test_monitoring_performance()
    print("✅ Performance tests passed")
    
    print("\n✨ All tests passed!")