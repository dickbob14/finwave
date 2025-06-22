#!/usr/bin/env python3

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import json
from pathlib import Path

def deep_analyze_sheet(sheet, sheet_name):
    """Perform deep analysis of a single sheet"""
    analysis = {
        "name": sheet_name,
        "structure": {
            "total_rows": sheet.max_row,
            "total_columns": sheet.max_column,
            "merged_cells": []
        },
        "data_regions": [],
        "potential_data_tables": [],
        "formulas": [],
        "pivot_references": []
    }
    
    # Check for merged cells (often used in headers)
    if hasattr(sheet, 'merged_cells'):
        for merged in sheet.merged_cells.ranges:
            analysis["structure"]["merged_cells"].append(str(merged))
    
    # Scan for data regions
    current_region = None
    empty_rows = 0
    
    for row_idx in range(1, min(sheet.max_row + 1, 200)):  # Limit scan to first 200 rows
        row_data = []
        non_empty = 0
        
        for col_idx in range(1, min(sheet.max_column + 1, 30)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                non_empty += 1
                row_data.append({
                    "col": get_column_letter(col_idx),
                    "value": str(cell.value)[:50],  # Truncate long values
                    "type": type(cell.value).__name__,
                    "has_formula": cell.data_type == 'f' if hasattr(cell, 'data_type') else False
                })
                
                # Track formulas
                if hasattr(cell, 'data_type') and cell.data_type == 'f':
                    analysis["formulas"].append({
                        "cell": f"{get_column_letter(col_idx)}{row_idx}",
                        "formula": cell.value if hasattr(cell, 'value') else ""
                    })
        
        # Detect data regions
        if non_empty > 2:  # Row has data
            if current_region is None:
                current_region = {
                    "start_row": row_idx,
                    "end_row": row_idx,
                    "sample_rows": [row_data]
                }
            else:
                current_region["end_row"] = row_idx
                if len(current_region["sample_rows"]) < 5:
                    current_region["sample_rows"].append(row_data)
            empty_rows = 0
        else:
            empty_rows += 1
            if empty_rows > 3 and current_region:
                analysis["data_regions"].append(current_region)
                current_region = None
    
    if current_region:
        analysis["data_regions"].append(current_region)
    
    # Identify potential data tables
    for region in analysis["data_regions"]:
        if region["end_row"] - region["start_row"] > 5:  # At least 6 rows
            # Check if first row looks like headers
            if region["sample_rows"]:
                first_row = region["sample_rows"][0]
                if all(item["type"] == "str" for item in first_row[:5] if item):
                    # This might be a data table
                    headers = [item["value"] for item in first_row]
                    analysis["potential_data_tables"].append({
                        "location": f"Rows {region['start_row']}-{region['end_row']}",
                        "headers": headers,
                        "row_count": region["end_row"] - region["start_row"]
                    })
    
    return analysis

def analyze_workbook_structure(file_path):
    """Analyze the entire workbook structure"""
    wb = openpyxl.load_workbook(file_path, data_only=False)  # Keep formulas
    
    workbook_analysis = {
        "file": Path(file_path).name,
        "sheets": {},
        "summary": {
            "total_sheets": len(wb.sheetnames),
            "has_data_sheets": False,
            "report_sheets": [],
            "data_source_sheets": [],
            "helper_sheets": []
        }
    }
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_analysis = deep_analyze_sheet(sheet, sheet_name)
        workbook_analysis["sheets"][sheet_name] = sheet_analysis
        
        # Categorize sheets
        if sheet_name.lower().startswith("data") or "data" in sheet_name.lower():
            workbook_analysis["summary"]["data_source_sheets"].append(sheet_name)
            workbook_analysis["summary"]["has_data_sheets"] = True
        elif sheet_analysis["potential_data_tables"]:
            workbook_analysis["summary"]["data_source_sheets"].append(sheet_name)
        elif len(sheet_analysis["formulas"]) > 20:
            workbook_analysis["summary"]["report_sheets"].append(sheet_name)
        elif sheet.max_row < 50 and len(sheet_analysis["formulas"]) > 0:
            workbook_analysis["summary"]["helper_sheets"].append(sheet_name)
    
    wb.close()
    return workbook_analysis

def create_data_schemas(workbook_analysis):
    """Create JSON schemas for identified data tables"""
    schemas = {}
    
    for sheet_name, sheet_data in workbook_analysis["sheets"].items():
        for table_idx, table in enumerate(sheet_data["potential_data_tables"]):
            schema_name = f"{sheet_name}_table_{table_idx}"
            schema = {
                "type": "object",
                "properties": {},
                "required": []
            }
            
            for header in table["headers"][:20]:  # Limit to 20 columns
                clean_header = header.strip().replace(" ", "_").replace("-", "_")
                schema["properties"][clean_header] = {
                    "type": "string",  # Default to string
                    "description": f"Column: {header}"
                }
                schema["required"].append(clean_header)
            
            schemas[schema_name] = {
                "schema": schema,
                "source": {
                    "sheet": sheet_name,
                    "location": table["location"],
                    "headers": table["headers"]
                }
            }
    
    return schemas

def main():
    files = [
        "/Users/alexandermillar/Downloads/Basic 3-Statement Model-2.xlsx",
        "/Users/alexandermillar/Downloads/Cube - KPI Dashboard-1.xlsx"
    ]
    
    for file_path in files:
        print(f"\n{'='*80}")
        print(f"DEEP ANALYSIS: {Path(file_path).name}")
        print('='*80)
        
        try:
            # Analyze structure
            analysis = analyze_workbook_structure(file_path)
            
            # Print summary
            print("\nWORKBOOK SUMMARY:")
            print(f"  Total Sheets: {analysis['summary']['total_sheets']}")
            print(f"  Report Sheets: {', '.join(analysis['summary']['report_sheets']) or 'None'}")
            print(f"  Data Source Sheets: {', '.join(analysis['summary']['data_source_sheets']) or 'None'}")
            print(f"  Helper Sheets: {', '.join(analysis['summary']['helper_sheets']) or 'None'}")
            
            # Print detailed sheet analysis
            print("\nDETAILED SHEET ANALYSIS:")
            for sheet_name, sheet_data in analysis["sheets"].items():
                print(f"\n  {sheet_name}:")
                print(f"    - Size: {sheet_data['structure']['total_rows']} rows x {sheet_data['structure']['total_columns']} columns")
                print(f"    - Merged Cells: {len(sheet_data['structure']['merged_cells'])}")
                print(f"    - Data Regions: {len(sheet_data['data_regions'])}")
                print(f"    - Formulas: {len(sheet_data['formulas'])}")
                
                if sheet_data["potential_data_tables"]:
                    print(f"    - Potential Data Tables:")
                    for table in sheet_data["potential_data_tables"]:
                        print(f"      * {table['location']} ({table['row_count']} rows)")
                        print(f"        Headers: {', '.join(table['headers'][:5])}")
                        if len(table['headers']) > 5:
                            print(f"        ... and {len(table['headers']) - 5} more columns")
            
            # Generate schemas
            schemas = create_data_schemas(analysis)
            if schemas:
                print("\nGENERATED DATA SCHEMAS:")
                for schema_name, schema_data in schemas.items():
                    print(f"\n  {schema_name}:")
                    print(f"    Source: {schema_data['source']['sheet']} - {schema_data['source']['location']}")
                    print(f"    Schema: {json.dumps(schema_data['schema'], indent=2)[:200]}...")
            
            # Save detailed analysis
            output_file = f"{Path(file_path).stem}_deep_analysis.json"
            with open(output_file, 'w') as f:
                json.dump({
                    "analysis": analysis,
                    "schemas": schemas
                }, f, indent=2)
            print(f"\nDetailed analysis saved to: {output_file}")
            
        except Exception as e:
            print(f"Error analyzing {file_path}: {str(e)}")
            import traceback
            traceback.print_exc()

if __name__ == "__main__":
    main()