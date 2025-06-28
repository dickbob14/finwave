"""
Excel utilities for adding named ranges to populated templates
"""

import logging
from pathlib import Path
from typing import Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

logger = logging.getLogger(__name__)

def add_metric_named_ranges(workbook: Workbook, metric_mappings: Dict[str, str]) -> None:
    """
    Add named ranges for metrics to Excel workbook
    
    Args:
        workbook: openpyxl Workbook object
        metric_mappings: Dict mapping metric_id to cell reference
                        e.g., {'revenue': 'Summary!B10', 'mrr': 'KPIs!C5'}
    """
    for metric_id, cell_ref in metric_mappings.items():
        # Create range name (prefix with rng_)
        range_name = f"rng_{metric_id}"
        
        # Check if sheet!cell format
        if '!' in cell_ref:
            sheet_name, cell_addr = cell_ref.split('!', 1)
            # Ensure sheet exists
            if sheet_name not in workbook.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' not found for metric {metric_id}")
                continue
        else:
            # Default to first sheet
            sheet_name = workbook.sheetnames[0]
            cell_addr = cell_ref
        
        # Create defined name
        defined_name = DefinedName(
            name=range_name,
            attr_text=f"'{sheet_name}'!${cell_addr.replace(':', ':$')}"
        )
        
        # Add to workbook
        workbook.defined_names.append(defined_name)
        logger.debug(f"Added named range: {range_name} -> {sheet_name}!{cell_addr}")

def add_period_named_range(workbook: Workbook, period_date: Any, cell_ref: str = "Summary!A1") -> None:
    """
    Add named range for period date
    """
    # Ensure the cell contains the period date
    if '!' in cell_ref:
        sheet_name, cell_addr = cell_ref.split('!', 1)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet[cell_addr] = period_date
    
    # Add named range
    defined_name = DefinedName(
        name="period_date",
        attr_text=f"'{sheet_name}'!${cell_addr}"
    )
    workbook.defined_names.append(defined_name)
    logger.debug(f"Added period_date named range: {cell_ref}")

# Standard metric mappings for different template types
STANDARD_MAPPINGS = {
    '3_statement_model': {
        # Income Statement metrics
        'revenue': 'Income Statement!B6',
        'cogs': 'Income Statement!B8', 
        'gross_profit': 'Income Statement!B10',
        'opex': 'Income Statement!B15',
        'ebitda': 'Income Statement!B17',
        'net_income': 'Income Statement!B25',
        
        # Balance Sheet metrics
        'cash': 'Balance Sheet!B6',
        'total_assets': 'Balance Sheet!B20',
        'total_liabilities': 'Balance Sheet!B35',
        'total_equity': 'Balance Sheet!B45',
        
        # Cash Flow metrics
        'operating_cash_flow': 'Cash Flow!B15',
        'free_cash_flow': 'Cash Flow!B25'
    },
    
    'kpi_dashboard': {
        # Financial KPIs
        'revenue': 'KPI Dashboard!C5',
        'gross_margin': 'KPI Dashboard!C7',
        'ebitda_margin': 'KPI Dashboard!C8',
        'burn_rate': 'KPI Dashboard!C10',
        
        # SaaS Metrics
        'mrr': 'KPI Dashboard!C13',
        'arr': 'KPI Dashboard!C14',
        'new_customers': 'KPI Dashboard!C15',
        'churn_rate': 'KPI Dashboard!C16',
        'ltv': 'KPI Dashboard!C18',
        'cac': 'KPI Dashboard!C17',
        
        # Operational
        'headcount': 'KPI Dashboard!C20',
        'revenue_per_employee': 'KPI Dashboard!C21',
        'runway_months': 'KPI Dashboard!C25'
    },
    
    'board_pack': {
        # Executive Summary metrics
        'revenue': 'Executive Summary!B10',
        'ebitda': 'Executive Summary!B12',
        'cash': 'Executive Summary!B14',
        'burn_rate': 'Executive Summary!B16',
        'runway_months': 'Executive Summary!B18',
        
        # P&L metrics
        'gross_profit': 'P&L!B10',
        'opex': 'P&L!B20',
        'net_income': 'P&L!B30'
    }
}

def get_template_type(workbook: Workbook) -> Optional[str]:
    """
    Detect template type based on sheet names
    """
    sheet_names = set(workbook.sheetnames)
    
    # Check for specific sheets
    if 'KPI Dashboard' in sheet_names:
        return 'kpi_dashboard'
    elif 'Income Statement' in sheet_names and 'Balance Sheet' in sheet_names:
        return '3_statement_model'
    elif 'Executive Summary' in sheet_names:
        return 'board_pack'
    
    return None

def auto_add_metric_ranges(workbook: Workbook) -> int:
    """
    Automatically add standard metric named ranges based on template type
    Returns number of ranges added
    """
    template_type = get_template_type(workbook)
    
    if not template_type:
        logger.warning("Could not detect template type")
        return 0
    
    mappings = STANDARD_MAPPINGS.get(template_type, {})
    if not mappings:
        logger.warning(f"No standard mappings for template type: {template_type}")
        return 0
    
    # Filter mappings to only include sheets that exist
    valid_mappings = {}
    for metric_id, cell_ref in mappings.items():
        if '!' in cell_ref:
            sheet_name = cell_ref.split('!')[0]
            if sheet_name in workbook.sheetnames:
                valid_mappings[metric_id] = cell_ref
    
    # Add the ranges
    add_metric_named_ranges(workbook, valid_mappings)
    
    logger.info(f"Added {len(valid_mappings)} metric named ranges for {template_type}")
    return len(valid_mappings)

def add_and_ingest(workspace_id: str, workbook: Workbook, filename: Path, 
                  period_date: Optional[Any] = None) -> Dict[str, int]:
    """
    Helper that adds named ranges AND ingests metrics in one step
    This ensures populators never forget to update the metric store
    
    Returns dict with ingestion results
    """
    from metrics.ingest import ingest_metrics
    
    # Add named ranges
    ranges_added = auto_add_metric_ranges(workbook)
    
    # Add period date range if provided
    if period_date:
        add_period_named_range(workbook, period_date)
        ranges_added += 1
    
    # Save the workbook
    workbook.save(filename)
    logger.info(f"Saved workbook with {ranges_added} named ranges: {filename}")
    
    # Ingest metrics
    try:
        results = ingest_metrics(workspace_id, str(filename), period_date)
        results['ranges_added'] = ranges_added
        return results
    except Exception as e:
        logger.error(f"Failed to ingest metrics: {e}")
        return {
            'extracted': 0,
            'inserted': 0,
            'updated': 0,
            'ranges_added': ranges_added,
            'error': str(e)
        }