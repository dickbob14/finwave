"""
Export API routes for financial reports
Handles Excel, Google Sheets, and other export formats
"""
import logging
import tempfile
from datetime import datetime, timedelta
from typing import Optional
import os

from fastapi import APIRouter, HTTPException, Query, Response
from fastapi.responses import StreamingResponse
import io

from templates.excel_templates import generate_financial_excel, export_to_google_sheets
from database import get_db_session

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/export", tags=["export"])

@router.get("/excel")
async def export_excel(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    filename: Optional[str] = Query(None, description="Custom filename for the export")
):
    """
    Export financial data to Excel format
    
    Returns a downloadable Excel file with comprehensive financial reports
    """
    try:
        # Validate dates
        datetime.fromisoformat(start_date)
        datetime.fromisoformat(end_date)
        
        # Generate filename if not provided
        if not filename:
            filename = f"financial_report_{start_date}_to_{end_date}.xlsx"
        elif not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        # Generate Excel file
        excel_file = generate_financial_excel(start_date, end_date)
        excel_file.seek(0)
        
        # Return as streaming response
        return StreamingResponse(
            io.BytesIO(excel_file.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Invalid date format: {str(e)}")
    except Exception as e:
        logger.error(f"Excel export failed: {e}")
        raise HTTPException(status_code=500, detail="Excel export failed")

@router.get("/google-sheets")
async def export_google_sheets(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    sheet_title: Optional[str] = Query(None, description="Title for the Google Sheet")
):
    """
    Export financial data to Google Sheets
    
    Returns a URL to the created Google Sheet
    """
    try:
        # Check if Google Sheets credentials are configured
        credentials_path = os.getenv("GOOGLE_SHEETS_CREDENTIALS")
        if not credentials_path or not os.path.exists(credentials_path):
            raise HTTPException(
                status_code=400, 
                detail="Google Sheets credentials not configured. Set GOOGLE_SHEETS_CREDENTIALS environment variable."
            )
        
        # Validate dates
        datetime.fromisoformat(start_date)
        datetime.fromisoformat(end_date)
        
        # Generate sheet title if not provided
        if not sheet_title:
            sheet_title = f"Financial Report {start_date} to {end_date}"
        
        # Generate Excel file first
        excel_file = generate_financial_excel(start_date, end_date)
        
        # Export to Google Sheets
        sheet_url = export_to_google_sheets(excel_file, sheet_title, credentials_path)
        
        return {
            "status": "success",
            "sheet_url": sheet_url,
            "sheet_title": sheet_title,
            "period": {"start_date": start_date, "end_date": end_date},
            "created_at": datetime.now().isoformat()
        }
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Invalid date format: {str(e)}")
    except Exception as e:
        logger.error(f"Google Sheets export failed: {e}")
        raise HTTPException(status_code=500, detail=f"Google Sheets export failed: {str(e)}")

@router.get("/template/{template_type}")
async def download_template(
    template_type: str
):
    """
    Download empty Excel template for specific financial statement
    
    Returns a downloadable Excel template file
    """
    try:
        # Map template types to filenames
        template_map = {
            "trial_balance": "trial_balance_template.xlsx",
            "pl": "profit_loss_template.xlsx", 
            "balance_sheet": "balance_sheet_template.xlsx",
            "cash_flow": "cash_flow_template.xlsx",
            "variance_analysis": "variance_analysis_template.xlsx"
        }
        
        if template_type not in template_map:
            raise HTTPException(
                status_code=400, 
                detail=f"Unknown template type: {template_type}. Available: {list(template_map.keys())}"
            )
        
        # Generate a basic template (simplified version)
        # In a full implementation, you'd have pre-built template files
        from templates.excel_templates import ExcelTemplateGenerator
        
        # Create template with mock data for structure
        mock_start = "2024-01-01"
        mock_end = "2024-01-31"
        
        generator = ExcelTemplateGenerator()
        excel_file = generator.generate_financial_summary(mock_start, mock_end)
        excel_file.seek(0)
        
        filename = template_map[template_type]
        
        return StreamingResponse(
            io.BytesIO(excel_file.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Template download failed: {e}")
        raise HTTPException(status_code=500, detail="Template download failed")

@router.post("/custom")
async def export_custom_data(
    data: dict,
    format: str = Query("excel", description="Export format (excel, csv, json)"),
    filename: Optional[str] = Query(None, description="Custom filename")
):
    """
    Export custom data in specified format
    
    Accepts arbitrary financial data and exports in requested format
    """
    try:
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"custom_export_{timestamp}"
        
        if format.lower() == "excel":
            # Create simple Excel export from data
            import openpyxl
            
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Custom Data"
            
            # Add headers
            if isinstance(data, dict) and data:
                first_key = list(data.keys())[0]
                if isinstance(data[first_key], list) and data[first_key]:
                    # Assume list of dictionaries
                    headers = list(data[first_key][0].keys())
                    for col, header in enumerate(headers, 1):
                        worksheet.cell(row=1, column=col, value=header)
                    
                    # Add data rows
                    for section_name, rows in data.items():
                        if isinstance(rows, list):
                            for row_idx, row_data in enumerate(rows, 2):
                                for col, header in enumerate(headers, 1):
                                    value = row_data.get(header, '')
                                    worksheet.cell(row=row_idx, column=col, value=value)
            
            # Save to BytesIO
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
            
        elif format.lower() == "csv":
            import csv
            import io
            
            output = io.StringIO()
            
            if isinstance(data, dict) and data:
                first_key = list(data.keys())[0]
                if isinstance(data[first_key], list) and data[first_key]:
                    headers = list(data[first_key][0].keys())
                    writer = csv.DictWriter(output, fieldnames=headers)
                    writer.writeheader()
                    
                    for section_name, rows in data.items():
                        if isinstance(rows, list):
                            writer.writerows(rows)
            
            if not filename.endswith('.csv'):
                filename += '.csv'
            
            return Response(
                content=output.getvalue(),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
            
        elif format.lower() == "json":
            import json
            
            if not filename.endswith('.json'):
                filename += '.json'
            
            return Response(
                content=json.dumps(data, indent=2, default=str),
                media_type="application/json",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
            
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported format: {format}")
            
    except Exception as e:
        logger.error(f"Custom export failed: {e}")
        raise HTTPException(status_code=500, detail="Custom export failed")

@router.get("/formats")
async def get_supported_formats():
    """
    Get list of supported export formats and their descriptions
    """
    return {
        "formats": [
            {
                "name": "excel",
                "description": "Microsoft Excel format with multiple worksheets and formatting",
                "file_extension": ".xlsx",
                "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            },
            {
                "name": "google_sheets",
                "description": "Export directly to Google Sheets (requires credentials)",
                "file_extension": None,
                "mime_type": None
            },
            {
                "name": "csv",
                "description": "Comma-separated values format",
                "file_extension": ".csv",
                "mime_type": "text/csv"
            },
            {
                "name": "json",
                "description": "JavaScript Object Notation format",
                "file_extension": ".json",
                "mime_type": "application/json"
            }
        ],
        "templates": [
            "trial_balance",
            "pl",
            "balance_sheet", 
            "cash_flow",
            "variance_analysis"
        ]
    }

@router.get("/status")
async def get_export_status():
    """
    Get export system status and configuration
    """
    # Check Google Sheets configuration
    google_sheets_configured = bool(os.getenv("GOOGLE_SHEETS_CREDENTIALS"))
    
    # Check database connectivity
    try:
        with get_db_session() as db:
            # Simple query to test connectivity
            db.execute("SELECT 1")
            db_connected = True
    except Exception:
        db_connected = False
    
    return {
        "status": "healthy" if db_connected else "degraded",
        "database_connected": db_connected,
        "google_sheets_configured": google_sheets_configured,
        "supported_formats": ["excel", "csv", "json"] + (["google_sheets"] if google_sheets_configured else []),
        "checked_at": datetime.now().isoformat()
    }