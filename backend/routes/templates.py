"""
Template API routes for financial report generation
Implements the finance-as-code playbook patterns
"""

import asyncio
import logging
import os
from datetime import datetime
from typing import Optional, List, Dict, Any

from fastapi import APIRouter, HTTPException, Query, BackgroundTasks, Response
from fastapi.responses import FileResponse, StreamingResponse
import boto3
from botocore.exceptions import ClientError
import pandas as pd
import openpyxl

# Import template manager
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from templates.template_manager import TemplateManager, TEMPLATE_REGISTRY

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/templates", tags=["templates"])

# Initialize template manager with company from env or default
COMPANY_SLUG = os.getenv("COMPANY_SLUG", "demo_corp")
template_manager = TemplateManager(COMPANY_SLUG)

# CRM configuration
CRM_TYPE = os.getenv("CRM_TYPE", "salesforce")  # Default to Salesforce

@router.get("/")
async def list_templates():
    """
    List all available financial templates with their metadata
    
    Returns template configurations including refresh frequency and delivery methods
    """
    templates = template_manager.list_templates()
    
    return {
        "templates": templates,
        "company": COMPANY_SLUG,
        "storage": {
            "s3_bucket": template_manager.s3_bucket,
            "base_path": f"s3://{template_manager.s3_bucket}/{COMPANY_SLUG}/"
        }
    }

@router.get("/{template_name}")
async def get_template_info(template_name: str):
    """
    Get detailed information about a specific template
    
    Includes schema, data sources, and sample command
    """
    if template_name not in TEMPLATE_REGISTRY:
        raise HTTPException(status_code=404, detail=f"Template not found: {template_name}")
    
    config = TEMPLATE_REGISTRY[template_name]
    
    # Get recent populated files
    history = template_manager.get_populated_files(template_name, limit=5)
    
    return {
        "name": config.name,
        "title": config.title,
        "description": config.description,
        "use_case": config.use_case,
        "version": config.version,
        "refresh_frequency": config.refresh_frequency.value,
        "delivery_methods": [m.value for m in config.delivery_methods],
        "schema": config.schema,
        "recent_files": history,
        "sample_command": f"make populate-{config.name.replace('_', '-')} SINCE=2024-01-01 UNTIL=2024-12-31"
    }

@router.post("/{template_name}/populate")
async def populate_template(
    template_name: str,
    background_tasks: BackgroundTasks,
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: Optional[str] = Query(None, description="End date in YYYY-MM-DD format"),
    sheet_id: Optional[str] = Query(None, description="Google Sheets ID for upload"),
    async_mode: bool = Query(False, description="Run population in background")
):
    """
    Populate a template with financial data
    
    Can run synchronously or asynchronously for large datasets
    """
    if template_name not in TEMPLATE_REGISTRY:
        raise HTTPException(status_code=404, detail=f"Template not found: {template_name}")
    
    # Validate dates
    try:
        datetime.fromisoformat(start_date)
        if end_date:
            datetime.fromisoformat(end_date)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Invalid date format: {str(e)}")
    
    # Run population
    if async_mode:
        # Queue for background processing
        task_id = datetime.now().strftime('%Y%m%d_%H%M%S')
        background_tasks.add_task(
            _populate_template_async,
            template_name, start_date, end_date, sheet_id, task_id
        )
        
        return {
            "status": "queued",
            "task_id": task_id,
            "message": f"Template population queued. Check /templates/{template_name}/status/{task_id}"
        }
    else:
        # Run synchronously
        try:
            result = template_manager.populate_template(
                template_name,
                start_date,
                end_date,
                sheet_id=sheet_id
            )
            
            return {
                "status": "success",
                "populated_file": os.path.basename(result['populated_file']),
                "download_url": f"/templates/{template_name}/download/{os.path.basename(result['populated_file'])}",
                "delivery_results": result['delivery_results'],
                "metadata": result['run_log']
            }
            
        except Exception as e:
            logger.error(f"Template population failed: {e}")
            raise HTTPException(status_code=500, detail=f"Population failed: {str(e)}")

@router.get("/{template_name}/download/{filename}")
async def download_populated_file(template_name: str, filename: str):
    """
    Download a populated template file
    
    Returns the Excel file for download
    """
    # Security check - ensure filename matches expected pattern
    if not filename.endswith('.xlsx') or '..' in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    
    # Check local file first
    local_path = template_manager.populated_path / filename
    if local_path.exists():
        return FileResponse(
            path=str(local_path),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename
        )
    
    # Try S3
    if template_manager.s3_bucket:
        try:
            s3_key = template_manager.get_s3_key('populated', filename)
            s3_client = template_manager._get_s3_client()
            
            # Get object
            response = s3_client.get_object(Bucket=template_manager.s3_bucket, Key=s3_key)
            
            return StreamingResponse(
                response['Body'],
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
            
        except ClientError:
            pass
    
    raise HTTPException(status_code=404, detail="File not found")

@router.get("/{template_name}/history")
async def get_template_history(
    template_name: str,
    limit: int = Query(10, description="Number of recent files to return")
):
    """
    Get history of populated files for a template
    
    Returns recent generated files with metadata
    """
    if template_name not in TEMPLATE_REGISTRY:
        raise HTTPException(status_code=404, detail=f"Template not found: {template_name}")
    
    files = template_manager.get_populated_files(template_name, limit)
    
    return {
        "template": template_name,
        "files": files,
        "total": len(files)
    }

@router.get("/{template_name}/snapshot")
async def get_template_snapshot(template_name: str):
    """
    Get the template file for download
    
    Returns the base template Excel file
    """
    # Map template names to actual file names
    template_files = {
        "3statement": "3statement_model.xlsx",
        "budget": "3statement_model.xlsx",
        "kpi": "kpi_dashboard.xlsx",
        "cash_flow": "3statement_model.xlsx",
        "revenue": "3statement_model.xlsx"
    }
    
    filename = template_files.get(template_name)
    if not filename:
        raise HTTPException(status_code=404, detail=f"Template not found: {template_name}")
    
    # Check multiple possible locations
    possible_paths = [
        f"templates/{filename}",
        f"assets/templates/{filename}",
        f"backend/assets/templates/{filename}"
    ]
    
    for path in possible_paths:
        file_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), path)
        if os.path.exists(file_path):
            return FileResponse(
                path=file_path,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=f"{template_name}_template.xlsx"
            )
    
    raise HTTPException(status_code=404, detail="Template file not found")

@router.post("/{template_name}/schedule")
async def schedule_template_refresh(
    template_name: str,
    cron_expression: str = Query(..., description="Cron expression for scheduling"),
    start_date: str = Query(..., description="Start date expression (e.g., 'MONTH_START')"),
    end_date: str = Query(..., description="End date expression (e.g., 'TODAY')"),
    delivery_config: Dict[str, Any] = None
):
    """
    Schedule automatic template refresh
    
    Sets up recurring population based on cron expression
    """
    if template_name not in TEMPLATE_REGISTRY:
        raise HTTPException(status_code=404, detail=f"Template not found: {template_name}")
    
    # In production, this would integrate with your scheduler (Airflow, etc.)
    # For now, return the configuration that would be used
    
    config = TEMPLATE_REGISTRY[template_name]
    
    schedule_config = {
        "template": template_name,
        "cron": cron_expression,
        "date_range": {
            "start": start_date,
            "end": end_date
        },
        "delivery": delivery_config or {
            "methods": [m.value for m in config.delivery_methods],
            "recipients": []
        },
        "created_at": datetime.now().isoformat()
    }
    
    # TODO: Save to scheduler database
    
    return {
        "status": "scheduled",
        "config": schedule_config,
        "next_run": "TBD based on cron expression"
    }

@router.post("/3statement/refresh")
async def refresh_3statement(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: Optional[str] = Query(None, description="End date in YYYY-MM-DD format"),
    include_prior_year: bool = Query(False, description="Include prior year data"),
    sheet_id: Optional[str] = Query(None, description="Google Sheets ID for upload")
):
    """
    Refresh 3-Statement Model with live QuickBooks data
    
    Returns the populated Excel file
    """
    try:
        # Import the v2 populator
        from templates.populate_3statement_v2 import ThreeStatementPopulator
        from pathlib import Path
        
        # Default end date to today
        if not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
        
        # Get template path
        template_path = Path(__file__).parent.parent / 'templates' / 'files' / 'Basic 3-Statement Model-2.xlsx'
        if not template_path.exists():
            # Fall back to template directory
            template_path = Path(__file__).parent.parent / 'assets' / 'templates' / 'Basic 3-Statement Model-2.xlsx'
        
        if not template_path.exists():
            raise HTTPException(status_code=404, detail="3-Statement template not found")
        
        # Initialize populator
        populator = ThreeStatementPopulator(str(template_path))
        populator.load_template()
        
        # Fetch QuickBooks data
        data = populator.fetch_quickbooks_data(start_date, end_date)
        
        # Include prior year if requested
        if include_prior_year:
            py_data = populator.fetch_prior_year_data(start_date, end_date)
            data['pl_py'] = py_data.get('pl', pd.DataFrame())
            data['bs_py'] = py_data.get('bs', pd.DataFrame())
        
        # Populate sheets
        if not data['pl'].empty:
            populator.populate_income_statement(data['pl'])
        if not data['bs'].empty:
            populator.populate_balance_sheet(data['bs'])
        
        # Save file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"3statement_live_{timestamp}.xlsx"
        output_path = template_manager.populated_path / output_filename
        template_manager.populated_path.mkdir(exist_ok=True)
        
        populated_path = populator.save_populated_file(str(output_path))
        
        # Upload to Google Sheets if requested
        if sheet_id:
            try:
                sheet_url = populator.upload_to_google_sheets(sheet_id)
                logger.info(f"Uploaded to Google Sheets: {sheet_url}")
            except Exception as e:
                logger.warning(f"Google Sheets upload failed: {e}")
        
        # Return file
        return FileResponse(
            path=populated_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=output_filename
        )
        
    except Exception as e:
        logger.error(f"3-Statement refresh failed: {e}")
        raise HTTPException(status_code=500, detail=f"Refresh failed: {str(e)}")

@router.get("/{template_name}/snapshot")
async def get_template_snapshot(
    template_name: str,
    format: str = Query("json", description="Output format (json, csv)"),
    metrics: Optional[str] = Query(None, description="Comma-separated list of metrics to include")
):
    """
    Get a lightweight snapshot of template data
    
    Perfect for Slack bots, dashboards, or quick KPI checks
    """
    if template_name not in TEMPLATE_REGISTRY:
        raise HTTPException(status_code=404, detail=f"Template not found: {template_name}")
    
    # Get most recent populated file
    files = template_manager.get_populated_files(template_name, limit=1)
    if not files:
        raise HTTPException(status_code=404, detail="No populated files found")
    
    latest_file = files[0]['filename']
    file_path = template_manager.populated_path / latest_file
    
    if not file_path.exists():
        # Try downloading from S3
        s3_key = files[0]['key']
        if template_manager.download_from_s3(s3_key, file_path):
            logger.info(f"Downloaded from S3: {s3_key}")
        else:
            raise HTTPException(status_code=404, detail="File not found")
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # Extract key metrics based on template type
        if template_name == "kpi_dashboard":
            # Extract from specific ranges
            kpi_data = {}
            
            # Try to find KPI sheet
            kpi_sheet = None
            for sheet_name in ['DASH_KPI', 'KPIs', 'Dashboard']:
                if sheet_name in wb.sheetnames:
                    kpi_sheet = wb[sheet_name]
                    break
            
            if kpi_sheet:
                # Extract specific KPIs (customize based on template)
                kpi_ranges = {
                    'revenue_mtd': 'B5',
                    'revenue_ytd': 'B6',
                    'expenses_mtd': 'B8',
                    'expenses_ytd': 'B9',
                    'ebitda_margin': 'B11',
                    'cash_balance': 'B13',
                    'headcount': 'B15'
                }
                
                for metric, cell_ref in kpi_ranges.items():
                    if not metrics or metric in metrics.split(','):
                        try:
                            value = kpi_sheet[cell_ref].value
                            kpi_data[metric] = value if value is not None else 0
                        except:
                            kpi_data[metric] = None
            
            # Add metadata
            kpi_data['_metadata'] = {
                'template': template_name,
                'generated_at': files[0]['last_modified'],
                'file': latest_file
            }
            
            if format == "csv":
                # Convert to CSV
                import io
                import csv
                
                output = io.StringIO()
                writer = csv.DictWriter(output, fieldnames=kpi_data.keys())
                writer.writeheader()
                writer.writerow(kpi_data)
                
                return Response(
                    content=output.getvalue(),
                    media_type="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=kpi_snapshot.csv"}
                )
            else:
                return kpi_data
                
        elif template_name == "3_statement_model":
            # Extract P&L summary
            pl_data = {}
            
            if 'Income Statement' in wb.sheetnames:
                pl = wb['Income Statement']
                
                # Get latest period column (assume it's the last populated column)
                last_col = 2  # Start from column B
                for col in range(2, pl.max_column + 1):
                    if pl.cell(row=3, column=col).value:  # Period header
                        last_col = col
                
                # Extract key metrics from last period
                metric_rows = {
                    'revenue': 5,
                    'cogs': 6,
                    'gross_profit': 7,
                    'opex': 9,
                    'ebitda': 11,
                    'net_income': 14
                }
                
                pl_data['period'] = pl.cell(row=3, column=last_col).value
                
                for metric, row in metric_rows.items():
                    if not metrics or metric in metrics.split(','):
                        try:
                            value = pl.cell(row=row, column=last_col).value
                            pl_data[metric] = float(value) if value else 0
                        except:
                            pl_data[metric] = None
            
            pl_data['_metadata'] = {
                'template': template_name,
                'generated_at': files[0]['last_modified'],
                'file': latest_file
            }
            
            return pl_data
            
        else:
            # Generic data extraction
            data = {
                'template': template_name,
                'sheets': wb.sheetnames,
                'generated_at': files[0]['last_modified'],
                'file': latest_file
            }
            return data
            
    except Exception as e:
        logger.error(f"Snapshot extraction failed: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to extract snapshot: {str(e)}")

@router.get("/health")
async def template_system_health():
    """
    Check health of template system
    
    Verifies S3 access, template availability, etc.
    """
    health_status = {
        "status": "healthy",
        "checks": {}
    }
    
    # Check template files exist
    template_count = 0
    for name, config in TEMPLATE_REGISTRY.items():
        template_path = template_manager.templates_path / config.template_file
        if template_path.exists() or template_manager.s3_bucket:
            template_count += 1
    
    health_status["checks"]["templates"] = {
        "available": template_count,
        "total": len(TEMPLATE_REGISTRY)
    }
    
    # Check S3 access
    if template_manager.s3_bucket:
        try:
            s3_client = template_manager._get_s3_client()
            s3_client.head_bucket(Bucket=template_manager.s3_bucket)
            health_status["checks"]["s3"] = "connected"
        except:
            health_status["checks"]["s3"] = "error"
            health_status["status"] = "degraded"
    
    # Check Google Sheets
    health_status["checks"]["google_sheets"] = "configured" if os.getenv("GOOGLE_SHEETS_JSON") else "not configured"
    
    return health_status

# Async helper function
async def _populate_template_async(template_name: str, start_date: str, 
                                  end_date: Optional[str], sheet_id: Optional[str],
                                  task_id: str):
    """Background task for template population"""
    try:
        result = template_manager.populate_template(
            template_name,
            start_date, 
            end_date,
            sheet_id=sheet_id
        )
        # In production, save result to cache/database with task_id
        logger.info(f"Async population completed for task {task_id}")
    except Exception as e:
        logger.error(f"Async population failed for task {task_id}: {e}")

# Include this router in main app
__all__ = ['router']